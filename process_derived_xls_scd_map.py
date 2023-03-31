# Standard libs
import os, argparse
from pathlib import Path

# Other necessary libs
import openpyxl, yaml               # pip install openpyxl (pyyaml)
from operator import itemgetter


def process_args():
    parser = argparse.ArgumentParser(description='Command Line Arguments')
    parser.add_argument('--rootdir', default=Path.cwd(),
                        help='The Working Directory for maps and Models (defaults to Current Directory)')
    parser.add_argument('--schema', default='CHANGEME',
                        help='if not default, becomes DEV_EDW_xxxxx with xxxxx passed as the schema name')
    # boolean
    parser.add_argument( '--yml', action='store_true', help='If flagged as tests, will only generate yml files')
    parser.add_argument( '--silent', action='store_true', help='If flagged will only print file names')
    parser.add_argument( '--debug', action='store_true', help='If flagged will add verbosity to output')

    args = parser.parse_args()

    if not args.schema == 'CHANGEME':
        devedw = f'DEV_EDW_{args.schema}'
        args.schema = devedw

    return args


def setup(args):
    #use global args to store the rootdir and subdirs
    dir_list = {'mapdir':'mappings', 'modeldir':'models', 'snapdir':'snapshots', 'archivedir':'models_archive'}

    for dir in dir_list.items():
        setattr(args, dir[0], args.rootdir / dir[1])
        if not getattr(args, dir[0]).exists(): getattr(args, dir[0]).mkdir()

    print( '--Setting up directories--')
    for dir in dir_list:
        print( f'{dir} directory: {getattr(args, dir)}' )
    print( f'{"-"*30}' )

    print( 'Mapping docs (XLS) in:', args.mapdir )
    print( 'Output (SQL, YML) going to:', args.modeldir )

    # Show existing XLS files in Mapdir
    for afile in args.mapdir.glob('*.xlsx'):
        print(f'\tMapping found: {afile}')


def sheet2dict(sheet):
    '''table:  {'Source Schema': 'STAGING', 'Source Table': 'DSV_CUSTOMER', 'Alias': 'C', 'Filter Conditions': None, 'Manual Filter Conditions': 'Exclude all the CUSTOMER_NUMBER  if Exists in DIM_PROVIDER & DIM_NETWORK\nPROV & NET Col Names : PROVIDER_PEACH_NUMBER,CORESUITE_CUSTOMER_NUMBER \n', 'Parent Join Number': None, 'Parent Table Join': None, 'Child Table Join': None, 'Join Type': None}
       column: {'Source Schema': 'STAGING', 'Source Table': 'C', 'Source Column': '(derived)', 'Datatype': None, 'Automated Logic': None, 'Manual Logic': 'hash:  CUSTOMER_NUMBER,RECORD_EFFECTIVE_DATE', 'Order#': 1, 'Staging Layer Column Name': 'CUSTOMER_HKEY', 'Staging Layer Datatype': 'CHAR(32)', 'CPI': None, 'Unique': 'unique', 'Not Null': 'not_null', 'Test Expression': None, 'Model Equality': None, ...}'''
    print(f'\t<<<< Processing {sheet}')

    fields = None
    for rownum, row in enumerate( sheet.iter_rows( values_only = True) ):
        if rownum == 0:
            fields = row
            continue
        row = dict(zip( fields, row ))

        yield row


def sheet2list(sheet):
    for row in sheet.iter_rows( values_only = True ):
        if not row:
            continue
        yield row

###### START OF SCD SPECIFIC ADDED CODE
def build_scd2( schema, table, cols, keycol ):

    #keycol = 'UNIQUE_DIM_KEY'
    schema = 'EDW_STAGING_SNAPSHOT'
    all_cols = cols + [ keycol ]
    
    cols_str = ', '.join( all_cols )
    src_table = table.replace( 'DIM_' , 'DSV_' )
    snap=f'''
    
    {{%% snapshot {table}_SNAPSHOT_STEP1 %%}}

    {{{{
        config(
        target_schema={schema!r},
        unique_key={keycol!r},
        strategy='check',
        check_cols={cols},
        )
    }}}}

    select {cols_str} 
        from  {{{{ref( {src_table!r}) }}}}

    {{%% endsnapshot %%}}
    '''

    return snap


def build_scd6 ( schema, table, keycol, scd1_cols = [], scd2_cols = [], scd3_cols = [], scd0_cols = [] ):
    #print(scd0_cols)
    keycol = keycol.upper()
    #keycol='UNIQUE_DIM_KEY'
    if not keycol: keycol = 'UNIQUE_ID_KEY'

    exclude_col = ['UNIQUE_ID_KEY', 'CURRENT_INDICATOR', 'EFFECTIVE_DATE', 'END_DATE', 'LOAD_DATETIME', 
    'UPDATE_DATETIME', 'EFFECTIVE_TIMESTAMP', 'END_TIMESTAMP',
    'LOAD_BATCH_ID', 'UPDATE_BATCH_ID', 'PRIMARY_SOURCE_SYSTEM']

    start_sql = f'''\n\n WITH  SCD AS ( \n\tSELECT  \n\t{keycol}'''

    sql2 = []
    for col in scd2_cols:
        if col in exclude_col: continue
        sql = f'''\n\t, {col}'''
        sql2.append( sql )

    sql3 = []
    for col in scd3_cols:
        if col in exclude_col: continue
        sql = f'''\n     lag({col}) over  
                  (partition by {keycol} 
                        order by dbt_updated_at
                   ) as {col}'''
                   #) as PREV_{col}'''
        sql3.append(sql)

    sql1 = []
    for idx, col in enumerate( scd1_cols ) :
        if col in exclude_col: continue
        if idx == 0: 
            sql = f'\n      last_value({col}) over '
        else:
            sql = f'\n    , last_value({col}) over '
        sql += f'''(partition by {keycol} 
        order by dbt_updated_at 
            ROWS BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING
            ) as {col}'''
        sql1.append(sql)
    
    sql0 = []
    for idx, col in enumerate( scd0_cols ):
        if col in exclude_col: continue
        if idx == 0:
            sql=f'\n     first_value({col}) over '
        else:
            sql=f'\n   , first_value({col}) over '
        sql += f'''(partition by {keycol} 
        order by dbt_updated_at 
            ROWS BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING
            ) as {col}'''
        sql0.append( sql )

    end_sql = f'''\n\t, DBT_VALID_FROM AS EFFECTIVE_TIMESTAMP \n\t, DBT_VALID_TO   AS END_TIMESTAMP \n\t, CASE WHEN CAST(DBT_VALID_FROM AS DATE) = '1901-01-01' then CAST(DBT_VALID_FROM AS DATE)
      WHEN CAST(DBT_VALID_FROM AS DATE) <> '1901-01-01' THEN dateadd(day,1,CAST(DBT_VALID_FROM AS DATE))
      else CAST(DBT_VALID_FROM AS DATE) end as EFFECTIVE_DATE\n\t, CAST(DBT_VALID_TO AS DATE) as END_DATE \n\t, CASE WHEN DBT_VALID_TO IS NULL THEN 'Y' ELSE 'N' END AS CURRENT_INDICATOR'''
    #CAST(DBT_VALID_FROM AS DATE) as EFFECTIVE_DATE\n,
    sql1[ 0 ] = start_sql + sql1[ 0 ]
    scd_all = sql1 + sql2 + sql3 + sql0
    #+ end_sql

    # Previous joined with comma
    scd_str = ' '.join( scd_all )
    scd_str = scd_str.strip(',')

    scd_str = scd_str + end_sql + '\n'+"     \n\tFROM {{ ref('" + table + "_SCDALL_STEP2') }})\n\n"
    scd_str += 'select * from SCD\n'

    return scd_str


def make_scd( scd_dict ):
    '''
    scd={
        'table':'STG_POLICYTEST',
        'schema':'X10057301',
        'scd1':['POLICY_ID','CLAIM_ID'],
        'scd2':['PTYPE'],
        'scd3':['OWNER_NAME','ADDR'],
        'keycol':'POLICY_ID'
    }
    '''

    keycol = 'KEY'

    #print(scd_dict)
    #cols=scd_dict['scd2']+scd_dict['scd3']+scd_dict['scd0']
    cols = scd_dict[ 'scd2' ] + scd_dict[ 'scd3' ] + scd_dict[ 'scd0' ]

    scd2_sql = build_scd2( scd_dict[ 'schema' ], scd_dict[ 'table' ], cols, scd_dict[ keycol ])

    scd6_sql = build_scd6( scd_dict[ 'schema' ], scd_dict[ 'table' ], scd_dict[ keycol ],
        scd1_cols = scd_dict[ 'scd1' ], scd2_cols = scd_dict[ 'scd2' ], scd3_cols = scd_dict[ 'scd3' ], scd0_cols = scd_dict[ 'scd0' ])

    return scd2_sql, scd6_sql


def write_scd_step1( table, snapdir, scd2_sql, modeldir, scd6_sql):
    snapfile = snapdir/(table+'_SNAPSHOT_STEP1.sql')
    print('\twriting to', snapfile)
    with open( snapfile, 'w') as fw:
        fw.write( scd2_sql )
    #table_split,split=table.split('_STEP1')
    if table.startswith('DIM'):
        model_file = modeldir/(table+'.sql')
    else:
        model_file = modeldir/(table+'_alt.sql')
    print('\twriting to', model_file)
    with open( model_file, 'w' ) as fw:
        fw.write( scd6_sql )


def get_scdinfo( schema, table, col_rows):
    '''
    {'Order#': None, 'SCD TYPE': None, 'Staging Layer Column Name': None, ' 'CPI': None, 
    'Unique': None, 'Not Null': None, 'Test Expression': None, 'Accepted Values': None}
    '''

    scd = {
        'table': table,
        'schema': schema,
        'KEY':'',
        'scd1':[], 'scd2':[], 'scd3':[], 'scd0':[],
    }

    for row in col_rows:
        col = row[ 'Staging Layer Column Name' ]
        if not col: continue
        if str(row[ 'SCD' ]).upper() == 'KEY':  scd[ 'KEY' ] = col.upper()

    for row in col_rows:
        col = row[ 'Staging Layer Column Name' ]
        #print('2',row)
        if not col: continue
        if scd[ 'KEY' ] == col.upper(): continue

        if not row[ 'SCD' ]: 
            if row[ 'SCD' ] != 0: 
                row[ 'SCD' ] = '1'
        if col not in scd[ 'scd'+ str(row[ 'SCD' ])]:
            scd['scd'+ str(row[ 'SCD' ])].append(col)
        
        # print('~~SCD:' , row['SCD'] )                                 # DEBUG
        # print('SCD APPEND >>> ',scd['scd' + str(row[ 'SCD' ])])       # DEBUG
        # print('2', scd['scd'+str(row[ 'SCD' ])])                      # DEBUG

    return scd


def write_scd_files( modeldir, tgt_table, scd1_sql, scd2_sql):
    scd1_file = modeldir/(tgt_table+'_SCD1.sql')
    print('\twriting to', scd1_file)
    with open( scd1_file, 'w' ) as fw:
        fw.write( scd1_sql )

    scd2_file = modeldir/(tgt_table+'_SCD2.sql')
    print('\twriting to', scd2_file)
    with open( scd2_file, 'w' ) as fw:
        fw.write( scd2_sql )


def write_final_scd_sql( args, tgt_table, scd_dict ):
    modeldir = args.modeldir
    
    scd1_r_cols = []
    scd1_cols = scd_dict[ 'scd1' ] 
    #+ scd_dict['scd0']

    exclude_col=[ 'UNIQUE_ID_KEY', 'CURRENT_INDICATOR', 'EFFECTIVE_DATE', 'END_DATE', 'LOAD_DATETIME', 'UPDATE_DATETIME', 'EFFECTIVE_TIMESTAMP', 'END_TIMESTAMP',
    'LOAD_BATCH_ID', 'UPDATE_BATCH_ID', 'PRIMARY_SOURCE_SYSTEM' ]
    scd1_cols = [ col.upper() for col in scd1_cols ]
    exclude_col = [ col.upper() for col in exclude_col ]

    scd1_remaining_cols = set( scd1_cols ) - set( exclude_col )

    scd1_str = '\n, '.join( scd1_remaining_cols )
    keycol = scd_dict[ 'KEY' ]
    
    src_table = tgt_table.replace( 'DIM_', 'DSV_' )

    # prefix = ''' {{ config(schema = 'EDW_STAGING', tags=["dim"] ) }}\n ----SRC LAYER----\nWITH\n'''
    prefix = '''\n ----SRC LAYER----\nWITH\n'''

    sql1 = prefix+f'''SCD1 as ( SELECT {scd1_str} \n, {keycol}    \n\t--, '1901-01-01' as DBT_VALID_FROM, '2099-12-31' as DBT_VALID_TO\n\tfrom      {{{{ ref( '{src_table}') }}}} )'''
    sql2 = f'''SCD2 as ( SELECT *    \n\tFROM      {{{{ ref('{tgt_table}_SNAPSHOT_STEP1') }}}} )'''
    sql3 = f'''FINAL as ( SELECT * 
            FROM  SCD2 
                INNER JOIN SCD1 USING( {keycol} )  )\nselect * from FINAL'''

    all_sql_str = '\n\t,'.join([ sql1, sql2, sql3 ])

    final_file = modeldir/(tgt_table+'_SCDALL_STEP2.sql')
    print( '\twriting to', final_file )
    with open( final_file, 'w' ) as fw:
        if not args.yml: fw.write( all_sql_str )  
    print('========== done')
###### END OF SCD SPECIFIC ADDED CODE


def get_source_tables( col_rows, alias):
    # This is visited for each table in the Tables tab.
    tables_raw = {}
    tables = {}
    schemas = {}
    cpi = {}
    select_cols = []
    missing_keys = []
    expected_keys = ['Source Schema', 'Source Table', 'Source Column', 'Datatype', 'Automated Logic', 'Manual Logic', 'Order#', 'Staging Layer Column Name', 
    'Staging Layer Datatype', 'CPI', 'Unique', 'Not Null', 'Test Expression', 'Accepted Values', 'SCD', 'Remove Column',
    'Equal Rowcount', 'Mutually Exclusive Ranges', 'Unique Combination of Columns', 'Accepted Range']
    RC = 'Remove Column'

    for idx, row in enumerate(col_rows):
        is_derived = False
        # print( f'>> Looping through rows: {row}')       
        # print( f'>> Rows have the following keys: {row.keys()}')
        # print( f'<< Compared to:                  {expected_keys}' )
        for each_value in expected_keys:
            if each_value in row.keys(): continue
                # print( f'The key {each_value} is present...')
            else:
                missing_keys.append( each_value )
        
        if not row.get( 'Source Table' ):
            continue
        indent = nl = ''
        #if idx % 5 == 0:
        #    nl = '\n\t\t\t\t'
        src_table = row[ 'Source Table' ].strip()

        if '(' in src_table:
            src_table = src_table[ :src_table.find('(') ].strip()
        if src_table not in tables:
            tables[ src_table ] = []
            tables_raw[ src_table ] = []
            cpi[ src_table ] = []

        tgt_col = row[ 'Staging Layer Column Name' ].strip()

        # if there is a value in the REMOVE COLUMN in XLS, do not include it in the list of source columns
        if RC in row.keys():
            if not row[ 'Remove Column' ]:
                # print( f'< ADDING {tgt_col.upper().strip()} to select_cols: {select_cols}' )            # DEBUG
                defaultval = row.get('Set Default') 
                if not defaultval == None:
                    modtgt_col = f"coalesce( {tgt_col}, MD5( '-1111' )) as {tgt_col}"
                    select_cols.append( modtgt_col.strip() )
                else:
                    select_cols.append( tgt_col.upper().strip() )


        src_test = row[ 'Source Column' ]
        # if there is no Source Column, assume (DERIVED)
        if not src_test:
            src_col = '(DERIVED)'
            is_derived = True
        else:
            src_col = row[ 'Source Column' ].strip()
            if '(DERIVED)' in src_col:
                is_derived = True
        # print( f'+++ New src_col: {src_col}')                         # DEBUG

        # if a Source Column is listed as:  (DERIVED) columnname
        if is_derived:
            derivedname = src_col.upper().replace('(DERIVED)','').strip()
            if derivedname == '':
                derivedname = tgt_col
        else:
            derivedname=''
        
        #If DERIVED, the source and target are set to the same value
        #unless there is Manual Logic driving the value
        if is_derived and row.get( 'Manual Logic','') :
            src_col = tgt_col
        elif is_derived and not row.get( 'Manual Logic',''):
            if not derivedname == '':
                src_col = derivedname
            else:
                src_col = tgt_col

        #if derived and has manual logic, use that, unless it's a HASH
        elif is_derived and 'HASH' not in row[ 'Manual Logic' ]:
            src_col = row[ 'Manual Logic' ].strip()
        elif src_col != tgt_col:
            src_col = src_col
        else:
            src_col = tgt_col

        final_col = f'{src_col:<50} as {tgt_col:>50}'
        
        tables[ src_table ].append( nl + final_col )
        tables_raw[ src_table ].append( final_col )
        for tablekey, tableval in alias.items():
            if tablekey == src_table:
                act_table = tableval
                schemas[act_table] = row[ 'Source Schema' ]
                continue

        if row[ 'CPI' ] == 'CPI':
            cpi[ src_table ].append( f'SHA256({src_col}) as {src_col}' )
        else:
            cpi[ src_table ].append( src_col )

    missing_keys = list(set( missing_keys ))
    print( f'~~~~ XLS file is Missing Key Columns: {missing_keys}\n')
    # print( f'\n\n<<<<< DEBUGGING get_source_tables:\nSCH: {schemas.items()}\n TR: {tables_raw.items()}\n  T: {tables.items()}\nCPI: {cpi.items()}\nCOL: {select_cols}' )      # DEBUG
        
    return schemas, tables_raw, tables, cpi, select_cols

##################################################
# Building out the Extract SQL statement
##################################################
def build_extract_sql( schemas, col_rows, alias ):
    '''
    COLUMN_ROWS: [{
    'Source Schema': 'STAGING', 
    'Source Table': 'INV', 
    'Source Column': '[DERIVED]', 
    'Datatype': None, 
    'Automated Logic': None, 
    'Manual Logic': 'Hash Key: INVOICE_NUMBER', 
    'Order#': 1, 
    'Staging Layer Column Name': 'UNIQUE_ID_KEY', 
    'Staging Layer Datatype': 'CHAR(32)', 
    'CPI': None, 
    'Unique': None, 
    'Not Null': 'not_null', 
    'Test Expression': None, 
    'Accepted Values': None, 
    'SCD': KEY, 1 or 2 
    None: None}]
    '''

    all_sql = []
    col_rows_final = {}

    for table, cols in col_rows.items():
        if table not in col_rows_final:
            col_rows_final[ table ] = []
        for col in cols:
            if col not in col_rows_final[ table ]:
                col_rows_final[ table ].append( col )

    comb_table = {}
    for table, cols in col_rows_final.items():
        for k, v in alias.items():
            if k == table:
                if v not in comb_table:
                    comb_table[ v ] = cols
                else:
                    for each in cols:
                        if each in comb_table[ v ]: continue
                        comb_table[ v ].append( each )
            elif v == table:
                if v not in comb_table:
                    comb_table[ v ] = cols
                else:
                    for each in cols:
                        if each in comb_table[ v ]: continue
                        comb_table[ v ].append( each )

    # print( f'=======\nCOMB_TABLE: {comb_table}')          # DEBUG
    # print( f'COL_ROWS_FINAL: {col_rows_final}')           # DEBUG

    for table, cols in comb_table.items():
        cols_str =  ', '.join( cols )

        sql = f'''SELECT {cols_str} \n\t\t\t\tFROM {schemas[ table ]}.{table};
            python /usr/local/etl/run/run_sql2csv.py  --fname /cifs/mswg9/groups/IT/ETL/snowflake/extracts/{table}.csv --operation vsql --t_schema BASE  --t_environment uat2 --tgtdb vertica --sql "SELECT {cols_str} from  {schemas[ table ]}.{table}"

        '''
        all_sql.append(sql)
    
    # print( f'\n\n<<<< BUILD EXTRACT\n{cols_str}' )                # DEBUG
    return all_sql


def build_ddl_sql( rows, alias ):
    '''
    COLUMN_ROWS: [{
    'Source Schema': 'STAGING', 
    'Source Table': 'INV', 
    'Source Column': '[DERIVED]', 
    'Datatype': None, 
    'Automated Logic': None, 
    'Manual Logic': 'Hash Key: INVOICE_NUMBER', 
    'Order#': 1, 
    'Staging Layer Column Name': 'UNIQUE_ID_KEY', 
    'Staging Layer Datatype': 'CHAR(32)', 
    'CPI': None, 
    'Unique': None, 
    'Not Null': 'not_null', 
    'Test Expression': None, 
    'Accepted Values': None, 
    'SCD TYPE': 1, 
    None: None}
    ]

    create or replace table POLICY_OWNERS_ADDR ( 
        POLICY_ID   number(38,0),
        OWNER_ADDR  text,
        ZIP         number(38,0),
        VOID_IND    text
    );
    '''

    tables = {}
    insert_cols = {}
    insert_vals = {}
    unique_cols = set()
    tbl = ''

    for idx, row in enumerate(rows):
        if not row.get('Source Table'):
            continue
        col = row['Source Column']
        col = col.replace('(DERIVED) ','')
        stg_tgt_dtype = row['Staging Layer Datatype']

        # print( f'\n====== ALIAS: {alias.items()}')              # DEBUG
        for tbl_alias, tbl_name in alias.items():
            tbl_alias = tbl_alias.strip()
            if tbl_alias == row['Source Table'].strip():
                act_table = tbl_name
                tbl = act_table
                continue

        tbl_col = tbl, col
            
        if tbl_col in unique_cols:
            continue
        unique_cols.add(tbl_col)

        if tbl not in tables:
            tables[tbl] = []
        if tbl not in insert_cols:
            insert_cols[tbl] = []
        cleandt = str(stg_tgt_dtype).replace('\xa0', '').strip()

        cleanstr = f'{str(col):<50} {str(cleandt):>50}'
        tables[tbl].append(f'{cleanstr}')

        # print( f'>>>> {tables[ tbl ]}')                    # DEBUG

        insert_cols[tbl].append(col)
        val = 'x'
        if stg_tgt_dtype == None:
            val = "'textval'"
        elif 'INT' in stg_tgt_dtype.upper() or 'NUM' in stg_tgt_dtype.upper():
            val = str(idx)
        elif 'DATE' in stg_tgt_dtype.upper():
            val = "'2222-02-22'"
        else:
            val = "'textval'"

        if tbl not in insert_vals:
            insert_vals[tbl] = []
        insert_vals[tbl].append(val)

    # print( f'\n\n >>>> TABLE DEFINITION\n{tables[ tbl ]}')                    # DEBUG

    ddl_sql = []
    for tbl, cols in tables.items():
        col_str = ',\n\t'.join(cols)

        sql = f"'create or replace TABLE {tbl} (\n\t{col_str}\n\t);'"
        ddl_sql.append(sql)

    # insert into mytable (col1, col3) select to_date('2222-02-22'), to_timestamp('2222-02-22T23:59:59');

    inserts_sql = ['\n']
    for tbl, cols in insert_cols.items():
        val_str = ', '.join(insert_vals[tbl])
        col_str = ', '.join(cols)                 # // debug

        sql = f"'-- insert into {tbl} ( {col_str} )  values ( {val_str} ); '"
        inserts_sql.append(sql)

    all_sql = ddl_sql + inserts_sql

    return all_sql


def get_config( alter_sql, tgt_table, dml_block ):
    '''
    Builds the CONFIG dynamically based on the file name
        # {{ config(   only if...:
        # ADD: materialized = 'view' if the tgt_table starts with DSV
        # ADD: tags = [ "fact" ]  if the tgt_table starts with FACT%
        # Only add post_hook = (" if there are alter statements.
        # ") <- close post_hook 
        # }} <- close config
    '''
    concat_config = '' 
    config = {}
    compileif = ['view','tags','alter']

    if tgt_table.startswith('DSV'):
        config['view'] = f"materialized = 'view'"
    elif tgt_table.startswith( 'FACT' ):
        config['tags'] = f'tags = [ "fact" ]'

    # Dimension, Fact and Factless Fact tables will get post hook (Alter statements for Referential Integrity)
    if tgt_table.startswith('DIM') or tgt_table.startswith('FACT') or tgt_table.startswith('FLF') :
        if len(alter_sql) > 1: config['alter'] = f'post_hook = ("\n{alter_sql.strip()} {dml_block}\n") '

    compile_cnt = 0
    for checkval in compileif:
        if checkval in config and compile_cnt == 0:
            concat_config = f'{{{{ config( \n{config.get(checkval)} '
            compile_cnt += 1
        elif checkval in config:
            concat_config += f'\n,\t {config.get(checkval)}'
            compile_cnt += 1
    if compile_cnt > 0:
        concat_config  += '\n) }}'

    return concat_config


def build_source_sql(args, tables, alias):
    all_sql = []
    act_table = ''
    # print( '--- BUILDING SOURCE SQL' )
    for table, cols in tables.items():
        cols_str = ', '.join(cols)
        for k, v in alias.items():
            if k == table:
                act_table = v
                continue
        # TODO:  remove entirely if all tables will be staged via DBT.  They will then all have model references.
        # if not alias['Schema'].upper() == 'STAGING' :
            # sql = f"SRC_{table} as ( SELECT *     from     {{{{ source('CORESUITE', {act_table!r}) }}}} ),"
            # sql = f"SRC_{table} as ( SELECT *     from     {{{{ ref( {act_table!r} ) }}}} ),"
        # else:
        sql = f"SRC_{table:<14} as ( SELECT *     FROM     {{{{ ref( {act_table!r} ) }}}} ),"
        all_sql.append(sql)

    for table, cols in tables.items():
        cols_str = ', '.join(cols)
        for k, v in alias.items():
            if k == table:
                act_table = v
                continue

        sql = f"//SRC_{table:<14} as ( SELECT *     FROM     {act_table}) ,"
        all_sql.append(sql)

    if not args.silent:
        print(f'\n=== Build SOURCE SQL\n{all_sql}')
    return all_sql


def build_logic_sql( args, col_rows, incl_trim):
    # Rewrite this whole process to build them as lists of dictionaries
    tables = {}
    
    for idx, row in enumerate(col_rows):
        if idx == 0: firstcol = True
        if idx == 0 and not row.get( 'Source Table' ):
            if row.get('Staging Layer Column Name','') == 'UNIQUE_ID_KEY':
                continue
        elif idx > 0 and row.get( 'Source Table') == None:
            tbl = 'tbd'
            tgt_col = row.get( 'Staging Layer Column Name','') 
            if row.get( 'Source Column','' ) == None:
                src_col =  row.get( 'Source Column','' ) 
            elif row.get( 'Source Column','')  =='(DERIVED)' :
                src_col = f'(DERIVED) {tgt_col}' 
                
            col = src_col
        else:            
            tbl = row.get( 'Source Table','' )
            if not row.get( 'Source Column','' ) == None:
                src_col = row.get( 'Source Column','')
            else:
                src_col = ''
            col = src_col
        tgt_col = row.get( 'Staging Layer Column Name','' )
        mlogic = row.get( 'Manual Logic','' )
        set_default = row.get('Default')

        hasname = ''

        # If the column is derived, check for a column name and assign to HASNAME variable.
        is_derived = False
        if col and '(DERIVED)' in col.upper(): 
            is_derived = True
            hasname = col.upper().replace('(DERIVED)','').strip()

        # If there is not Manual Logic, checkf or Automated logic and use that in its place.  Manual Trumps automatic.
        if not row.get( 'Manual Logic' ):
            # Get Automated logic as a substitute
            if row.get( 'Automated Logic'):
                mlogic = row.get( 'Automated Logic' )
                mlogic = mlogic.upper()
            else:
                mlogic = ''      
        else:
            mlogic = row.get( 'Manual Logic' )
    
        # If the Logic is a HASH/COMPOSITE for derived columns, create the field as SURROGATE_KEY
        dtype = row[ 'Datatype' ]
        if is_derived:
            # col = '--- NEW ' + row[ 'Staging Layer Column Name' ]
            if dtype == None:
                mlogic = mlogic.upper()
                if 'HASH' in mlogic or 'COMPOSITE' in mlogic:
                    hashcol = ''
                    hashcol = mlogic.replace('HASH:', '').replace('COMPOSITE:', '').replace('\n','').replace(' ','').strip()
                    temp1 = hashcol.split(',')
                    joincol = (', ').join(temp1)
                    hashcol = ("','").join(temp1)
                    col1 = f"{{ dbt_utils.generate_surrogate_key ( [ '{hashcol}' ] ) }}" 
                    surrogate = f"{{{col1}}}"
                elif mlogic:
                    # If not HASH or COMPOSITE, but has logic, just use Manual Logic as is
                    col = mlogic.strip()
                else:
                    col = src_col
                    # col = 'NULL'
            # If the field is DERIVED and a DATE field, then it will get a special case statement built for the logic.
            # This requires a column-name after (DERIVED)
            elif dtype == 'DATE':
                col = f"CASE WHEN {hasname} is null then '-1' \n\t\t\tWHEN {hasname} < '1901-01-01' then '-2' \n\t\t\tWHEN {hasname} > '2099-12-31' then '-3' \n\t\t\tELSE regexp_replace( {hasname}, '[^0-9]+', '') \n\t\t\t\tEND :: INTEGER {' '*40}"

        # Check to see if there is a 'Set Default' for NULL values and wrap col in it.  If it's HASH or COMPOSITE, test for any null and use the set default MD5.
        # If there is no column name, default to just MD5 default.  
        # If there is a default/use replace nulls with the MD5 using coalesce
        if set_default and not ( col == 'NULL' or col == ''):
            if is_derived:
                if 'HASH' in mlogic or 'COMPOSITE' in mlogic:
                    if ',' in joincol:
                        col = f"CASE WHEN  nullif(array_to_string(array_construct_compact( {joincol} ),''),'') is NULL \n\t\t\tthen MD5( '{set_default}' ) ELSE {surrogate} \n\t\t\t\tEND {' '*40}"
                    else:
                        col = f"CASE WHEN {joincol} is NULL \n\t\t\tthen MD5( '{set_default}' ) ELSE {surrogate} \n\t\t\t\tEND {' '*40}"
                elif hasname  == '':
                    col = f"MD5( '{set_default}' )"
                else:
                    col = hasname
                    col = f"COALESCE( {hasname},  MD5( '{set_default}' ) ) "
            else:
                col = f"COALESCE( {col}, MD5( '{set_default}' ) ) "
        else:
            if is_derived:
                if 'HASH' in mlogic or 'COMPOSITE' in mlogic:
                    col = f" {surrogate} "
            
                    
        # tgt_col = str(row['Staging Layer Column Name']).strip()


        #Only do these for Automated Logic
        logic = row.get( 'Automated Logic' )
        if logic:
            logic = logic.strip()

        # INVOICE_HEADER > HDR from alias lookup
        # REF (R1) -> R1
        if '(' in tbl:
            tbl = tbl[ :tbl.find('(') ]

        src_tbl = tbl
        tgt_tbl = f'LOGIC_{tbl}'

        if src_tbl not in tables:
            firstcol = True
            tables[ src_tbl ] = []

            indent = False
        if dtype and ( 'VARCHAR' in dtype.upper() or 'TEXT' in dtype.upper() ):
            indent = True
            if incl_trim: 
                col = f"NULLIF( TRIM( {col} ),'')"
            else: 
                col = f'{col}'

        if logic:
            indent = True
            if logic.lower() == 'cast date':
                col = f'cast( {col} as DATE )'
            elif logic.lower() == 'cast text':
                col = f'cast( {col} as TEXT )'
            elif logic.lower() == 'upper_trim' or logic.lower() == 'upper,trim' or logic.lower() == 'trim':         # Newly added
                col = f'upper( trim( {col} ))'
            elif logic.lower() == 'upper':
                col = f'upper( {col} )'
            else:
                print(f'!! Unsupported Operation !! <{logic}>')

        # Add the AS ... layer here to alias back to the column
        src_col ='' if not src_col else src_col
        src_col = src_col.upper().replace('(DERIVED)','')

        col = '' if not col else col
        if firstcol == True or col.startswith('--- NEW'):
            if len(col)>50:
                col = f"\n\t\t  {col}\n{' ':<60} as {src_col:>50}"
            else:
                col = f"\n\t\t  {col:<50} as {src_col:>50}"
                firstcol = False
        else:
            if len(col)>50:
                col = f"\n\t\t, {col:<50}\n{' ':<60} as {src_col:>50}"
            else:
                col = f"\n\t\t, {col:<50} as {src_col:>50}"

        # Only 1st time col is mentioned, it gets appended
        if col not in tables[ src_tbl ]:
            if not tgt_col:
                pass
            else:
                tables[ src_tbl ].append(col)

    final_sql = []

    has_derived = False
    for tbl, cols in tables.items():
        # if 'tbd' is in the tables, then we have derived columns and need to move them to the final select
        if tbl == 'tbd':
            has_derived = True
            # Skip the writing of the logic layer and only write for final select
            pass
        else:
            # if not defined, select everything via asterisk (*)
            if not cols:
                cols_str = ' * '
            else:
                cols_str = ' '.join(cols)
            tgt_tbl = f'LOGIC_{tbl}'            # Removing 'SRC_'
            sql = f'''\n{tgt_tbl} as ( SELECT {cols_str} \n\t\tFROM SRC_{tbl}
                )'''
            final_sql.append( sql )

            cleanprint = ' '.join( final_sql )                           # DEBUG
            if not args.silent:
                print(f'\n==== Build LOGIC SQL\n {cleanprint}')          # DEBUG

    joined_sql = ',\n'.join( final_sql )
    if has_derived:
        derived_sql = '\n---- DERIVED COLUMNS need to be finalized and moved into the right position in FINAL SQL\n'
        # for col in tables['tbd']:
        #     derived_cols.append( col )
        derived_sql += ''.join( tables['tbd'] )
        derived_sql += '\n---END of DERIVED Columns'
    else: derived_sql = ''
    
    return joined_sql, derived_sql


def build_rename_sql( args, tables):
    rename_sql = []

    # print(f'\n\n<<<<< DEBUGGING:  BUILD RENAME\n{tables}')        # DEBUG
    for table, cols in tables.items():
        cols_str = '\n\t\t, '.join(cols)
        sql = f'\nRENAME_{table:<10} as ( SELECT \n\t\t  {cols_str.strip() } \n\t\t\t\tFROM     LOGIC_{table}   )'
        rename_sql.append(sql)

    all_sql = ', '.join(rename_sql)                                # DEBUG
    if not args.silent:
        print(f'\n==== Build RENAME SQL\n{all_sql}')                   # DEBUG
    return all_sql


def build_filter_sql(table_rows):
    filter_sql = []
    for row in table_rows:
        if not row['Filter Conditions']:
            filters = []
        else:
            filters = row['Filter Conditions'].split(';')
        filters_str = ''
        if filters:
            filters_str = f"\n{'WHERE ':>50}"
            filters_str += ' AND '.join(filters)

        alias = row['Alias']
        from_table = alias or row['Source Table']

        if not alias:
            if not row['Source Table']:
                continue
        source_table = f'RENAME_{from_table}'

        sql = f'FILTER_{from_table:<30} as ( SELECT * FROM    {source_table} {filters_str}  ),'
        filter_sql.append(sql)

    # cleanprint = ' '.join( filter_sql )                             # DEBUG
    # print( f'\n==== Build FILTER SQL\n{cleanprint}')                # DEBUG
    return filter_sql


def build_join_sql(tgt, table_rows, select_cols):
    join_sql = []

    #{ 'POLICY_CLAIMS':'P', 'STG_POLICYTEST':'T', 'POLICY_OWNERS':'O', 'POLICY_OWNERS_ADDRESS':'POA' }
    final_alias = ''
    aliases = {}
    for row in table_rows:
        if not row['Source Schema']:
            continue
        alias = row.get('Alias','')
        aliases[row.get('Source Table','tbd')] = alias
        if not row['Join Type']:
            final_alias = alias

    join_lvls = build_join_levels(table_rows)

    all_sql = []

    made_cte = set()
    # print( f'\n<< JOIN LEVELS: {join_lvls}\n')                               # DEBUG
    # All join_lvls should have a value.  If not, uncomment the above and see where the issue is.

    for lvl in sorted(join_lvls):
        # print( f'<< Currently on lvl: {lvl}')                               # DEBUG
        parent_table, join_sql = do_join( made_cte, join_lvls[ lvl ] )
        made_cte.add( parent_table )
        join_sql_str = '\n\t\t'.join(join_sql)
        
        if 'COALESCE(' in parent_table.upper():
            wascoalesce = parent_table.upper().replace('COALESCE(','').strip()
            sql = f'{wascoalesce} as ( SELECT * \n\t\t\t\tFROM  FILTER_{wascoalesce}\n{join_sql_str} )'
        else:
            sql = f'{parent_table} as ( SELECT * \n\t\t\t\tFROM  FILTER_{parent_table}\n{join_sql_str} )'
        all_sql.append(sql)

    # Used by final select
    if select_cols:
        select = ''
        for col in select_cols:
            if col == select_cols[ 0 ]:
                select += f'\n\t\t  {col}'
            else:
                select += f'\n\t\t, {col}'
        # select = select.strip(',')
    else:
        select = '*'

    # If final SQL, then append the tdb columns


    final_sql = f'\nSELECT {select} \nFROM {final_alias}'
    if len( all_sql ) > 0:
        if not all_sql:
            sql = f' JOIN_{alias:<10}  as  ( SELECT * \n\t\t\t\tFROM  FILTER_{alias} )'
            all_sql = [sql + '\n' + f' SELECT * FROM  JOIN_{alias}']
        else:
            all_sql[-1] = all_sql[-1] + final_sql

    all_sql_str = ',\n'.join(all_sql)

    # print( f'\n\n=== Build JOIN SQL\n{all_sql}')                   # DEBUG
    return all_sql_str


def do_join(made_cte, table_rows):
    '''
        select *
            from FPO
                LEFT JOIN FPAD USING( POLICY_ID )
                LEFT JOIN FPAD2 on FPO.POLICY_ID = FPAD2.POLICY_IDX;
    '''
    join_sql = []
    parent_join_list = []
    # The sorted_rows group items together by Parent Join Alias with INNER JOINS being outdented
    sorted_rows = []
    sorted_rows = sorted( table_rows, key = itemgetter( 'Parent Join Alias' ))
    
    # for row in table_rows:
    for row in sorted_rows:
        join_type = row[ 'Join Type' ]
        if not join_type:
            continue
        join_type = join_type.replace('_', ' ').upper()
        source_table = row[ 'Alias' ]
        parent_join = row[ 'Parent Table Join' ]
        child_join = row[ 'Child Table Join' ]
        parent_table = parent_join.split('.')[ 0 ]

        using_join = ''
        parent_join_col = parent_join.split('.')[ -1 ]
        child_join_col = child_join.split('.')[ -1 ]

        if parent_join_col == child_join_col:
            using_join = f' using( {parent_join_col} ) '

        if 'coalesce' in parent_table:
            tempparent = parent_join.upper().replace('COALESCE(','').strip()
            parent_join = f'coalesce( FILTER_{tempparent}'
        else:
            parent_join = f'FILTER_{parent_join}'
            
        if source_table in made_cte:
            sql = f'\t\t\t\t{join_type} {source_table} ON  {parent_join} = {child_join} '
            if using_join:
                sql = f'\t\t\t\t{join_type} {source_table} {using_join} '
        else:
            if parent_table in parent_join_list and not join_type.lower().startswith('inner'):
                sql = f'\t\t\t\t\t\t{join_type} FILTER_{source_table} ON  {parent_join} =  FILTER_{child_join} '
            else:
                sql = f'\t\t\t\t{join_type} FILTER_{source_table} ON  {parent_join} =  FILTER_{child_join} '
                if using_join:
                    sql = f'\t\t\t\t{join_type} {join_type}  FILTER_{source_table} {using_join} '
                parent_join_list.append( parent_table )

        join_sql.append( sql )

        # print( f'\n\n==== DO JOIN\nPT:{parent_table}\nSQL:{join_sql}')
    return parent_table, join_sql


def build_join_levels(table_rows):
    lvls = {}
    for row in table_rows:
        if not row[ 'Parent Table Join' ]:
            continue
        lvl = row[ 'Parent Join Number' ]
        if not lvl: lvl = 0
        if lvl not in lvls:
            lvls[ lvl ] = []
        # Add Parent Join Alias to determine if it's a sub join
        row[ 'Parent Join Alias' ] = row[ 'Parent Table Join'].split('.')[0]
        lvls[ lvl ].append(row)

        # print( f'\n\n---- Build JOIN LEVELS\n{lvls[ lvl ]} ')                   # DEBUG
    return lvls


def make_yml(tgt, col_rows, table_rows):
    '''
    version: 2

models:
    - name: orders
    tests:
    - dbt_utils.expression_is_true:
            expression: "colA + colB = totalAB"
    - dbt_utils.equality:
            compare_model: ref( 'other_table_name' )
    columns:
    - name: order_id
        tests:
            - unique
            - not_null
            - dbt_utils.expression_is_true:
                expression: "colA + colB = totalAB"
    - name: status
        tests:
            - accepted_values:
                values: [ 'placed', 'shipped', 'completed', 'returned' ]
    - name: REVENUE_CENTER_HKEY
        tests:
            - dbt_utils.relationships_where:
                to: ref('DIM_REVENUE_CENTER')
                field: REVENUE_CENTER_HKEY
                from_condition: REVENUE_CENTER_HKEY <> '40c5dea533476acdd01f7ef0e84de22f'
    '''

    modeltests = []
    for row in table_rows:
        modeltests += get_model_tests(row)
        
    print( f'<<< Going in with modeltests: {modeltests}' )        # DEBUG

    schema_dict = {
        'version': 2,
        'models': [
            {'name': tgt,
             'tests': modeltests,
             'columns': [],
             }
        ]
    }
    
    # print( f'@@@ ADDED MODELTEST: {schema_dict}' )                  # DEBUG


    ymlcols = []
    cols_processed = set()
    for row in table_rows:
        col = row['Staging Layer Column Name']

        if not col:
            continue
        if col in cols_processed:
            continue
        cols_processed.add(col)

        tests = get_column_tests(row, tgt)

        if tests:
            ymlrow = {
                'name': col,
                'tests': tests,
            }
            ymlcols.append(ymlrow)

    y = {}
    y = schema_dict['models'][0]
    y['columns'] = ymlcols

    return schema_dict


def get_model_tests(row):
    '''
    https://github.com/fishtown-analytics/dbt-utils
    '''
    test_dict = {}
    test_cols = ['Unique', 'Model Equality', 'Model Equal Rowcount', 'Test Expression',  ]
    # Globally replacing the relationship with relationship_where

    tests = []
    add = {'severity': 'warn'}

    for atest in test_cols:
        if atest not in row:
            continue
        if not row[atest]:
            continue

        if atest == 'Unique':
            if not row.get(atest).lower() == 'unique' :
                test =   str(row[ atest ]).replace('COMPOSITE:','').strip()
                row[atest] = {
                    'dbt_utils.unique_combination_of_columns':
                    {
                        'combination_of_columns': [each.strip() for each in test.split(',')],   
                        'severity': 'warn'
                    }
                }
            else: continue

        if atest == 'Model Equality':
            # print(f' $$$$$$$$ FOUND Model Equality ->  {row[ atest ]}')               # DEBUG
            test =   str(row[ atest ]).strip()
            row[atest] = {
                'dbt_utils.equality': 
                {
                    'compare_model': f"ref( {row[ 'Staging Layer Column Name' ]} )",
                    'compare_columns': [each.strip() for each in test.split(',')]                    
                    # 'compare_columns': [ f'ref({row[ atest ]!r})' ]
                }
            }

        if atest == 'Model Equal Rowcount':
            # print(f' $$$$$$$$ FOUND Model Equal Rowcount -=>  {row[ atest ]}')        # DEBUG
            row[atest] = {
                'dbt_utils.equal_rowcount': 
                {
                    'compare_model': f'ref({row[ atest ]!r})'
                }
            }

        # Relationship moved to column test...

        if atest == 'Test Expression':
            if row.get('Staging Layer Column Name','x') in row.get('Test Expression','y'):
                # If true, then should be a model test, else column test
                test =   str(row[ atest ]).strip()
                row[atest] = {
                    'dbt_utils.expression_is_true':
                    {
                        'severity': 'warn',
                        'expression': test
                    }
                }
            else: continue
        tests.append(row[atest])

    return tests


def get_column_tests(row, tgt):
    '''
    https://github.com/fishtown-analytics/dbt-utils
    '''
    test_dict = {}
    test_cols = [ 'Unique', 'Not Null', 'Test Expression', 'Accepted Values', 'Relationship', 'Set Default']
    # 'Equal Rowcount',  'Mutually Exclusive Ranges', 'Unique Combination of Columns', 'Accepted Range']

    tests = []
    add = {'severity': 'warn'}

    for atest in test_cols:
        if atest not in row:
            continue
        if not row[atest]:
            continue

        # If a FLF or FACT build, do not include the has_default test.  Will require knowing the layer during build.
        if atest == 'Set Default':
            test = str( row[ atest ] ).strip()
            row['Default']= test
            if tgt.startswith('F'):
                row[atest]=''
            else:
                row[atest] = {
                    'has_default':
                    {
                        'severity': 'warn',
                        'hashval': [each.strip() for each in test.split(',')]
                    }
                }

        if atest == 'Unique':
            row[atest] = {
                'unique':
                {
                    'severity': 'warn'
                }
            }

        if atest == 'Not Null':
            if 'where' in row[atest]:
                # TODO: Fix quoting around where
                test = str( row[ atest ] ).upper().replace( 'WHERE','').strip()
                row[atest] = {
                    'dbt_utils.not_null_where':
                    {
                        'where': test
                    }
                }                
            else:
                row[atest] = {
                    'not_null':
                    {
                        'severity': 'warn'
                    }
                }

        if atest == 'Accepted Values':
            if 'not:' in row[atest]:
            # Split to break them into separate rows...
                test = str( row[ atest ] ).upper().replace( 'NOT:','').strip()
                row[atest] = {
                    'not_accepted_values':
                    {
                        'severity': 'warn',
                        'values': [each.strip() for each in test.split(',')]
                    }
                }
            else:
                test = str( row[ atest ] )
                row[atest] = {
                    'accepted_valids':
                    {
                        'ignore': '',
                        'severity': 'warn',
                        'values': [each.strip() for each in test.split(',')]
                    }
                }

        if atest == 'Test Expression':
            # If the column name is not listed, then it can be a column test, else model test
            if not row.get('Staging Layer Column Name','x') in str(row[ atest ]).strip():
                test =   str(row[ atest ]).strip()
                row[atest] = {
                    'expression_is_true_ignore':
                    {
                        'ignore': '',
                        'severity': 'warn',
                        'expression': test
                    }
                }
            else: continue

        if atest == 'Relationship':
            ignore_in = ''
            # Changed to a relationships_where instead to handle ignoring dummy value:  md5('-1111') 
            # Determine if the field is different (Based on TableName.ColumnName )
            if '.' in row[atest]:
                ref_table = row[ atest ].split('.')[0]
                ref_field = row[ atest ].split('.')[1]
                ref_ignore = row[ 'Staging Layer Column Name' ]
            else:
                ref_table = row[ atest ]
                ref_field = row[ 'Staging Layer Column Name' ]
                ref_ignore = row[ 'Staging Layer Column Name' ]

            if ref_field == 'DATE_KEY':
                ignore_in = " -1, -2, -3 "
            else: 
                ignore_in = " '40c5dea533476acdd01f7ef0e84de22f', 'fcbcdcb8f6b1c597c5fdc7a54cd321ae' "

            row[atest] = {
                'dbt_utils.relationships_where':
                {
                    'to': f'ref( {ref_table!r} )',
                    'field': f"{ref_field}",
                    'from_condition': f"{ref_ignore} not in ( {ignore_in} ) "
                }
            }

        if not row[atest] == '':
            tests.append(row[atest])

    return tests


def build( args, modeldir, tgt_table, table_rows, col_rows, alter_sql, dml_block ):
    yml = make_yml( tgt_table, table_rows, col_rows )
    alias = {}
    newvalue = ''
    incl_trim = False

    for thisrow in table_rows:
        '''{'Source Schema': None, 'Source Table': None, 'Alias': None, 'Filter Conditions': None, 'Parent Join Number': None, 'Parent Table Join': None, 'Child Table Join': None, 'Join Type': None}'''
        thisrow_dict = {}
        for key, value in thisrow.items():
            if not value: value = ''
            if not key: key = ''
            thisrow_dict[key] = value
        thisrow = thisrow_dict

        for key, value in sorted( thisrow.items() ):
            if not key: continue
            if key not in ['Source Table', 'Alias', 'Source Schema']: continue
            if value not in alias:
                if 'Alias' in key:
                    alias[ value ] = []
                    newvalue = value
            if 'Source Table' in key:
                alias[ newvalue ] = value
            if 'Source Schema' in key:
                alias[ 'Schema' ] = value
            if 'Set Default' in key:
                alias[ 'Set Default' ] = value 

    schemas, table_cols, table_cols_formatted, cpi, select_cols = get_source_tables( col_rows, alias)

    # SQL to pull from oracle/db2/vertica
    extract_sql = build_extract_sql( schemas, cpi, alias )
    extract_sql_str = '\n'.join( extract_sql )

    ##################################################
    # IF NOT a STG_ file do not generate the EXTRACT_FROM_SRC file
    ##################################################    
    if tgt_table.startswith('STG') and not args.yml:
        with open(modeldir/(tgt_table + '_extract_from_src.sql'), 'w') as fw:
            fw.write( extract_sql_str )

    # SQL to pull from oracle/db2/vertica
    ddl_sql = build_ddl_sql( col_rows, alias )
    ddl_sql_str = '\n'.join( ddl_sql )
    if not args.silent:
        print(f'\n~~~~ DDL :\n{str( ddl_sql_str )}')  # DEBUG

    ##################################################
    # Potentially not used and can be removed.
    ##################################################
    # with open(modeldir/(tgt_table + '_create_in_snowflake_ddl.sql'), 'w') as fw:
    #     fw.write(
    #         f'USE WAREHOUSE WH_ETL;\nUSE ROLE ETL_ADMIN;\nUSE DATABASE DBTEST;\nUSE SCHEMA {args.schema} !!!;\n\n')
    #     fw.write(extract_sql_str)

    ################################
    # SRC LAYER
    config = get_config( alter_sql, tgt_table, dml_block )
    all_sql = build_source_sql( args, table_cols_formatted, alias )
    if all_sql:
        all_sql[ 0 ] = config + '\n\n---- SRC LAYER ----\nWITH\n' + all_sql[ 0 ]

    # LOGIC LAYER                   $$$ FIX FORMATTING
    # if tgt_table.startswith( 'DST' ) or tgt_table.startswith( 'STG' ): incl_trim = True
    if tgt_table.startswith( 'STG' ): incl_trim = True
    if all_sql:
        all_sql[ -1 ] += '\n\n---- LOGIC LAYER ----\n'
    else:
        all_sql = []
    logic_sql, derived_sql  = build_logic_sql( args, col_rows, incl_trim )
    all_sql.append( logic_sql )

    # RENAME LAYER
    all_sql[-1] += '\n\n---- RENAME LAYER ----\n,'
    # print( f'\n\n<<<<< DEBUGGING table_cols_formatted  going into RENAME LAYER\n{ table_cols_formatted }' )
    # {'INVOICE': ['\n\t\t\t\n\t\t\tINVOICE_NUMBER as MEDICAL_INVOICE_NUMBER']}    <- Table_cols_formatted.  Only one column
    # For every wb('Tables')['Alias'], walk through and find matching wb('Columns')['Source Table'].
    # for each match, bring ['Source Column'] as ['Staging Layer Column Name']
    rename_sql = build_rename_sql( args, table_cols_formatted )
    # all_sql += rename_sql
    all_sql.append( rename_sql )

    # FILTER LAYER
    all_sql[-1] += '\n\n---- FILTER LAYER (uses aliases) ----\n,'
    filter_sql = build_filter_sql( table_rows )
    all_sql += filter_sql

    # JOIN LAYER
    all_sql[-1] += '\n\n---- JOIN LAYER ----\n'
    join_sql = build_join_sql( tgt_table, table_rows, select_cols )
    all_sql.append( join_sql )
    # print( f'\n\n<<<<< JOINs LAYER \n{all_sql}')                     # DEBUG

    sql_str = '\n'.join( all_sql )
    sql_str +=  derived_sql
    
    ################################
    # Write out the final finals...
    # May have to loop back through col_rows to determine if there are any tbl: tbd 

    with open(modeldir/(tgt_table + '.yml'), 'w') as fw:
        # print( f'\n\n~~~~> YML:\n{yml}' )                          # DEBUG
        fw.write(yaml.dump(yml))

    ##### Output SQL for all other non-DIM sql (STG, DST, DSV)
    if not tgt_table.startswith( 'DIM' ) and not args.yml:
        with open(modeldir/(tgt_table + '.sql'), 'w') as fw:
            fw.write(sql_str)

    # print(f'\n\n===== CURRENT SQL =\n{sql_str}')                   # DEBUG


def process_unified_map( args, xls ):
    #  CHECK for new Layout:  DSV_TABLES, DST_TABLES, DIM_TABLES,
    #    DST_COLUMNS, DSV_COLUMS, DIM_COLUMNS
    new_sheets = [ 'STG Tables', 'STG Tables', 'DST Tables', 'DSV Tables', 'DIM Tables', 'DST Columns', 'DSV Columns', 'DIM Columns', 'DML']
    missing_sheets =[]
    expected_cols = [ 'Source Schema', 'Source Table', 'Source Column', 'Datatype', 'Automated Logic', 'Manual Logic', 'Order#' ,'Staging Layer Column Name' ,'Staging Layer Datatype', 'CPI' ,'Unique' ,'Not Null' ,'Test Expression' ,'Accepted Values', 'SCD' ]

    # If the workbook contains a DML tab, capture it to append to the CONFIG block
    print( '-'*30 )
    try:
        wb = openpyxl.load_workbook(xls)
    except ValueError as e:
        print(f'### ERROR: {e}\n ## > Please check if the following file is OPEN:\n\t {xls}')
        sys.exit(1)

    print(f'\n\n==== FOUND WORKBOOKS: {wb.sheetnames}')         # DEBUG
    for each_sheet in new_sheets:
        if each_sheet in wb.sheetnames: 
            continue 
        else:
            missing_sheets.append( each_sheet )

    if len( missing_sheets ) > 0:
        missing_sheets =  list(set( missing_sheets ))
        print( f'\n\n~~~ There are missing sheets for the UNIFIED format: {missing_sheets}')

    buildorder = ['STG','DST','DSV','DIM']
    for currentorder in buildorder:
        tgt_table = xls.stem.replace('UNIFIED', currentorder )
        curtable = f"{currentorder} Tables"
        curcol =   f"{currentorder} Columns"
        maxrow = wb[curtable].max_row
        print( f'Processing Unified Mapping for {tgt_table}' )
        print(f'{curtable} contains {maxrow} rows')
        
        if maxrow <= 2: 
            # Check to see if the First Cell of the table has a value, if not, append to missing_sheets
            testfield = wb[curtable].cell(2,1).value
            if testfield == None:
                missing_sheets.append( curtable )
                missing_sheets.append( curcol )
                continue
        else:
            # It is possible that all of the rows are empty, so check to see if the first cell of the first row has a value
            # If not, append to missing_sheets
            print( f'Checking {curtable} for empty rows' )
            testrows = wb[curtable].rows
            for idx, row in enumerate(testrows):
                if idx == 0: continue
                # Check to see if the first column is empty
                if row[0].value == None:
                    missing_sheets.append( curtable )
                    missing_sheets.append( curcol )
                    table_rows, col_rows = [],[]
                    break
                else:
                    continue

        # Moved to be outside of DIM table creation script
        alter_sql = ''
        args.currentorder = currentorder

        # If the workbook contains a DML tab, capture it to append to the CONFIG block
        dml_block, dml_sql = '',''
        if 'DML' in wb.sheetnames:
            dml_tab = wb['DML']
            #replace all newlines in the cell value with a space
            # save the contents of the first column of the dml tab to a string
            try:
                for row in dml_tab.rows:
                    if row[0].value is not None:
                        dml_sql += row[0].value.replace('\n', ' ')
                # dml_sql = '\n'.join([' '.join( row[0].value.replace('\n',' ') ) for row in dml_tab.rows])
                dml_block = f'\n{dml_sql}'
            except Exception as e:
                print(f'\t\tERROR: {e}')
                continue

        # print (f"<< Testing for {eachval} Table for >> {tgt_table}'")
        if not curtable in missing_sheets:
            table_rows = [row for row in sheet2dict(wb[ curtable ])]
            col_rows = [row for row in sheet2dict(wb[ curcol ])]

            # Only check on SCD for the DIM tables/columns
            is_scd=False
            if currentorder == 'DIM':
                if not args.silent:
                    print( f'\n\n~~Current tab/cols are: {curtable}; {curcol}')         # DEBUG

                    missing_cols =[]        
                    for each_col in expected_cols:
                        if each_col in col_rows[0].keys(): 
                            continue 
                        else:
                            missing_cols.append( each_col )
                    if len(missing_cols)>0:
                        missing_cols =  list(set( missing_cols ))
                        print( f'\n\n~~~ There are missing colums for the current table: {missing_cols}')

                for row in col_rows:
                #print(row)
                    if row.get( 'SCD' ): is_scd = True; break
                    # print('NOT DIM MAPPING, skipping',xls)
                    break
                if not is_scd: continue
            
            # replaced tgt_table with curtable < If this fails
                scd_dict = get_scdinfo( args.schema, tgt_table, col_rows)
                scd2_sql, scd6_sql = make_scd( scd_dict )

                # Workbook name is Case Sensitive.  Need to correct
                if 'Alter' in wb:
                    alter_sql = '\n'
                    for sql in sheet2list(wb['Alter']):
                        if not sql[0]:
                            continue
                        alter_sql += sql[0] + '\n'
                    alter_sql = alter_sql.replace('\\xa0', ' ')

                # Put config in front of scd6_sql
                config = get_config( alter_sql, tgt_table, dml_block )
                scd6_sql = config + scd6_sql

                # write_dbt: replaced scd_dict['table'] w/ curtable
                if not args.yml:
                    write_scd_step1( scd_dict[ 'table' ], args.modeldir, scd2_sql, args.modeldir, scd6_sql )
                    write_final_scd_sql( args, tgt_table, scd_dict )

            # tgt_table = xls.stem
            print(f'Tgt table: {tgt_table} ')  # Should be the XLS filename

            # Start to process the build of each Layer
            # fix table_rows reference
        build(args, args.modeldir, tgt_table, table_rows, col_rows, alter_sql, dml_block)


def process_std_map( args, xls ):
    # If the workbook contains a DML tab, capture it to append to the CONFIG block
    tgt_table = xls.stem
    print( '-'*30 )
    print( f'Processing Standard Mapping for {tgt_table}' )
    missing_sheets =[]
    expected_cols = [ 'Source Schema', 'Source Table', 'Source Column', 'Datatype', 'Automated Logic', 'Manual Logic', 'Order#' ,'Staging Layer Column Name' ,'Staging Layer Datatype', 'CPI' ,'Unique' ,'Not Null' ,'Test Expression' ,'Accepted Values', 'SCD' ]

##############
# Standard file processing...
    wb = openpyxl.load_workbook(xls)

    print(f'Tgt table: {tgt_table} ')  # Should be the XLS filename
    alter_sql, config, dml_block, dml_sql = '','','',''

    if 'DML' in wb.sheetnames:
        dml_tab = wb['DML']
        #replace all newlines in the cell value with a space
        # save the contents of the first column of the dml tab to a string
        try:
            for row in dml_tab.rows:
                if row[0].value is not None:
                    dml_sql += row[0].value.replace('\n', ' ')
            # dml_sql = '\n'.join([' '.join( row[0].value.replace('\n',' ') ) for row in dml_tab.rows])
            dml_block = f'\n{dml_sql}'
        except Exception as e:
            print(f'\t\tERROR: {e}')

    # args.currentorder = {tgt_table}.split()

    table_rows = [row for row in sheet2dict(wb['Tables'])]
    col_rows = [row for row in sheet2dict(wb['Columns'])]
    missing_cols =[]

    for each_col in expected_cols:
        if each_col in col_rows[0].keys(): 
            continue 
        else:
            missing_cols.append( each_col )
    if len(missing_cols)>0:
        missing_cols =  list(set( missing_cols ))
        print( f'\n\n~~~ There are missing colums for the current table: {missing_cols}')

# Workbook name is Case Sensitive.  Need to correct
# Build the CONFIG block
    if 'Alter' in wb:
        alter_sql = '\n'
        for sql in sheet2list(wb['Alter']):
            if not sql[0]:
                continue
            alter_sql += sql[0] + '\n'
        alter_sql = alter_sql.replace('\\xa0', ' ')
        config = get_config( alter_sql, tgt_table, dml_block )

    # Only check on SCD for the DIM tables/columns
    if xls.name.startswith('DIM'):
        is_scd = False
        for row in col_rows:
            #print(row)
            if row.get( 'SCD' ): 
                is_scd = True; break
            # print('NOT DIM MAPPING, skipping',xls)

        # replaced tgt_table with curtable < If this fails
        # Put config in front of scd6_sql
        scd_dict = get_scdinfo( args.schema, tgt_table, col_rows)
        
        if is_scd:
            scd2_sql, scd6_sql = make_scd( scd_dict )
            scd6_sql = config + scd6_sql
        else:
            scd6_sql = config

        # write_dbt: replaced scd_dict['table'] w/ curtable
        if not args.yml:
            write_scd_step1( scd_dict[ 'table' ], args.modeldir, scd2_sql, args.modeldir, scd6_sql )
            write_final_scd_sql( args, tgt_table, scd_dict)

    # Start to process the build of each Layer
    build(args, args.modeldir, tgt_table, table_rows, col_rows, alter_sql, dml_block)



def main():
    args = process_args()
    setup(args)
    count = 0
    plural, scd6_sql, dml_block, dml_sql  = '', '', '',''

    print('\n' + '-'*30)

    #Process each xls file in the directory
    for xls in args.mapdir.iterdir():
        print(f'\t<< Processing: {xls}')
        if xls.name.startswith('XDIM_'):
            print(f'\t\tNOT STAGE MAPPING -- skipping: {xls.name}')
            continue
        elif 'xls' not in xls.suffix.lower():
            print(f'\t-- Skipping non-XLS file: {xls}')
            continue

        print(f'-'*30)
        print(f"Currently processing {os.path.basename(xls)}")

        # Determine if it is UNIFIED or not
        if 'UNIFIED' in xls.name.upper():
            print(f'\t-- UNIFIED MAPPING: {xls.name}')
            process_unified_map( args, xls )
        else:
            print(f'\t-- Standard MAPPING: {xls.name}')
            process_std_map( args, xls )
        
        count += 1

    if count > 1: plural = 's.'
    else: plural = '.'

    print(f'Processed: {count} file{plural}')
    if count > 0 : print(f'Models written to: {args.modeldir}')

    input('[ Press Enter to exit ] ')


if __name__ == '__main__':
    main()
