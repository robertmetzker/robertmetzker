import os, argparse
from pathlib import Path

# Other necessary libs
import openpyxl, yaml               # pip install openpyxl (pyyaml)
from openpyxl.utils import get_column_letter
from operator import itemgetter

from io import StringIO
import argparse
import sys, traceback

# Global variables
DEBUG = False

# Functions
def process_args():
    parser = argparse.ArgumentParser(description='Command Line Arguments')
    parser.add_argument('--rootdir', default=Path.cwd(),
                        help='The Working Directory for maps and Models (defaults to Current Directory)')
    # boolean
    parser.add_argument( '--yml', action='store_true', help='If flagged as tests, will only generate yml files')
    parser.add_argument( '--silent', action='store_true', help='If flagged will only print file names')
    parser.add_argument( '--debug', action='store_true', help='If flagged will add verbosity to output')

    args = parser.parse_args()
    args.schema = 'DEV_EDW'

    if args.rootdir:
        args.rootdir = Path(args.rootdir)

    global DEBUG
    DEBUG = args.debug

    return args


def setup(args):
    #use global args to store the rootdir and subdirs
    dir_list = {'mapdir':'mappings', 'modeldir':'models', 'snapdir':'snapshots', 'archivedir':'models_archive'}

    for dir in dir_list.items():
        setattr(args, dir[0], args.rootdir / dir[1])
        if not getattr(args, dir[0]).exists(): getattr(args, dir[0]).mkdir()

    print( '--Setting up directories--')
    for dir in dir_list:
        print( f'{dir} directory: {getattr(args, dir)}' )
    print( f'{"-"*30}' )

    print( 'Mapping docs (XLS) in:', args.mapdir )
    print( 'Output (SQL, YML) going to:', args.modeldir )

    # Show existing XLS files in Mapdir
    for afile in args.mapdir.glob('*.xlsx'):
        print(f'\tMapping found: {afile}')


def sheet2dict(sheet):
    print(f'\t<<<< Processing {sheet}')

    fields = None
    for rownum, row in enumerate( sheet.iter_rows( values_only = True) ):
        if rownum == 0:
            fields = row
            continue
        row = dict(zip( fields, row ))

        yield row

def sheet2list(sheet):
    for row in sheet.iter_rows( values_only = True ):
        if not row:
            continue
        yield row

def get_derived_names(wb):
    derived_names = {}
    index_sheet = wb['Index']
    
    for row_index, row in enumerate(index_sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] and isinstance(row[0], str):
            cell = index_sheet.cell(row=row_index, column=1)
            if cell.hyperlink:
                tab_name = cell.hyperlink.location.split('!')[0].replace("'", "")
                # derived_name = row[0].lower().rstrip('tables').strip()
                derived_name = cell.hyperlink.display.lower().rstrip('tables').strip()
                derived_names[tab_name] = derived_name
    
    return derived_names

def get_output_name(tab_name, derived_names):
    return derived_names.get(tab_name, tab_name)

def debug_print(*args, **kwargs):
    if DEBUG:
        print(*args, **kwargs)

def write_snap_step1( table, snapdir, scd2_sql, modeldir, scd6_sql):
    snapfile = snapdir/(table+'_SNAPSHOT_STEP1.sql')
    print('\t-- Writing to', snapfile)
    with open( snapfile, 'w') as fw:
        fw.write( scd2_sql )

    #table_split,split=table.split('_STEP1')
    if table.startswith('DIM'):
        model_file = modeldir/(table+'.sql')
    else:
        model_file = modeldir/(table+'_alt.sql')

    print('\t   Writing to', model_file)
    with open( model_file, 'w' ) as fw:
        fw.write( scd6_sql )
    # print( '========== done' )


def write_scd_step2( args, tgt_table, scd_dict ):
    modeldir = args.modeldir
    
    scd1_cols = scd_dict[ 'scd1' ] 
    exclude_col=[ 'UNIQUE_ID_KEY', 'CURRENT_INDICATOR', 'EFFECTIVE_DATE', 'END_DATE', 'LOAD_DATETIME', 'UPDATE_DATETIME', 'EFFECTIVE_TIMESTAMP', 'END_TIMESTAMP',
    'LOAD_BATCH_ID', 'UPDATE_BATCH_ID', 'PRIMARY_SOURCE_SYSTEM' ]

    scd1_cols = [ col.upper() for col in scd1_cols ]
    exclude_col = [ col.upper() for col in exclude_col ]

    # determine which columns are in scd1 but not in exclude_col without using a set
    scd1_remaining_cols = [ col for col in scd1_cols if col not in exclude_col ]

    scd1_str = ', '.join( scd1_remaining_cols )
    keycol = scd_dict[ 'KEY' ]
    
    src_table = tgt_table.replace( 'DIM_', 'DSV_' )

    # prefix = ''' {{ config(schema = 'EDW_STAGING', tags=["dim"] ) }}\n ----SRC LAYER----\nWITH\n'''
    prefix = '''\n ----SRC LAYER----\nWITH\n'''

    sql1 = prefix+f'''SCD1 as ( SELECT {scd1_str} , {keycol}    \n\t--, '1901-01-01' as DBT_VALID_FROM, '2099-12-31' as DBT_VALID_TO\n\tfrom      {{{{ ref( '{src_table.lower()}') }}}} )'''
    sql2 = f'''SCD2 as ( SELECT *    \n\tFROM      {{{{ ref('{tgt_table}_SNAPSHOT_STEP1') }}}} )'''
    sql3 = f'''FINAL as ( SELECT * 
            FROM  SCD2 
                INNER JOIN SCD1 USING( {keycol} )  )\nselect * from FINAL'''

    all_sql_str = ',\n'.join([ sql1, sql2, sql3 ])

    final_file = modeldir/(tgt_table+'_SCDALL_STEP2.sql')

    print( '\t   Writing to', final_file )
    with open( final_file, 'w' ) as fw:
        if not args.yml: fw.write( all_sql_str )  
    # print('========== done')


def process_std_map(args, xls):
    src_file = xls.name.split('.')[0]
    schemas = {'HUB':'RAW_VAULT', 'SAT':'RAW_VAULT', 'LNK':'RAW_VAULT', 'REF':'RAW_VAULT', 'BASE':'STAGING', 'STG':'STAGING', 'PIT':'BUSINESS_VAULT', 'BRIDGE':'BUSINESS_VAULT'}

    wb = openpyxl.load_workbook(xls)
    found_sheets = wb.sheetnames

    # Group sheets by layer type
    layer_groups = {}
    for sheet in found_sheets:
        parts = sheet.split()
        if len(parts) >= 2:
            layer = parts[0]
            if layer not in layer_groups:
                layer_groups[layer] = []
            layer_groups[layer].append(sheet)

    # Print grouped sheets
    print(f"Sheets found in {xls.name}:")
    for layer, sheets in layer_groups.items():
        print(f"\n{layer} Layers:")
        for sheet in sheets:
            print(f"  - {sheet}")

    # Process each base layer
    for layer, sheets in layer_groups.items():
        for sheet in sheets:
            parts = sheet.split()
            table_name = ' '.join(parts[1:-1])
            sheet_type = parts[-1]
            
            if sheet_type == 'Tables':
                print(f'\t\t-- Processing {layer} {table_name} in: {xls.name}')
                tables_data = process_tables(args, wb, sheet)
                columns_sheet = f"{layer} {table_name} Columns"
                if columns_sheet in found_sheets:
                    columns_data = process_columns(args, wb, columns_sheet)

                    print(f'\t\t-- Validating Table and Column aliases for {layer} {table_name}')
                    is_valid = validate_tables_columns(xls, tables_data, columns_data, layer, table_name)

                    if not is_valid:
                        print(f'\t## ERROR: Table and Column aliases are not valid in: {xls.name} for {layer} {table_name}\n\n')
                        continue

                    print(f'\n{layer} table: {layer}_{table_name}')

                    alter_sql = ''
                    alter_sheet = f"{layer} {table_name} Alter"
                    if alter_sheet in found_sheets:
                        alter_sql = process_alter(args, wb, alter_sheet)

                    # Generate SCD dictionary
                    scd_dict = get_scdinfo(schemas[layer], f"{layer}_{table_name}", columns_data)

                    # Generate config block
                    config = get_config(f"{layer}_{table_name}", alter_sql=alter_sql, scd_dict=scd_dict)

                    # build(args, args.modeldir, f"{layer}_{table_name}", tables_data, columns_data, wb, alter_sql=alter_sql, config_sql=config)
                    build(args, args.modeldir, f"{layer}_{table_name}", tables_data, columns_data, wb, alter_sql=alter_sql, scd_dict=config)

    # Process DML if present
    if 'DML' in found_sheets:
        dml_block = process_dml(args, wb, 'DML')

    return 0


def process_unified(args, xls):
    print(f'\t-- UNIFIED MAPPING: {xls.name}')
    src_file = xls.name.split('.')[0]
    layers = {'HUB', 'SAT', 'LNK', 'REF', 'BASE', 'STG', 'PIT', 'BRIDGE'}
    schemas = {'HUB':'RAW_VAULT', 'SAT':'RAW_VAULT', 'LNK':'RAW_VAULT', 'REF':'RAW_VAULT', 'BASE':'STAGING', 'STG':'STAGING', 'PIT':'BUSINESS_VAULT', 'BRIDGE':'BUSINESS_VAULT'}

    wb = openpyxl.load_workbook(xls)
    derived_names = get_derived_names(wb)
    found_sheets = wb.sheetnames

    # Group sheets by layer type and table name
    layer_groups = {}
    for sheet in found_sheets:
        parts = sheet.split()
        if len(parts) >= 3:  # Layer, Table Name, Sheet Type
            layer = parts[0]
            sheet_type = parts[-1]
            table_name = ' '.join(parts[1:-1])
            
            if layer not in layer_groups:
                layer_groups[layer] = {}
            if table_name not in layer_groups[layer]:
                layer_groups[layer][table_name] = {'Tables': None, 'Columns': None, 'Alter': None}
            
            layer_groups[layer][table_name][sheet_type] = sheet
        else:
            if 'Other' not in layer_groups:
                layer_groups['Other'] = {}
            layer_groups['Other'][sheet] = {'sheet': sheet}

    # Print grouped sheets
    print(f"Sheets found in {xls.name}:")
    for layer, tables in layer_groups.items():
        print(f"\n{layer} Layers:")
        for table, sheets in tables.items():
            for sheet_type, sheet in sheets.items():
                if sheet:
                    print(f"  - {sheet}")

    print("\nProcessing layers:")
    for this_layer in layers:
        if this_layer not in layer_groups:
            print(f'\t-- No sheets found for {this_layer} layer. Skipping.')
            continue

        print(f'\t-- Processing {this_layer} Layer of UNIFIED mapping: {xls.name}')

        for table_name, sheets in layer_groups[this_layer].items():
            tables_tab = sheets.get('Tables')
            columns_tab = sheets.get('Columns')
            alter_tab = sheets.get('Alter')

            if not (tables_tab and columns_tab):
                print(f'\t\t-- Skipping {this_layer} {table_name} due to missing Tables or Columns sheet.')
                continue

            print(f'\t\t-- Processing {this_layer} {table_name} in: {xls.name}')
            
            tables = process_tables(args, wb, tables_tab)
            columns = process_columns(args, wb, columns_tab)

            print(f'\t\t-- Validating Table and Column aliases for {this_layer} {table_name}')
            is_valid = validate_tables_columns(xls, tables, columns, this_layer, table_name)

            if not is_valid:
                print(f'\t## WARNING: Table and Column aliases are not valid in: {xls.name} for {this_layer} {table_name}. Continuing processing...\n')

            alter_sql = ''
            if alter_tab:
                alter_sql = process_alter(args, wb, alter_tab)

            # Generate SCD dictionary
            scd_dict = get_scdinfo(schemas[this_layer], table_name, columns)

            # Use the derived name for the output file
            output_name = get_output_name(f"{this_layer} {table_name}", derived_names)
            full_table_name = output_name.replace(' ', '_')
            print(f'{this_layer} table: {full_table_name}')
            
            # Call build with scd_dict instead of config_sql
            build(args, args.modeldir, full_table_name, tables, columns, wb, alter_sql=alter_sql, scd_dict=scd_dict)
            print(f'\t\t-- Finished processing {this_layer} layer for {full_table_name}\n\n')

    # Process additional sheets if necessary
    if 'Other' in layer_groups:
        for sheet, sheet_info in layer_groups['Other'].items():
            if sheet == 'DML':
                print(f'\t-- Processing DML sheet in: {xls.name}')
                dml_block = process_dml(args, wb, 'DML')
                # You might want to do something with dml_block here

    return 0


def build_scd2( schema, table, cols, keycol ):
    schema = 'EDW_STAGING_SNAPSHOT'
    all_cols = cols + [ keycol ]
    
    
    cols_str = ', '.join( all_cols )
    src_table = table.replace( 'DIM_' , 'DSV_' )
    snapshot_sql=f'''
    
    {{%% snapshot %s_SNAPSHOT_STEP1 %%}}

    {{{{
        config(
        target_schema='%s',
        unique_key='%s',
        strategy='check',
        check_cols={cols},
        )
    }}}}

    select {cols_str} 
        from  {{{{ref('%s') }}}}

    {{%% endsnapshot %%}}
    '''%(table,schema,keycol,src_table)

    return snapshot_sql

def build_scd6 ( schema, table, keycol, scd1_cols = [], scd2_cols = [], scd3_cols = [], scd0_cols = [] ):
    keycol = keycol.upper()
    if not keycol: keycol = 'UNIQUE_ID_KEY'

    exclude_col = ['UNIQUE_ID_KEY', 'CURRENT_INDICATOR', 'EFFECTIVE_DATE', 'END_DATE', 'LOAD_DATETIME', 
    'UPDATE_DATETIME', 'EFFECTIVE_TIMESTAMP', 'END_TIMESTAMP',
    'LOAD_BATCH_ID', 'UPDATE_BATCH_ID', 'PRIMARY_SOURCE_SYSTEM']

    start_sql = f'''\n\n WITH  SCD as ( \n\tSELECT  {keycol}'''

    sql2 = []
    for col in scd2_cols:
        if col in exclude_col: continue
        sql = f'''\n {col}'''
        sql2.append( sql )

    sql3 = []
    for col in scd3_cols:
        if col in exclude_col: continue
        sql = f'''\n\t, lag({col}) over  
           (partition by {keycol} 
             order by dbt_updated_at
        ) as {col}'''
        sql3.append(sql)

    sql1 = []
    for col in scd1_cols:
        if col in exclude_col: continue
        sql=f'''\n\t, last_value({col}) over 
            (partition by {keycol} 
              order by dbt_updated_at 
                ROWS BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING
            ) as {col}'''
        sql1.append(sql)
    
    sql0 = []
    for col in scd0_cols:
        if col in exclude_col: continue
        sql=f'''\n\t, first_value({col}) over 
            (partition by {keycol} 
              order by dbt_updated_at 
                ROWS BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING
            ) as {col}'''
        sql0.append( sql )

    end_sql = f'''\n\t    , DBT_VALID_FROM as EFFECTIVE_TIMESTAMP
    \t, DBT_VALID_TO   as END_TIMESTAMP
    \t,
    \t  CASE WHEN CAST(DBT_VALID_FROM as DATE) = '1901-01-01' then cast( DBT_VALID_FROM as DATE )
    \t    WHEN cast( DBT_VALID_FROM as DATE ) <> '1901-01-01' THEN dateadd(day,1,CAST(DBT_VALID_FROM as DATE))
    \t  else cast( DBT_VALID_FROM as DATE ) end as EFFECTIVE_DATE
    \t, cast( DBT_VALID_TO as DATE ) as END_DATE
    \t, CASE WHEN DBT_VALID_TO IS NULL THEN 'Y' ELSE 'N' END as CURRENT_INDICATOR'''
    #CAST(DBT_VALID_FROM AS DATE) as EFFECTIVE_DATE\n,
    sql1[ 0 ] = start_sql + sql1[ 0 ]
    scd_all = sql1 + sql2 + sql3 + sql0
    #+ end_sql

    scd_str = ' '.join( scd_all )
    scd_str = scd_str.strip(',')

    scd_str = scd_str + end_sql + '\n'+"     \n\tFROM {{ ref('" + table.lower() + "_SCDALL_STEP2') }})\n\n"
    scd_str += 'select * from SCD\n'

    return scd_str

def get_scdinfo( schema, table, cols_dict):
    '''
    {'Order#': None, 'SCD TYPE': None, 'STAGING LAYER COLUMN NAME': None, ' 'HASHDIFF': None, 
    'UNIQUE': None, 'NOT NULL': None, 'TEST EXPRESSION': None, 'ACCEPTED VALUES': None}
    '''
    scd_dict = {
        'table': table,
        'schema': schema,
        'KEY':'',
        'scd1':[], 'scd2':[], 'scd3':[], 'scd0':[],
    }

    keycol = 'KEY'
    for row in cols_dict.values():
        if '1' in str(row.get( 'SCD' )).split(','):
            scd_dict[ 'scd1' ].append( row.get( 'STAGING LAYER COLUMN NAME' ))
        if '2' in str(row.get( 'SCD' )).split(','):
            scd_dict[ 'scd2' ].append( row.get( 'STAGING LAYER COLUMN NAME' ))
        if '3' in str(row.get( 'SCD' )).split(','):
            scd_dict[ 'scd3' ].append( row.get( 'STAGING LAYER COLUMN NAME' ))
        if '0' in str(row.get( 'SCD' )).split(','):
            scd_dict[ 'scd0' ].append( row.get( 'STAGING LAYER COLUMN NAME' ))
        if row.get( 'SCD' ) == 'KEY':
            scd_dict[ 'KEY' ] = row.get( 'STAGING LAYER COLUMN NAME' ).upper() 
            
    return scd_dict


def process_tables(args, wb, tables_tab):
    print(f'\t\t-- Processing {tables_tab} ...')
    expected_cols = ['SOURCE SCHEMA', 'SOURCE TABLE', 'ALIAS', 'FILTER CONDITIONS', 'FILTER RESTRICTION RULE','PARENT JOIN NUMBER', 'PARENT TABLE JOIN', 'CHILD TABLE JOIN','FINAL LAYER FILTER', 'JOIN TYPE']

    ws = wb[tables_tab]
    tables_dict = {}

    # Get derived names
    derived_names = get_derived_names(wb)

    # Convert found column names to uppercase
    found_cols = [cell.value.upper() if cell.value else None for cell in ws[1]]
    missing_cols = list(set(expected_cols) - set(found_cols))
    extra_cols = list(set(found_cols) - set(expected_cols))

    if extra_cols:
        debug_print(f'\t\t\t-- Found Extra columns: {extra_cols}')
    if missing_cols:
        debug_print(f'\t\t\t-- Missing columns: {missing_cols}')

    for row in ws.iter_rows(min_row=2, values_only=True):
        if 'SOURCE TABLE' in found_cols and row[found_cols.index('SOURCE TABLE')] is not None:
            source_table = row[found_cols.index('SOURCE TABLE')]
            tables_dict[source_table] = {}
            for i, col in enumerate(found_cols):
                if col:  # Only process non-None column names
                    tables_dict[source_table][col] = row[i] if i < len(row) else ''
            
            # Add derived name to the table dictionary
            derived_name = derived_names.get(tables_tab, tables_tab)
            tables_dict[source_table]['DERIVED_NAME'] = derived_name

    # Debug: Print processed tables
    debug_print(f"DEBUG: Processed tables: {tables_dict}")

    return tables_dict


def process_columns(args, wb, columns_tab):
    print(f'\t\t-- Processing {columns_tab} ...')
    
    # read the tables tab and convert to a dictionary
    expected_cols = ['SOURCE SCHEMA', 'SOURCE TABLE', 'SOURCE COLUMN', 'DATATYPE', 'AUTOMATED LOGIC', 'MANUAL LOGIC' , 'ORDER#', 'STAGING LAYER COLUMN NAME', 'STAGING LAYER DATATYPE', 'HASHDIFF', 'UNIQUE','SET DEFAULT', 'NOT NULL', 'REMOVE COLUMN', 'TEST EXPRESSION', 'ACCEPTED VALUES', 'RELATIONSHIP', 'MAPPING NOTES', 'PK','SCD']
    dim_cols = ['MODEL EQUALITY', 'MODEL EQUAL ROWCOUNT', 'SCD']
    dbt_cols_list = ['DBT_SCD_ID', 'DBT_VALID_FROM', 'DBT_VALID_TO', 'DBT_UPDATED_AT']
    cur_layer = columns_tab.split(' ')[0]

    if columns_tab.startswith('DIM'):
        expected_cols = expected_cols + dim_cols

    ws = wb[columns_tab]
    columns_dict = {}

    # Check for missing columns and convert to a list
    # missing_cols=[x for x in expected_cols if x not in found_cols]
    # extra_cols=[x for x in found_cols if x not in expected_cols]

    # Convert found column names to uppercase
    found_cols = [cell.value.upper() if cell.value else None for cell in ws[1]]
    
    missing_cols = list(set(expected_cols) - set(found_cols))
    extra_cols = list(set(found_cols) - set(expected_cols))

    if extra_cols:
        print(f'\t\t\t-- Found Extra columns: {extra_cols}')
    if missing_cols:
        print(f'\t\t\t-- Missing columns: {missing_cols}')
        
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        staging_col_name = row[found_cols.index('STAGING LAYER COLUMN NAME')]
        if staging_col_name is not None and staging_col_name not in dbt_cols_list:
            columns_dict[idx] = {}
            for i, col in enumerate(found_cols):
                if col:  # Only process non-None column names
                    columns_dict[idx][col] = row[i] if i < len(row) else ''

    # Debug: Print processed columns
    debug_print(f"DEBUG: Processed columns: {columns_dict}")

    return columns_dict


def validate_tables_columns(xls, tables, columns, current_layer, table_name):
    is_valid = True
    
    if current_layer == 'STG':
        # For STG layer, we don't validate aliases
        return True

    # Create a mapping of Source Table to Alias
    table_alias_mapping = {table['SOURCE TABLE']: table['ALIAS'] for table in tables.values()}

    # Get the set of aliases used in the Tables sheet
    table_aliases = set(table_alias_mapping.values())

    # Get the set of aliases used in the Columns sheet
    column_aliases = set(col['SOURCE TABLE'] for col in columns.values() if col.get('SOURCE TABLE') is not None)

    missing_tables = list(column_aliases - table_aliases)
    missing_columns = list(table_aliases - column_aliases)

    if missing_tables:
        print(f'\t\t\t-- Missing Table references on Columns Tab for {current_layer} {table_name}: {missing_tables}')
        is_valid = False
    if missing_columns:
        print(f'\t\t\t-- Missing Columns alias in Tables tab for {current_layer} {table_name}: {missing_columns}')
        is_valid = False

    return is_valid


def process_alter( args, wb, alter_tab ):
    print(f'\t\t-- Processing {alter_tab} ...')

    ws = wb[alter_tab]
    alter_sql = []

    # read each row of the alter_tab and append to the alter_sql list if it's not blank
    for row in ws.iter_rows(min_row=0, values_only=True):
        if row[0]:
            if row[0].strip() != '':
                alter_sql.append(row[0])
    
    return alter_sql

def process_dml( args, wb, dml_tab ):
    print(f'\t\t-- Processing DML Statements ...' )

    ws = wb[dml_tab]
    dml_sql = []

    # read the dml_tab and append each row to the dml_sql list
    for row in ws.iter_rows(min_row=0, values_only=True):
        if row != '':
            dml_sql.append(row[0])

    return dml_sql

def get_scdinfo(schema, table, cols_dict):
    scd_dict = {
        'table': table,
        'schema': schema,
        'KEY': [],
        'scd1': [], 'scd2': [], 'scd3': [], 'scd0': [],
        'config': {
            'target_schema': schema,
            'unique_key': None,
            'strategy': None,
            'updated_at': None,
            'check_cols': None
        }
    }

    check_cols = []
    for row in cols_dict.values():
        scd_value = str(row.get('SCD', '')).lower().strip()
        column_name = row.get('STAGING LAYER COLUMN NAME')

        if 'key' in scd_value:
            scd_dict['KEY'].append(column_name)
            if scd_dict['config']['unique_key'] is None:
                scd_dict['config']['unique_key'] = []
            scd_dict['config']['unique_key'].append(column_name)
        elif 'timestamp' in scd_value:
            scd_dict['config']['strategy'] = 'timestamp'
            scd_dict['config']['updated_at'] = column_name
        elif 'check:' in scd_value:
            scd_dict['config']['strategy'] = 'check'
            if 'all' in scd_value.replace(' ', ''):
                scd_dict['config']['check_cols'] = 'all'
            else:
                check_cols.append(column_name)

        # Existing SCD type assignments
        if '1' in scd_value:
            scd_dict['scd1'].append(column_name)
        if '2' in scd_value:
            scd_dict['scd2'].append(column_name)
        if '3' in scd_value:
            scd_dict['scd3'].append(column_name)
        if '0' in scd_value:
            scd_dict['scd0'].append(column_name)

    # If check_cols is not 'all', set it to the list of check columns
    if scd_dict['config']['check_cols'] != 'all' and check_cols:
        scd_dict['config']['check_cols'] = check_cols

    # If there's only one key, convert it from a list to a string
    if len(scd_dict['KEY']) == 1:
        scd_dict['KEY'] = scd_dict['KEY'][0]
        scd_dict['config']['unique_key'] = scd_dict['config']['unique_key'][0]

    return scd_dict

def generate_dbt_snapshot_config(scd_dict):
    """
    Generates a dictionary containing dbt snapshot configuration based on SCD information.
    """
    config = scd_dict.get('config', {})
    snapshot_config = {
        'target_schema': config.get('target_schema'),
        'unique_key': config.get('unique_key'),
        'strategy': config.get('strategy'),
    }

    if config.get('strategy') == 'timestamp':
        snapshot_config['updated_at'] = config.get('updated_at')
    elif config.get('strategy') == 'check':
        snapshot_config['check_cols'] = config.get('check_cols')

    return snapshot_config

def get_config(cur_layer, alter_sql='', scd_dict=None):
    """
    Builds the CONFIG dynamically based on the layer, alter SQL, and SCD information
    """
    config = {}

    # Basic config setup
    if cur_layer.startswith(('SAT', 'HUB', 'LNK')):
        config['tags'] = ['raw_vault']

    if alter_sql:
        alter_sql_str = '\n'.join(alter_sql) if isinstance(alter_sql, list) else alter_sql
        config['post_hook'] = f'("{alter_sql_str}")'

    # Incorporate snapshot config if SCD info is provided
    if scd_dict:
        snapshot_config = generate_dbt_snapshot_config(scd_dict)
        config.update(snapshot_config)

    # Generate the config string
    config_items = []
    for key, value in config.items():
        if isinstance(value, list):
            config_items.append(f"{key} = {value}")
        elif isinstance(value, str):
            config_items.append(f"{key} = '{value}'")
        else:
            config_items.append(f"{key} = {value}")

    if config_items:
        config_str = "{{ config(\n    " + ",\n    ".join(config_items) + "\n) }}"
    else:
        config_str = ""

    return config_str

# Example usage:
# scd_dict = get_scdinfo(schema, table, cols_dict)
# config_sql = get_config(cur_layer, alter_sql, scd_dict)


def get_model_tests(row):
    '''
    https://github.com/fishtown-analytics/dbt-utils
    '''
    test_dict = {}
    test_cols = ['UNIQUE', 'MODEL EQUALITY', 'MODEL EQUAL ROWCOUNT', 'TEST EXPRESSION', 'PK' ]
    # Globally replacing the relationship with relationship_where

    tests = []
    add = {'severity': 'warn'}

    for atest in test_cols:
        if atest not in row:
            continue
        if not row[atest]:
            continue

        if atest == 'UNIQUE':
            if ':' in str(row[atest]):
                test = str(row[atest]).replace('COMPOSITE:', '').strip()
                test = test.replace("unique:", "").replace("Unique:", "").strip()
                row[atest] = {
                    'dbt_utils.unique_combination_of_columns': {
                        'combination_of_columns': [each.strip() for each in test.split(',')],
                        'severity': 'warn'
                    }
                }
            else:
                continue

        if atest == 'PK':
            if ':' in str(row[atest]):
                test = str(row[atest]).replace('PK:', '').strip()
                row[atest] = {
                    'dbt_constraints.primary_key': {
                        'column_names': [each.strip() for each in test.split(',')],
                    }
                }
            else:
                continue

        if atest == 'MODEL EQUALITY':
            # print(f' $$$$$$$$ FOUND Model Equality ->  {row[ atest ]}')               # DEBUG
            test =   str(row[ atest ]).strip()
            row[atest] = {
                'dbt_utils.equality': 
                {
                    'compare_model': f"ref( {row[ 'STAGING LAYER COLUMN NAME' ]} )",
                    'compare_columns': [each.strip() for each in test.split(',')]                    
                    # 'compare_columns': [ f'ref({row[ atest ]!r})' ]
                }
            }

        if atest == 'MODEL EQUAL ROWCOUNT':
            # print(f' $$$$$$$$ FOUND Model Equal Rowcount -=>  {row[ atest ]}')        # DEBUG
            row[atest] = {
                'dbt_utils.equal_rowcount': 
                {
                    'compare_model': f'ref({row[ atest ]!r})'
                }
            }

        # Relationship moved to column test...

        if atest == 'TEST EXPRESSION':
            if row.get('STAGING LAYER COLUMN NAME','x') in row.get('TEST EXPRESSION','y'):
                # If true, then should be a model test, else column test
                test =   str(row[ atest ]).strip()
                row[atest] = {
                    'dbt_utils.expression_is_true':
                    {
                        'severity': 'warn',
                        'expression': test
                    }
                }
            else: continue
        tests.append(row[atest])

    return tests

def get_column_tests(row, table_name):
    '''
    https://github.com/fishtown-analytics/dbt-utils
    '''
    test_dict = {}
    test_cols = [ 'UNIQUE', 'NOT NULL', 'TEST EXPRESSION', 'ACCEPTED VALUES', 'RELATIONSHIP', 'SET DEFAULT']
    # 'Equal Rowcount',  'Mutually Exclusive Ranges', 'Unique Combination of Columns', 'Accepted Range']

    tests = []
    add = {'severity': 'warn'}

    for atest in test_cols:
        if atest not in row:
            continue
        if not row[atest]:
            continue

        # If a FLF or FACT build, do not include the has_default test.  Will require knowing the layer during build.
        if atest == 'SET DEFAULT':
            test = str( row[ atest ] ).strip()
            row['DEFAULT']= test
            if table_name.startswith('F'):
                row[atest]=''
            else:
                row[atest]=''
                # row[atest] = {
                #     'has_default':
                #     {
                #         'severity': 'warn',
                #         'hashval': [each.strip() for each in test.split(',')]
                #     }
                # }

        if atest == 'UNIQUE':
            if ':' not in str(row[atest]):
                row[atest] = {
                    'unique':
                    {
                        'severity': 'warn'
                    }
                }
            else:
                continue

        if atest == 'NOT NULL':
            if 'where' in row[atest]:
                # TODO: Fix quoting around where
                test = str( row[ atest ] ).upper().replace( 'WHERE','').strip()
                row[atest] = {
                    'dbt_utils.not_null_where':
                    {
                        'where': test
                    }
                }                
            else:
                row[atest] = {
                    'not_null':
                    {
                        'severity': 'warn'
                    }
                }

        if atest == 'ACCEPTED VALUES':
            if 'not:' in row[atest]:
            # Split to break them into separate rows...
                test = str( row[ atest ] ).upper().replace( 'NOT:','').strip()
                row[atest] = {
                    'not_accepted_values':
                    {
                        'severity': 'warn',
                        'values': [each.strip() for each in test.split(',')]
                    }
                }
            else:
                test = str( row[ atest ] )
                row[atest] = {
                    'accepted_valids':
                    {
                        'severity': 'warn',
                        'values': [each.strip() for each in test.split(',')]
                    }
                }

        if atest == 'TEST EXPRESSION':
            # If the column name is not listed, then it can be a column test, else model test
            if not row.get('STAGING LAYER COLUMN NAME','x') in str(row[ atest ]).strip():
                test =   str(row[ atest ]).strip()
                row[atest] = {
                    'expression_is_true_ignore':
                    {
                        'ignore': '',
                        'severity': 'warn',
                        'expression': test
                    }
                }
            else: continue

        if atest == 'RELATIONSHIP':
            ignore_in = ''
            # Changed to a relationships_where instead to handle ignoring dummy value:  md5('-1111') 
            # Determine if the field is different (Based on TableName.ColumnName )
            if '.' in row[atest]:
                ref_table = row[ atest ].split('.')[0]
                ref_field = row[ atest ].split('.')[1]
                ref_ignore = row[ 'STAGING LAYER COLUMN NAME' ]
            else:
                ref_table = row[ atest ]
                ref_field = row[ 'STAGING LAYER COLUMN NAME' ]
                ref_ignore = row[ 'STAGING LAYER COLUMN NAME' ]

            if ref_field == 'DATE_KEY':
                ignore_in = " -1, -2, -3 "
            else: 
                ignore_list = row.get('SET DEFAULT')
                if ignore_list != None:
                    ignore_vals = str(ignore_list).split(',')
                    ignore_md5 = [f'{val}' for val in ignore_vals ]
                    ignore_in = ', '.join( ignore_md5 )
                else:
                    ignore_in = " '40c5dea533476acdd01f7ef0e84de22f', 'fcbcdcb8f6b1c597c5fdc7a54cd321ae' "
                    
                

            row[atest] = {
                'dbt_utils.relationships_where':
                {
                    'to': f'ref( {ref_table.lower()!r} )',
                    'field': f"{ref_field}"
                }
            }

        if not row[atest] == '':
            tests.append(row[atest])

    return tests


def get_source_sql(target_table, cols, tables):
    current_layer = target_table.split('_')[0]
    columns = cols.copy()

    for idx, row in enumerate(columns.values()):
        src_col = row.get('SOURCE COLUMN', '')
        manual_logic = row.get('MANUAL LOGIC', '')
        datatype = row.get('DATATYPE')
        automated_logic = row.get('AUTOMATED_LOGIC')
        staging_col_name = row.get('STAGING LAYER COLUMN NAME', '')

        if (current_layer == 'STG' or src_col == '(DERIVED)') and manual_logic:
            if 'HASH:' in manual_logic.upper() or 'COMPOSITE:' in manual_logic.upper():
                hashcol = manual_logic.replace('HASH:', '').replace('COMPOSITE:', '').replace('\n', '').replace(' ', '').strip()
                temp1 = [comp.strip() for comp in hashcol.split(',')]
                row['LOGIC_NAME'] = staging_col_name
                
                # Build the custom hash key with improved formatting
                hash_components = []
                for component in temp1:
                    hash_components.append(f"        IFNULL(NULLIF(UPPER(TRIM(CAST({component} as VARCHAR))), ''), '^^')")
                
                hash_concatenation = "\n        || '||' ||\n".join(hash_components)
                final_hash = f"        )) as BINARY(16))                                           "
                src_sql = f"""CAST(MD5_BINARY(CONCAT(
{hash_concatenation}
{final_hash}"""
            else:
                src_sql = manual_logic.strip()
        elif src_col == '(DERIVED)':
            if datatype == 'DATE':
                src_sql = f"""CASE WHEN {staging_col_name} is null then '-1' 
                                   WHEN {staging_col_name} < '1901-01-01' then '-2' 
                                   WHEN {staging_col_name} > '2099-12-31' then '-3' 
                                   ELSE regexp_replace({staging_col_name}, '[^0-9]+', '') 
                              END :: INTEGER"""
            else:
                src_sql = staging_col_name
            row['LOGIC_NAME'] = staging_col_name
        else:
            src_sql = src_col
            if automated_logic:
                automated_logic = automated_logic.lower()
                if 'trim' in automated_logic:
                    src_sql = f"TRIM({src_sql})"
                if 'upper' in automated_logic:
                    src_sql = f"UPPER({src_sql})"
            src_sql = f"NULLIF(TRIM({src_sql}), '')"
            row['LOGIC_NAME'] = src_col

        row['SQL'] = src_sql
        row['RENAME'] = staging_col_name
        row['IS_DERIVED'] = src_col == '(DERIVED)'

    return columns


def make_scd( schema, table_name, scd_dict):
    '''
    scd={
        'table':'STG_POLICYTEST',
        'schema':'X10057301',
        'scd1':['POLICY_ID','CLAIM_ID'],
        'scd2':['PTYPE'],
        'scd3':['OWNER_NAME','ADDR'],
        'keycol':'POLICY_ID'
    }
    '''

    keycol = 'KEY'

    #print(scd_dict)
    #cols=scd_dict['scd2']+scd_dict['scd3']+scd_dict['scd0']
    combined_scd_cols = scd_dict[ 'scd2' ] + scd_dict[ 'scd3' ] + scd_dict[ 'scd0' ]

    scd2_sql = build_scd2( schema, table_name,  combined_scd_cols, scd_dict[ keycol ])

    scd6_sql = build_scd6( schema, table_name,  scd_dict[ keycol ],
        scd1_cols = scd_dict[ 'scd1' ], scd2_cols = scd_dict[ 'scd2' ], scd3_cols = scd_dict[ 'scd3' ], scd0_cols = scd_dict[ 'scd0' ])

    return scd2_sql, scd6_sql


def make_yml(table_name, tables, columns):
    print(f'==== Building YAML for {table_name} ...')
    
    model_tests = []
    column_tests = {}

    # Get the derived name from the first table in the tables dictionary
    derived_name = next(iter(tables.values()))['DERIVED_NAME']
    output_name = derived_name.replace(' ', '_').lower()

    for row in columns.values():
        col_name = row.get('STAGING LAYER COLUMN NAME')
        if col_name:
            model_tests.extend(get_model_tests(row))
            column_tests[col_name] = get_column_tests(row, table_name)

    schema_dict = {
        'version': 2,
        'models': [
            {
                'name': output_name,  # Use the derived name here
                'tests': model_tests,
                'columns': []
            }
        ]
    }

    for col_name, tests in column_tests.items():
        if tests:
            column_entry = {
                'name': col_name,
                'tests': tests
            }
            schema_dict['models'][0]['columns'].append(column_entry)

    debug_print(f"Generated YAML for {output_name}:")
    debug_print(yaml.dump(schema_dict, sort_keys=False))

    return schema_dict


def format_select_cols(columns, layer):
    select_cols = []
    for idx, col in enumerate(columns.values()):
        col_name = col.get('STAGING LAYER COLUMN NAME')
        src_sql = col.get('SQL')
        
        if col_name:
            if layer == 'FINAL':
                if 'HASHDIFF' in src_sql:
                    pass
                else:
                    col_str = col_name
            else:
                if '\n' in src_sql:
                    src_sql = format_manual_logic(src_sql, ' ' * 12)
                col_str = f"{src_sql:<60} as {col_name:>50}"

            if idx == 0:
                select_cols.append(f"        {col_str}")
            else:
                select_cols.append(f"      , {col_str}")
    
    return '\n'.join(select_cols)


def format_manual_logic(logic, indent):
    if not logic or '\n' not in logic:
        return logic
    lines = logic.split('\n')
    return f"{lines[0]}\n" + '\n'.join([f"{indent}{line.strip()}" for line in lines[1:]])


def build(args, modeldir, table_name, tables, columns, wb, alter_sql='', scd_dict=None):
    print(f'\t\t-- Building {table_name} ...')
    yml = make_yml(table_name, tables, columns)
    all_sql = []

    try:
        current_layer = table_name.split('_')[0]
        columns = get_source_sql(table_name, columns, tables)

        # Use the derived name from the first table in the tables dictionary
        derived_name = next(iter(tables.values()))['DERIVED_NAME']
        output_name = derived_name.replace(' ', '_').lower()

        # Determine if it's a snapshot based on scd_dict
        is_snapshot = scd_dict and 'config' in scd_dict and scd_dict['config'].get('strategy') is not None

        # Add snapshot opening tag if it's a snapshot
        if is_snapshot:
            all_sql.append(f"{{% snapshot {output_name} %}}\n")

        # Config and SRC LAYER
        config_sql = get_config(output_name, alter_sql, scd_dict)
        all_sql.append(config_sql)
        if config_sql:
            all_sql.append('\n')  # Add a newline after the config block
        all_sql.append('---- SRC LAYER ----\nWITH')
        src_sql = build_src_layer(args, tables, current_layer)
        all_sql.append(src_sql)

        # LOGIC LAYER
        all_sql.append('---- LOGIC LAYER ----')
        logic_sql = build_logic_layer(columns, tables, current_layer)
        all_sql.append(logic_sql)

        # RENAME LAYER
        all_sql.append('---- RENAME LAYER ----')
        rename_sql = build_rename_layer(columns)
        all_sql.append(rename_sql)

        # FILTER LAYER
        all_sql.append('---- FILTER LAYER ----')
        filter_sql = build_filter_layer(tables)
        all_sql.append(filter_sql)

        # JOIN LAYER
        join_sql, base_join_alias = build_join_layer(tables, columns, current_layer)
        all_sql.append(join_sql)

        # FINAL LAYER
        final_sql = build_final_layer(tables, columns, current_layer, base_join_alias)
        all_sql.append(final_sql)

        # Add snapshot closing tag if it's a snapshot
        if is_snapshot:
            all_sql.append("\n{% endsnapshot %}")

        sql_str = '\n'.join(all_sql)

        # Write files using the derived name
        yml_file = modeldir / (output_name + '.yml')
        sql_file = modeldir / (output_name + '.sql')

        with open(yml_file, 'w') as fw:
            yaml.dump(yml, fw, sort_keys=False)
            print(f'\t\t-- Created YAML file: {yml_file}')

        with open(sql_file, 'w') as fw:
            fw.write(sql_str)
            print(f'\t\t-- Created SQL file: {sql_file}')

    except Exception as e:
        print(f'\t\t-- Error building {output_name}: {str(e)}')
        traceback.print_exc()

    print(f'\t\t-- Finished building {output_name}')


def build_src_layer(args, tables, current_layer):
    debug_print(f'==== Build SOURCE SQL for {current_layer} layer.')
    all_sql = []

    for row in tables.values():
        source_schema = row.get('SOURCE SCHEMA')
        source_table = row.get('SOURCE TABLE')
        table = row.get('ALIAS')
        
        if source_schema and source_table:
            # Get the original filter conditions
            filter_conditions = row.get('FILTER CONDITIONS') or  ''
            filter_condition_upper = filter_conditions.upper()
            
            # Check if DBT_VALID_TO is in the filter conditions (case-insensitive)
            if 'DBT_VALID_TO' in filter_condition_upper and 'NULL' in filter_condition_upper:
                # Use the entire filter condition in the WHERE clause
                dbt_where = f"WHERE {filter_conditions}"
                # Remove this condition from the FILTER_CONDITIONS for later use
                row['FILTER CONDITIONS'] = ''
            else:
                dbt_where = ''
            
            if current_layer == 'STG':
                sql = f"SRC_{table:<14} as ( SELECT * FROM {{{{ source('{source_schema}', '{source_table}') }}}} {dbt_where})"
            else:
                sql = f"SRC_{table:<14} as ( SELECT * FROM {{{{ ref('{source_table.lower()}') }}}} {dbt_where})"
            
            all_sql.append(sql)
        else:
            debug_print(f"WARNING: Missing SOURCE SCHEMA or SOURCE TABLE for {current_layer} layer table {table}")

    all_sql_str = ',\n'.join(all_sql)

    # Generate commented-out source references
    comment_sql = []
    for row in tables.values():
        table = row.get('ALIAS')
        schema = row.get('SOURCE SCHEMA')
        source_table = row.get('SOURCE TABLE')
        if table is not None and schema is not None and source_table is not None:
            if current_layer == 'STG':
                comment_sql.append(f"SRC_{table:<14} as ( SELECT * FROM {schema}.{source_table} )")
            else:
                comment_sql.append(f"SRC_{table:<14} as ( SELECT * FROM {schema}.{source_table} )")

    comment_sql_str = '\n'.join(comment_sql)

    # Combine both parts
    final_sql = f"{all_sql_str}\n\n/*\n{comment_sql_str}\n*/"

    debug_print(f"DEBUG: Generated Source SQL:\n{final_sql}")

    return final_sql


def build_logic_layer(columns, tables, current_layer):
    print(f'==== Build LOGIC SQL ..')
    alias_list = []
    final_sql = []

    # Collect distinct table aliases
    for row in tables.values():
        tbl = row.get('ALIAS')
        if tbl and tbl not in alias_list:
            alias_list.append(tbl)

    # Build SQL for each table
    for tbl in alias_list:
        src_tbl = f'SRC_{tbl}'
        tgt_tbl = f'LOGIC_{tbl}'

        table_cols = []
        for idx, row in enumerate(columns.values()):
            if row.get('SOURCE TABLE') == tbl:
                staging_col_name = row.get('STAGING LAYER COLUMN NAME', '')
                is_derived = row.get('IS_DERIVED', False)
                col_str = row['SQL']

                # For STG layer, add the NULLIF(TRIM()) wrapper if not already present and not a derived column
                if current_layer == 'STG' and not is_derived and not col_str.startswith('NULLIF(TRIM('):
                    col_str = f"NULLIF(TRIM({col_str}), '')"

                # Format the column string with perfect alignment
                col_str = f"{col_str:<60} as {staging_col_name:>50}"

                # Adjust the starting comma logic based on the index within the table's context
                if idx == 0 or not table_cols:
                    table_cols.append(f"        {col_str}")
                else:
                    table_cols.append(f"      , {col_str}")

        # Construct the SQL block if there are columns for the table
        if table_cols:
            cols_str = '\n'.join(table_cols)
            sql = f"""
, {tgt_tbl} as (
    SELECT
{cols_str}
    FROM {src_tbl}
)"""
            final_sql.append(sql)

    return '\n'.join(final_sql)


def build_rename_layer(columns):
    print(f'==== Build RENAME SQL ...')
    alias_list = []
    final_sql = []

    # Collect distinct table names
    for row in columns.values():
        tbl = row.get('SOURCE TABLE')
        if tbl and tbl not in alias_list:
            alias_list.append(tbl)

    # Build SQL for each table
    for tbl in alias_list:
        src_tbl = f'LOGIC_{tbl}'
        tgt_tbl = f'RENAME_{tbl}'
        table_cols = []

        # Enumerate columns specifically for the current table to reset the IDX
        for idx, row in enumerate(columns.values()):
            if row.get('SOURCE TABLE') == tbl:
                original_col = row.get('SOURCE COLUMN', '')
                target_col = row.get('STAGING LAYER COLUMN NAME', '')
                is_derived = row.get('IS_DERIVED', False)
                
                # Determine column renaming or use as is
                if is_derived:
                    col_str = target_col  # Derived columns already have their final name from the LOGIC layer
                elif original_col != target_col:
                    col_str = f"{original_col:<60} as {target_col:>50}"
                else:
                    col_str = original_col
                
                # Adjust the starting comma logic based on the index within the table's context
                if idx == 0 or not table_cols:
                    table_cols.append(f"        {col_str}")
                else:
                    table_cols.append(f"      , {col_str}")

        # Construct the SQL block if there are columns for the table
        if table_cols:
            cols_str = '\n'.join(table_cols)
            sql = f"""
, {tgt_tbl} as (
    SELECT
{cols_str}
    FROM {src_tbl}
)"""
            final_sql.append(sql)

    return '\n'.join(final_sql)


def build_filter_layer(tables):
    print(f'==== Build FILTER SQL ....')
    filter_sql = []
    
    for row in tables.values():
        alias = row['ALIAS']
        from_table = alias or row['SOURCE TABLE']
        source_table = f'RENAME_{from_table}'

        filters = row['FILTER CONDITIONS'].split(';') if row['FILTER CONDITIONS'] else []
        filters = [f for f in filters if 'DBT_VALID_TO' not in f.upper() or 'NULL' not in f.upper()]
        filters_str = ' AND '.join(filters)

        sql = f"""
, FILTER_{from_table} as (
    SELECT *
    FROM {source_table}"""

        if filters_str:
            sql += f"""
    WHERE {filters_str}"""

        sql += "\n)"
        filter_sql.append(sql)

    return '\n'.join(filter_sql)


def build_join_levels(tables):
    lvls = {}
    for row in tables.values():
        if not row[ 'PARENT TABLE JOIN' ]:
            continue
        lvl = row[ 'PARENT JOIN NUMBER' ]
        if not lvl: lvl = 0
        if lvl not in lvls:
            lvls[ lvl ] = []
        # Add Parent Join Alias to determine if it's a sub join
        row[ 'PARENT JOIN ALIAS' ] = row[ 'PARENT TABLE JOIN'].split('.')[0]
        lvls[ lvl ].append(row)

    return lvls


def build_join_layer(tables, columns, current_layer):
    debug_print("Entering build_join_layer function")
    join_sql = []

    # Sort tables based on Parent Join Number, defaulting to 1 if not specified
    sorted_tables = sorted(tables.values(), key=lambda x: int(x.get('PARENT JOIN NUMBER', '1')))

    # Check if all joins are UNION ALL
    all_union_all = all(table.get('JOIN TYPE', '').strip().upper() == 'UNION ALL' for table in sorted_tables if table.get('JOIN TYPE'))

    if all_union_all:
        join_sql.append("---- JOIN LAYER ----")
        union_cte = ", JOIN_RESULT as (\n"
        for idx, table in enumerate(sorted_tables):
            alias = table['ALIAS']
            union_cte += f"    SELECT * FROM FILTER_{alias}"
            if idx < len(sorted_tables) - 1:
                union_cte += "\n    UNION ALL\n"
        union_cte += "\n)"
        join_sql.append(union_cte)
    else:
        # Identify the base table (the one with lowest join number and no parent join)
        base_table = next((table for table in sorted_tables if table.get('PARENT TABLE JOIN') is None), sorted_tables[0])
        base_alias = base_table['ALIAS']

        join_cte = f"""
---- JOIN LAYER ----
, JOIN_RESULT as (
    SELECT *
    FROM FILTER_{base_alias}"""

        # Process joins
        for table in sorted_tables:
            if table == base_table:
                continue  # Skip the base table
            
            join_type = table.get('JOIN TYPE', '').strip().upper()
            if not join_type:
                continue
            
            alias = table['ALIAS']
            parent_join = table.get('PARENT TABLE JOIN')
            child_join = table.get('CHILD TABLE JOIN')
            
            if parent_join and child_join:
                # Prepend 'FILTER_' to table aliases in join condition
                parent_join_parts = parent_join.split('.')
                child_join_parts = child_join.split('.')
                
                if len(parent_join_parts) > 1:
                    parent_join = f"FILTER_{parent_join_parts[0]}.{'.'.join(parent_join_parts[1:])}"
                if len(child_join_parts) > 1:
                    child_join = f"FILTER_{child_join_parts[0]}.{'.'.join(child_join_parts[1:])}"
                
                join_condition = f"    ON {parent_join} = {child_join}"
                join_cte += f"""
    {join_type} FILTER_{alias}
    {join_condition}"""
            else:
                debug_print(f"Warning: Incomplete join information for table {alias}. Skipping this join.")

        join_cte += "\n)"
        join_sql.append(join_cte)

    debug_print(f"Exiting build_join_layer function. Generated SQL:\n{' '.join(join_sql)}")
    return '\n'.join(join_sql), "JOIN_RESULT"


def build_final_layer(tables, columns, current_layer, base_join_alias):
    debug_print("Entering build_final_layer function")
    
    final_columns = []
    hashdiff_cols = []
    has_hashdiff = False

    for col in columns.values():
        colname = col.get('STAGING LAYER COLUMN NAME', '')
        remove_column = (col.get('REMOVE COLUMN') or '').strip().lower()
        hashdiff = (col.get('HASHDIFF') or '').strip().lower()
        src_col = col.get('SOURCE COLUMN', '')
        is_derived = src_col == '(DERIVED)'
        manual_logic = col.get('MANUAL LOGIC', '')
        source_table = col.get('SOURCE TABLE')

        if remove_column not in ('yes', 'y') and colname not in ('HASHDIFF'):
            if is_derived and not source_table:
                # For derived columns without a source table, use the MANUAL LOGIC and alias
                final_columns.append((colname, manual_logic, True))
            else:
                # For all other columns, use the Staging Layer Column Name
                final_columns.append((colname, colname, False))

            if hashdiff in ('yes', 'y'):
                has_hashdiff = True
                hashdiff_cols.append(f"IFNULL(NULLIF(TRIM({colname}), ''), '^^')")

    # Build the SELECT statement with leading commas and aligned columns
    select_columns = []
    for idx, (col, logic, needs_alias) in enumerate(final_columns):
        if needs_alias:
            if idx == 0:
                select_columns.append(f"          {logic:<60} as {col}")
            else:
                select_columns.append(f"        , {logic:<60} as {col}")
        else:
            if idx == 0:
                select_columns.append(f"          {col}")
            else:
                select_columns.append(f"        , {col}")
    
    # Add HASHDIFF column if necessary
    if has_hashdiff and current_layer not in ['STG', 'BASE']:
        hashdiff_str = '\n            || '.join(hashdiff_cols)
        select_columns.append(f"""        , SHA2(
              {hashdiff_str}
          ) as HASHDIFF""")

    formatted_cols = '\n'.join(select_columns)

    final_sql = f"""
---- FINAL LAYER ----
SELECT
{formatted_cols}
FROM {base_join_alias}\n"""

    # Add any final layer filters as comments
    final_filter_str, filter_comment, final_filter = '','',''
    for row in tables.values():
        filter_comment += row.get('FILTER RESTRICTION RULE') or ''
        final_filter += row.get('FINAL LAYER FILTER') or  ''
        # filter_comment = row.get('FINAL LAYER FILTER', '')
    if filter_comment:
        filter_comment = filter_comment.replace("\n", " ")
        final_filter_str += f'\n/* {filter_comment} */\n'
    if final_filter:
        final_filter_str = final_filter.strip()
        # final_filter_str += f'\n {final_filter_str} \n'


    final_sql += final_filter_str

    debug_print(f"Exiting build_final_layer function. Generated SQL:\n{final_sql}")
    return final_sql


def test_dot( str_val ):
    valid = False
    pos_dot = str_val.find(".")
    pos_ubar = str_val.find("_")

    if pos_dot != -1:
        if pos_dot < pos_ubar:
            valid = True
    return valid


def do_join(made_cte, tables ):
    '''
        select *
            from FPO
                LEFT JOIN FPAD USING( POLICY_ID )
                LEFT JOIN FPAD2 on FPO.POLICY_ID = FPAD2.POLICY_IDX;
    '''
    join_sql = []
    parent_join_list = []
    # The sorted_rows group items together by Parent Join Alias with INNER JOINS being outdented
    sorted_rows = []
    sorted_rows = sorted(tables, key = itemgetter( 'PARENT JOIN ALIAS' ))
    
    # for row in table_rows:
    for row in sorted_rows:
        join_type = row[ 'JOIN TYPE' ]
        if not join_type:
            continue
        join_type = join_type.replace('_', ' ').upper()
        source_table = row[ 'ALIAS' ]
        parent_join = row[ 'PARENT TABLE JOIN' ]
        child_join = row[ 'CHILD TABLE JOIN' ]
        parent_table = parent_join.split('.')[ 0 ]

        using_join = ''
        parent_join_col = parent_join.split('.')[ -1 ]
        child_join_col = child_join.split('.')[ -1 ]

        if parent_join_col == child_join_col:
            using_join = f' using( {parent_join_col} ) '

        if 'coalesce' in parent_table:
            tempparent = parent_join.upper().replace('COALESCE(','').strip()
            parent_join = f'coalesce( FILTER_{tempparent}'
        else:
            parent_join = f'FILTER_{parent_join}'

        # Check to see if the child table join is an alias or column reference
        alias_child = test_dot( child_join)
        if alias_child:
            child_join = f'FILTER_{child_join}'
            
        if source_table in made_cte:
            sql = f'\t\t\t\t{join_type} {source_table} ON  {parent_join} = {child_join} '
            if using_join:
                sql = f'\t\t\t\t{join_type} {source_table} {using_join} '
        else:
            if parent_table in parent_join_list and not join_type.lower().startswith('inner'):
                sql = f'\t\t\t\t\t\t{join_type} FILTER_{source_table} ON  {parent_join} =  {child_join} '
            else:
                sql = f'\t\t\t\t{join_type} FILTER_{source_table} ON  {parent_join} =  {child_join} '
                if using_join:
                    sql = f'\t\t\t\t{join_type} {join_type}  FILTER_{source_table} {using_join} '
                parent_join_list.append( parent_table )

        join_sql.append( sql )

        # print( f'\n\n==== DO JOIN\nPT:{parent_table}\nSQL:{join_sql}')
    return parent_table, join_sql



def main():
    args.count, args.err = 0, 0
    plural, scd6_sql, dml_block, dml_sql  = '', '', '',''
    print('\n' + '-'*30)

    #Process each xls file in the directory
    # for xls in args.mapdir.iterdir():
    for xls in args.mapdir.glob('*.xls*'):
        print(f'\t<< Processing: {xls}')
        if 'xls' not in xls.suffix.lower():
            print(f'\t-- Skipping non-XLS file: {xls}')
            continue

        print(f'-'*30)
        print(f"Currently processing {os.path.basename(xls)}")

        if 'UNIFIED' in xls.name.upper():
            print(f'\t-- UNIFIED MAPPING: {xls.name}')
            err = process_unified(args, xls)
        else:
            print(f'\t-- Standard MAPPING: {xls.name}')
            err = process_std_map(args, xls)
            
        args.count += 1
        args.err += err

    plural = 's.' if args.count > 1 else '.'

    print(f'Processed: {args.count} file{plural}')
    if args.count > 0 : print(f'Mapping docs (XLS) in: {args.mapdir} \nModels written to: {args.modeldir}')
    if args.err > 0:
        plural =  'S.' if args.err > 1 else '.'
        print(f'\n\t!!! Encountered {args.err} ERROR{plural}... ')

    input('[ Press Enter to exit ] ')


if __name__ == '__main__':
    args = process_args()
    setup(args)

    main()
