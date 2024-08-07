import os, argparse
from pathlib import Path

# Other necessary libs
import openpyxl, yaml               # pip install openpyxl (pyyaml)
from openpyxl.utils import get_column_letter
from operator import itemgetter

from io import StringIO
import argparse
import sys, traceback

# Global variables
DEBUG = False

# Functions
def process_args():
    parser = argparse.ArgumentParser(description='Command Line Arguments')
    parser.add_argument('--rootdir', default=Path.cwd(),
                        help='The Working Directory for maps and Models (defaults to Current Directory)')
    # boolean
    parser.add_argument( '--yml', action='store_true', help='If flagged as tests, will only generate yml files')
    parser.add_argument( '--silent', action='store_true', help='If flagged will only print file names')
    parser.add_argument( '--debug', action='store_true', help='If flagged will add verbosity to output')

    args = parser.parse_args()
    args.schema = 'DEV_EDW'

    if args.rootdir:
        args.rootdir = Path(args.rootdir)

    global DEBUG
    DEBUG = args.debug

    return args


def setup(args):
    #use global args to store the rootdir and subdirs
    dir_list = {'mapdir':'mappings', 'modeldir':'models', 'snapdir':'snapshots', 'archivedir':'models_archive'}

    for dir in dir_list.items():
        setattr(args, dir[0], args.rootdir / dir[1])
        if not getattr(args, dir[0]).exists(): getattr(args, dir[0]).mkdir()

    print( '--Setting up directories--')
    for dir in dir_list:
        print( f'{dir} directory: {getattr(args, dir)}' )
    print( f'{"-"*30}' )

    print( 'Mapping docs (XLS) in:', args.mapdir )
    print( 'Output (SQL, YML) going to:', args.modeldir )

    # Show existing XLS files in Mapdir
    for afile in args.mapdir.glob('*.xlsx'):
        print(f'\tMapping found: {afile}')


def sheet2dict(sheet):
    print(f'\t<<<< Processing {sheet}')

    fields = None
    for rownum, row in enumerate( sheet.iter_rows( values_only = True) ):
        if rownum == 0:
            fields = row
            continue
        row = dict(zip( fields, row ))

        yield row

def sheet2list(sheet):
    for row in sheet.iter_rows( values_only = True ):
        if not row:
            continue
        yield row


def debug_print(*args, **kwargs):
    if DEBUG:
        print(*args, **kwargs)

def write_snap_step1( table, snapdir, scd2_sql, modeldir, scd6_sql):
    snapfile = snapdir/(table+'_SNAPSHOT_STEP1.sql')
    print('\t-- Writing to', snapfile)
    with open( snapfile, 'w') as fw:
        fw.write( scd2_sql )

    #table_split,split=table.split('_STEP1')
    if table.startswith('DIM'):
        model_file = modeldir/(table+'.sql')
    else:
        model_file = modeldir/(table+'_alt.sql')

    print('\t   Writing to', model_file)
    with open( model_file, 'w' ) as fw:
        fw.write( scd6_sql )
    # print( '========== done' )


def write_scd_step2( args, tgt_table, scd_dict ):
    modeldir = args.modeldir
    
    scd1_cols = scd_dict[ 'scd1' ] 
    exclude_col=[ 'UNIQUE_ID_KEY', 'CURRENT_INDICATOR', 'EFFECTIVE_DATE', 'END_DATE', 'LOAD_DATETIME', 'UPDATE_DATETIME', 'EFFECTIVE_TIMESTAMP', 'END_TIMESTAMP',
    'LOAD_BATCH_ID', 'UPDATE_BATCH_ID', 'PRIMARY_SOURCE_SYSTEM' ]

    scd1_cols = [ col.upper() for col in scd1_cols ]
    exclude_col = [ col.upper() for col in exclude_col ]

    # determine which columns are in scd1 but not in exclude_col without using a set
    scd1_remaining_cols = [ col for col in scd1_cols if col not in exclude_col ]

    scd1_str = ', '.join( scd1_remaining_cols )
    keycol = scd_dict[ 'KEY' ]
    
    src_table = tgt_table.replace( 'DIM_', 'DSV_' )

    # prefix = ''' {{ config(schema = 'EDW_STAGING', tags=["dim"] ) }}\n ----SRC LAYER----\nWITH\n'''
    prefix = '''\n ----SRC LAYER----\nWITH\n'''

    sql1 = prefix+f'''SCD1 as ( SELECT {scd1_str} , {keycol}    \n\t--, '1901-01-01' as DBT_VALID_FROM, '2099-12-31' as DBT_VALID_TO\n\tfrom      {{{{ ref( '{src_table}') }}}} )'''
    sql2 = f'''SCD2 as ( SELECT *    \n\tFROM      {{{{ ref('{tgt_table}_SNAPSHOT_STEP1') }}}} )'''
    sql3 = f'''FINAL as ( SELECT * 
            FROM  SCD2 
                INNER JOIN SCD1 USING( {keycol} )  )\nselect * from FINAL'''

    all_sql_str = ',\n'.join([ sql1, sql2, sql3 ])

    final_file = modeldir/(tgt_table+'_SCDALL_STEP2.sql')

    print( '\t   Writing to', final_file )
    with open( final_file, 'w' ) as fw:
        if not args.yml: fw.write( all_sql_str )  
    # print('========== done')


def process_std_map( args, xls ):
    src_file = xls.name.split('.')[0]
    table_name = src_file
    # TODO: VALIDATE THE ADDITIONAL LAYERS
    schemas = {'HUB':'RAW_VAULT', 'SAT':'RAW_VAULT', 'LNK':'RAW_VAULT', 'REF':'RAW_VAULT', 'BASE':'STAGING', 'STG':'STAGING', 'PIT':'BUSINESS_VAULT', 'BRIDGE':'BUSINESS_VAULT' }

    wb = openpyxl.load_workbook(xls)
    found_sheets = wb.sheetnames

    print(f'Sheets found: {found_sheets}\n\n')

    if 'Tables' in found_sheets:
        print(f'\t\t-- Processing Tables  in: {xls.name}' )
        tables  = process_tables( args, wb, tables_tab='Tables' )

    if 'Columns' in found_sheets:
        print(f'\t\t-- Processing Columns in: {xls.name}' )
        columns = process_columns( args, wb, columns_tab='Columns' )

    if tables and columns:
        print(f'\t\t-- Validating Table and Column aliases in: {xls.name}' )
        is_valid = validate_tables_columns( xls, tables, columns )

    if not is_valid:
        print(f'\t## ERROR: Table and Column aliases are not valid in: {xls.name}\n\n' )
        return 1

    # schemas = {'BASE':'STAGING', 'STG':'STAGING', 'HUB':'RAW_VAULT', 'SAT':'RAW_VAULT', 'LINK':'RAW_VAULT', 'REF':'RAW_VAULT',  'PIT':'BUSINESS_VAULT', 'BRIDGE':'BUSINESS_VAULT' }
    if src_file.startswith( 'STG'):
        print(f'STAGE table: {table_name} ')  # Should be the XLS filename
        build(args, args.modeldir, table_name, tables, columns, wb )
        
    if src_file.startswith( 'BASE'):
        print(f'\nBASE table: {table_name} ')  # Should be the XLS filename
        build(args, args.modeldir, table_name, tables, columns, wb )

    elif src_file.startswith( 'LNK' ):
        print(f'\nLINK table: {table_name} ')  # Should be the XLS filename
        build(args, args.modeldir, table_name, tables, columns, wb )
        
    elif src_file.startswith( 'REF' ):
        print(f'\nREFERENCE table: {table_name} ')  # Should be the XLS filename
        build(args, args.modeldir, table_name, tables, columns, wb )

    elif src_file.startswith( 'HUB' ):
        print(f'\nHUB table: {table_name} ')  # Should be the XLS filename
        build(args, args.modeldir, table_name, tables, columns, wb )

    elif src_file.startswith( 'SAT' ):
        print(f'\nSATELLITE table: {table_name} ')  # Should be the XLS filename
        build(args, args.modeldir, table_name, tables, columns, wb )

    elif src_file.startswith( 'PIT' ):
        print(f'\nPIT table: {table_name} ')  # Should be the XLS filename
        build(args, args.modeldir, table_name, tables, columns, wb )

    elif src_file.startswith( 'BRIDGE' ):
        print(f'\nBRIDGE table: {table_name} ')  # Should be the XLS filename
        build(args, args.modeldir, table_name, tables, columns, wb )

    # DIM and FACT (and FLF) should all be treated the same way    
    elif src_file.startswith( 'DIM' ):
                schema = schemas[ 'DIM' ]
                print(f'\nDIM table: {table_name} ')  # Should be the XLS filename
                # Need to check for SCD1, SCD2 columns
                # Need to process Alter tab for constraints
                alter_sql = process_alter( args, wb, 'Alter' )
                # Need to process DML tab for INSERT statements
                dml_block = process_dml( args, wb, 'DML' )

                # Combine the alter/dml to build the configuration
                config = get_config( cur_layer= table_name, alter_sql= alter_sql, dml_block= dml_block )

                # Determine if there are any SCD columns
                is_scd = False
                for col in columns.values():
                    if col.get('SCD','') == None:
                        is_scd = True
                        break
                
                scd_dict = get_scdinfo( schema, table_name, columns )
                if is_scd:
                    scd2_sql, scd6_sql = make_scd( schema, table_name, scd_dict )
                    scd6_sql = config + scd6_sql
                else:
                    scd6_sql = config

                write_snap_step1( table_name, args.modeldir, scd2_sql, args.modeldir, scd6_sql )
                write_scd_step2( args, table_name, scd_dict )

                # print( config_sql )
                # build(args, args.modeldir, table_name, tables, columns, alter_sql= alter_sql, dml_block= dml_block, config_sql = config_sql )
                build(args, args.modeldir, table_name, tables, columns, wb )
                
    elif src_file.startswith( 'FACT' ) or src_file.startswith( 'FLF' ):
        schema = schemas[ 'FCT' ]
        print(f'\nFACT table: {table_name} ')  # Should be the XLS filename
                # Need to process Alter tab for constraints
        build(args, args.modeldir, table_name, tables, columns, wb )
        
    else:
        print(f'\n\n==== Unknown mapping type for: {xls} ===')

    return 0       

    
def process_unified( args, xls ):
    print(f'\t-- UNIFIED MAPPING: {xls.name}')
    src_file = xls.name.split('.')[0]
    layers = {'HUB', 'SAT', 'LNK', 'REF', 'BASE', 'STG', 'PIT', 'BRIDGE' }
    schemas = {'HUB':'RAW_VAULT', 'SAT':'RAW_VAULT', 'LNK':'RAW_VAULT', 'REF':'RAW_VAULT', 'BASE':'STAGING', 'STG':'STAGING', 'PIT':'BUSINESS_VAULT', 'BRIDGE':'BUSINESS_VAULT' }
    addtl = ['Alter', 'DML']

    # read the XLS file and list the found sheets
    # use pandas to read the xls file
    # wb = pd.ExcelFile(xls)
    # get the sheet names using pandas
    # found_sheets = wb.sheet_names

    # use pandas to read the xls file
    wb = openpyxl.load_workbook(xls)
    found_sheets = wb.sheetnames

    print(f'Sheets found: {found_sheets}\n\n')

    for this_layer in layers:
        print(f'\t-- Attempting to Process {this_layer} Layer of UNIFIED mapping: {xls.name}')

        # Find matching sheets for this layer
        matching_sheets = [sheet for sheet in found_sheets if sheet.startswith(this_layer)]

        if not matching_sheets:
            print(f'\t\t-- No sheets found for {this_layer} layer. Skipping.')
            continue

        # Extract the table name from the sheet names
        table_name_parts = matching_sheets[0].split(' ')[1:-1]  # Remove layer and 'Tables'/'Columns'/'Alter'
        table_name = f"{this_layer}_{'_'.join(table_name_parts)}"
        
        tables_tab = next((sheet for sheet in matching_sheets if sheet.endswith('Tables')), None)
        columns_tab = next((sheet for sheet in matching_sheets if sheet.endswith('Columns')), None)
        alter_tab = next((sheet for sheet in matching_sheets if sheet.endswith('Alter')), None)
        
        schema = schemas.get(this_layer, 'UNKNOWN')

        tables = {}
        columns = {}
        alter_sql = []

        if tables_tab:
            print(f'\t\t-- Processing {tables_tab} in: {xls.name}')
            tables = process_tables(args, wb, tables_tab)
        else:
            print(f'\t\t-- No Tables sheet found for {this_layer} layer.')

        if columns_tab:
            print(f'\t\t-- Processing {columns_tab} in: {xls.name}')
            columns = process_columns(args, wb, columns_tab)
        else:
            print(f'\t\t-- No Columns sheet found for {this_layer} layer.')

        if alter_tab:
            print(f'\t\t-- Processing {alter_tab} in: {xls.name}')
            alter_sql = process_alter(args, wb, alter_tab)
        else:
            print(f'\t\t-- No Alter sheet found for {this_layer} layer.')

        if tables and columns:
            print(f'\t\t-- Validating Table and Column aliases in: {xls.name}')
            is_valid = validate_tables_columns(xls, tables, columns)

            if not is_valid:
                print(f'\t## WARNING: Table and Column aliases are not valid in: {xls.name}. Continuing processing...\n')
            
            # BUILD EACH LAYER
            print(f'{this_layer} table: {table_name}')
            config = get_config(table_name, alter_sql=alter_sql)
            build(args, args.modeldir, table_name, tables, columns, wb, alter_sql=alter_sql, config_sql=config)
            print(f'\t\t-- Finished processing {this_layer} layer for {table_name}')
        else:
            print(f'\t\t-- Skipping {this_layer} layer processing due to missing Tables or Columns data.')

    return 0

def build_scd2( schema, table, cols, keycol ):
    schema = 'EDW_STAGING_SNAPSHOT'
    all_cols = cols + [ keycol ]
    
    
    cols_str = ', '.join( all_cols )
    src_table = table.replace( 'DIM_' , 'DSV_' )
    snapshot_sql=f'''
    
    {{%% snapshot %s_SNAPSHOT_STEP1 %%}}

    {{{{
        config(
        target_schema='%s',
        unique_key='%s',
        strategy='check',
        check_cols={cols},
        )
    }}}}

    select {cols_str} 
        from  {{{{ref('%s') }}}}

    {{%% endsnapshot %%}}
    '''%(table,schema,keycol,src_table)

    return snapshot_sql

def build_scd6 ( schema, table, keycol, scd1_cols = [], scd2_cols = [], scd3_cols = [], scd0_cols = [] ):
    keycol = keycol.upper()
    if not keycol: keycol = 'UNIQUE_ID_KEY'

    exclude_col = ['UNIQUE_ID_KEY', 'CURRENT_INDICATOR', 'EFFECTIVE_DATE', 'END_DATE', 'LOAD_DATETIME', 
    'UPDATE_DATETIME', 'EFFECTIVE_TIMESTAMP', 'END_TIMESTAMP',
    'LOAD_BATCH_ID', 'UPDATE_BATCH_ID', 'PRIMARY_SOURCE_SYSTEM']

    start_sql = f'''\n\n WITH  SCD AS ( \n\tSELECT  {keycol}'''

    sql2 = []
    for col in scd2_cols:
        if col in exclude_col: continue
        sql = f'''\n {col}'''
        sql2.append( sql )

    sql3 = []
    for col in scd3_cols:
        if col in exclude_col: continue
        sql = f'''\n\t, lag({col}) over  
           (partition by {keycol} 
             order by dbt_updated_at
        ) as {col}'''
        sql3.append(sql)

    sql1 = []
    for col in scd1_cols:
        if col in exclude_col: continue
        sql=f'''\n\t, last_value({col}) over 
            (partition by {keycol} 
              order by dbt_updated_at 
                ROWS BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING
            ) as {col}'''
        sql1.append(sql)
    
    sql0 = []
    for col in scd0_cols:
        if col in exclude_col: continue
        sql=f'''\n\t, first_value({col}) over 
            (partition by {keycol} 
              order by dbt_updated_at 
                ROWS BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING
            ) as {col}'''
        sql0.append( sql )

    end_sql = f'''\n\t    , DBT_VALID_FROM AS EFFECTIVE_TIMESTAMP
    \t, DBT_VALID_TO   AS END_TIMESTAMP
    \t,
    \t  CASE WHEN CAST(DBT_VALID_FROM AS DATE) = '1901-01-01' then cast( DBT_VALID_FROM AS DATE )
    \t    WHEN cast( DBT_VALID_FROM AS DATE ) <> '1901-01-01' THEN dateadd(day,1,CAST(DBT_VALID_FROM AS DATE))
    \t  else cast( DBT_VALID_FROM AS DATE ) end as EFFECTIVE_DATE
    \t, cast( DBT_VALID_TO AS DATE ) as END_DATE
    \t, CASE WHEN DBT_VALID_TO IS NULL THEN 'Y' ELSE 'N' END AS CURRENT_INDICATOR'''
    #CAST(DBT_VALID_FROM AS DATE) as EFFECTIVE_DATE\n,
    sql1[ 0 ] = start_sql + sql1[ 0 ]
    scd_all = sql1 + sql2 + sql3 + sql0
    #+ end_sql

    scd_str = ' '.join( scd_all )
    scd_str = scd_str.strip(',')

    scd_str = scd_str + end_sql + '\n'+"     \n\tFROM {{ ref('" + table + "_SCDALL_STEP2') }})\n\n"
    scd_str += 'select * from SCD\n'

    return scd_str

def get_scdinfo( schema, table, cols_dict):
    '''
    {'Order#': None, 'SCD TYPE': None, 'STAGING LAYER COLUMN NAME': None, ' 'HASHDIFF': None, 
    'UNIQUE': None, 'NOT NULL': None, 'TEST EXPRESSION': None, 'ACCEPTED VALUES': None}
    '''
    scd_dict = {
        'table': table,
        'schema': schema,
        'KEY':'',
        'scd1':[], 'scd2':[], 'scd3':[], 'scd0':[],
    }

    keycol = 'KEY'
    for row in cols_dict.values():
        if '1' in str(row.get( 'SCD' )).split(','):
            scd_dict[ 'scd1' ].append( row.get( 'STAGING LAYER COLUMN NAME' ))
        if '2' in str(row.get( 'SCD' )).split(','):
            scd_dict[ 'scd2' ].append( row.get( 'STAGING LAYER COLUMN NAME' ))
        if '3' in str(row.get( 'SCD' )).split(','):
            scd_dict[ 'scd3' ].append( row.get( 'STAGING LAYER COLUMN NAME' ))
        if '0' in str(row.get( 'SCD' )).split(','):
            scd_dict[ 'scd0' ].append( row.get( 'STAGING LAYER COLUMN NAME' ))
        if row.get( 'SCD' ) == 'KEY':
            scd_dict[ 'KEY' ] = row.get( 'STAGING LAYER COLUMN NAME' ).upper() 
            
    return scd_dict


def process_tables(args, wb, tables_tab):
    print(f'\t\t-- Processing {tables_tab} ...')
    # read the tables tab and convert to a dictionary
    # FILTER RESTRICTION RULE should be added as a comment to the filter layer for the appropriate table, after the condition
    # Final Layer filter should be added as a comment to the final layer (after the from clause)
    expected_cols = ['SOURCE SCHEMA', 'SOURCE TABLE', 'ALIAS', 'FILTER CONDITIONS', 'FILTER RESTRICTION RULE','PARENT JOIN NUMBER', 'PARENT TABLE JOIN', 'CHILD TABLE JOIN','FINAL LAYER FILTER', 'JOIN TYPE']

    ws = wb[tables_tab]
    tables_dict = {}

    # Convert found column names to uppercase
    found_cols = [cell.value.upper() if cell.value else None for cell in ws[1]]
    missing_cols = list(set(expected_cols) - set(found_cols))
    extra_cols = list(set(found_cols) - set(expected_cols))

    if extra_cols:
        debug_print(f'\t\t\t-- Found Extra columns: {extra_cols}')
    if missing_cols:
        debug_print(f'\t\t\t-- Missing columns: {missing_cols}')

    for row in ws.iter_rows(min_row=2, values_only=True):
        if 'SOURCE TABLE' in found_cols and row[found_cols.index('SOURCE TABLE')] is not None:
            source_table = row[found_cols.index('SOURCE TABLE')]
            tables_dict[source_table] = {}
            for i, col in enumerate(found_cols):
                if col:  # Only process non-None column names
                    tables_dict[source_table][col] = row[i] if i < len(row) else ''

    # Debug: Print processed tables
    debug_print(f"DEBUG: Processed tables: {tables_dict}")

    return tables_dict


def process_columns(args, wb, columns_tab):
    print(f'\t\t-- Processing {columns_tab} ...')
    
    # read the tables tab and convert to a dictionary
    expected_cols = ['SOURCE SCHEMA', 'SOURCE TABLE', 'SOURCE COLUMN', 'DATATYPE', 'AUTOMATED LOGIC', 'MANUAL LOGIC' ,'MAPPING RULE', 'ORDER#', 'STAGING LAYER COLUMN NAME', 'STAGING LAYER DATATYPE', 'HASHDIFF', 'UNIQUE','SET DEFAULT', 'NOT NULL', 'REMOVE COLUMN', 'TEST EXPRESSION', 'ACCEPTED VALUES']
    dim_cols = ['MODEL EQUALITY', 'MODEL EQUAL ROWCOUNT', 'SCD']
    cur_layer = columns_tab.split(' ')[0]

    if columns_tab.startswith('DIM'):
        expected_cols = expected_cols + dim_cols

    ws = wb[columns_tab]
    columns_dict = {}

    # Check for missing columns and convert to a list
    # missing_cols=[x for x in expected_cols if x not in found_cols]
    # extra_cols=[x for x in found_cols if x not in expected_cols]

    # Convert found column names to uppercase
    found_cols = [cell.value.upper() if cell.value else None for cell in ws[1]]
    
    missing_cols = list(set(expected_cols) - set(found_cols))
    extra_cols = list(set(found_cols) - set(expected_cols))

    if extra_cols:
        print(f'\t\t\t-- Found Extra columns: {extra_cols}')
    if missing_cols:
        print(f'\t\t\t-- Missing columns: {missing_cols}')
        
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if row[found_cols.index('STAGING LAYER COLUMN NAME')] is not None:
            columns_dict[idx] = {}
            for i, col in enumerate(found_cols):
                if col:  # Only process non-None column names
                    columns_dict[idx][col] = row[i] if i < len(row) else ''

    # Debug: Print processed columns
    debug_print(f"DEBUG: Processed columns: {columns_dict}")

    return columns_dict


def validate_tables_columns(xls, tables, columns):
    is_valid = True
    current_layer = next(iter(tables.values())).get('SOURCE TABLE', '').split('_')[0]

    if current_layer == 'STG':
        # For STG layer, we don't validate aliases
        return True

    table_aliases = set(tables.keys())
    column_aliases = set(col['SOURCE COLUMN'] for col in columns.values() if col.get('SOURCE COLUMN') is not None)

    missing_tables = list(table_aliases - column_aliases)
    missing_columns = list(column_aliases - table_aliases)

    if missing_tables:
        print(f'\t\t\t-- Missing Table references on Columns Tab: {missing_tables}')
        is_valid = False
    if missing_columns:
        print(f'\t\t\t-- Missing Columns alias in Tables tab: {missing_columns}')
        is_valid = False

    return is_valid


def process_alter( args, wb, alter_tab ):
    print(f'\t\t-- Processing {alter_tab} ...')

    ws = wb[alter_tab]
    alter_sql = []

    # read each row of the alter_tab and append to the alter_sql list if it's not blank
    for row in ws.iter_rows(min_row=0, values_only=True):
        if row[0]:
            if row[0].strip() != '':
                alter_sql.append(row[0])
    
    return alter_sql

def process_dml( args, wb, dml_tab ):
    print(f'\t\t-- Processing DML Statements ...' )

    ws = wb[dml_tab]
    dml_sql = []

    # read the dml_tab and append each row to the dml_sql list
    for row in ws.iter_rows(min_row=0, values_only=True):
        if row != '':
            dml_sql.append(row[0])

    return dml_sql

def get_config( cur_layer, alter_sql='', dml_block='' ):
    '''
    Builds the CONFIG dynamically based on the file name
        # {{ config(   only if...:
        # ADD: materialized = 'view' if the tgt_table starts with DSV
        # ADD: tags = [ "fact" ]  if the tgt_table starts with FACT%
        # Only add post_hook = (" if there are alter statements.
        # ") <- close post_hook 
        # }} <- close config
    '''
    print( '-- COMBINING ALTER/DML statements ')
    concat_config = '' 
    config = {}
    compileif = ['view', 'tags', 'alter']
    
    if cur_layer.startswith('SAT') or cur_layer.startswith('HUB') or cur_layer.startswith('LNK'):
        config['tags'] = f'tags = [ "raw_vault" ]'

    if alter_sql:
        alter_sql = '\n'.join(alter_sql)
        config['alter'] = f'post_hook = ("\n{alter_sql}\n") '

    concat_config = ''
    compile_cnt = 0
    for checkval in compileif:
        if checkval in config and compile_cnt == 0:
            concat_config = f'{{{{ config( \n{config.get(checkval)} '
            compile_cnt += 1
        elif checkval in config:
            concat_config += f'\n,\t {config.get(checkval)}'
            compile_cnt += 1
    if compile_cnt > 0:
        concat_config += '\n) }}'

    config_sql = ''.join(concat_config)

    return config_sql


def get_model_tests(row):
    '''
    https://github.com/fishtown-analytics/dbt-utils
    '''
    test_dict = {}
    test_cols = ['UNIQUE', 'MODEL EQUALITY', 'MODEL EQUAL ROWCOUNT', 'TEST EXPRESSION',  ]
    # Globally replacing the relationship with relationship_where

    tests = []
    add = {'severity': 'warn'}

    for atest in test_cols:
        if atest not in row:
            continue
        if not row[atest]:
            continue

        if atest == 'UNIQUE':
            if ':' in str(row[atest]):
                test = str(row[atest]).replace('COMPOSITE:', '').strip()
                test = test.replace("unique:", "").replace("Unique:", "").strip()
                row[atest] = {
                    'dbt_utils.unique_combination_of_columns': {
                        'combination_of_columns': [each.strip() for each in test.split(',')],
                        'severity': 'warn'
                    }
                }
            else:
                continue

        if atest == 'MODEL EQUALITY':
            # print(f' $$$$$$$$ FOUND Model Equality ->  {row[ atest ]}')               # DEBUG
            test =   str(row[ atest ]).strip()
            row[atest] = {
                'dbt_utils.equality': 
                {
                    'compare_model': f"ref( {row[ 'STAGING LAYER COLUMN NAME' ]} )",
                    'compare_columns': [each.strip() for each in test.split(',')]                    
                    # 'compare_columns': [ f'ref({row[ atest ]!r})' ]
                }
            }

        if atest == 'MODEL EQUAL ROWCOUNT':
            # print(f' $$$$$$$$ FOUND Model Equal Rowcount -=>  {row[ atest ]}')        # DEBUG
            row[atest] = {
                'dbt_utils.equal_rowcount': 
                {
                    'compare_model': f'ref({row[ atest ]!r})'
                }
            }

        # Relationship moved to column test...

        if atest == 'TEST EXPRESSION':
            if row.get('STAGING LAYER COLUMN NAME','x') in row.get('TEST EXPRESSION','y'):
                # If true, then should be a model test, else column test
                test =   str(row[ atest ]).strip()
                row[atest] = {
                    'dbt_utils.expression_is_true':
                    {
                        'severity': 'warn',
                        'expression': test
                    }
                }
            else: continue
        tests.append(row[atest])

    return tests

def get_column_tests(row, table_name):
    '''
    https://github.com/fishtown-analytics/dbt-utils
    '''
    test_dict = {}
    test_cols = [ 'UNIQUE', 'NOT NULL', 'TEST EXPRESSION', 'ACCEPTED VALUES', 'RELATIONSHIP', 'SET DEFAULT']
    # 'Equal Rowcount',  'Mutually Exclusive Ranges', 'Unique Combination of Columns', 'Accepted Range']

    tests = []
    add = {'severity': 'warn'}

    for atest in test_cols:
        if atest not in row:
            continue
        if not row[atest]:
            continue

        # If a FLF or FACT build, do not include the has_default test.  Will require knowing the layer during build.
        if atest == 'SET DEFAULT':
            test = str( row[ atest ] ).strip()
            row['DEFAULT']= test
            if table_name.startswith('F'):
                row[atest]=''
            else:
                row[atest]=''
                # row[atest] = {
                #     'has_default':
                #     {
                #         'severity': 'warn',
                #         'hashval': [each.strip() for each in test.split(',')]
                #     }
                # }

        if atest == 'UNIQUE':
            if ':' not in str(row[atest]):
                row[atest] = {
                    'unique':
                    {
                        'severity': 'warn'
                    }
                }
            else:
                continue

        if atest == 'NOT NULL':
            if 'where' in row[atest]:
                # TODO: Fix quoting around where
                test = str( row[ atest ] ).upper().replace( 'WHERE','').strip()
                row[atest] = {
                    'dbt_utils.not_null_where':
                    {
                        'where': test
                    }
                }                
            else:
                row[atest] = {
                    'not_null':
                    {
                        'severity': 'warn'
                    }
                }

        if atest == 'ACCEPTED VALUES':
            if 'not:' in row[atest]:
            # Split to break them into separate rows...
                test = str( row[ atest ] ).upper().replace( 'NOT:','').strip()
                row[atest] = {
                    'not_accepted_values':
                    {
                        'severity': 'warn',
                        'values': [each.strip() for each in test.split(',')]
                    }
                }
            else:
                test = str( row[ atest ] )
                row[atest] = {
                    'accepted_valids':
                    {
                        'severity': 'warn',
                        'values': [each.strip() for each in test.split(',')]
                    }
                }

        if atest == 'TEST EXPRESSION':
            # If the column name is not listed, then it can be a column test, else model test
            if not row.get('STAGING LAYER COLUMN NAME','x') in str(row[ atest ]).strip():
                test =   str(row[ atest ]).strip()
                row[atest] = {
                    'expression_is_true_ignore':
                    {
                        'ignore': '',
                        'severity': 'warn',
                        'expression': test
                    }
                }
            else: continue

        if atest == 'RELATIONSHIP':
            ignore_in = ''
            # Changed to a relationships_where instead to handle ignoring dummy value:  md5('-1111') 
            # Determine if the field is different (Based on TableName.ColumnName )
            if '.' in row[atest]:
                ref_table = row[ atest ].split('.')[0]
                ref_field = row[ atest ].split('.')[1]
                ref_ignore = row[ 'STAGING LAYER COLUMN NAME' ]
            else:
                ref_table = row[ atest ]
                ref_field = row[ 'STAGING LAYER COLUMN NAME' ]
                ref_ignore = row[ 'STAGING LAYER COLUMN NAME' ]

            if ref_field == 'DATE_KEY':
                ignore_in = " -1, -2, -3 "
            else: 
                ignore_list = row.get('SET DEFAULT')
                if ignore_list != None:
                    ignore_vals = str(ignore_list).split(',')
                    ignore_md5 = [f'{val}' for val in ignore_vals ]
                    ignore_in = ', '.join( ignore_md5 )
                else:
                    ignore_in = " '40c5dea533476acdd01f7ef0e84de22f', 'fcbcdcb8f6b1c597c5fdc7a54cd321ae' "
                    
                

            row[atest] = {
                'dbt_utils.relationships_where':
                {
                    'to': f'ref( {ref_table!r} )',
                    'field': f"{ref_field}"
                }
            }

        if not row[atest] == '':
            tests.append(row[atest])

    return tests


def get_source_sql(target_table, cols, tables):
    current_layer = target_table.split('_')[0]
    columns = cols.copy()

    for idx, row in enumerate(columns.values()):
        src_col = row.get('SOURCE COLUMN', '')
        manual_logic = row.get('MANUAL LOGIC', '')
        datatype = row.get('DATATYPE')
        automated_logic = row.get('AUTOMATED LOGIC')
        staging_col_name = row.get('STAGING LAYER COLUMN NAME', '')

        if current_layer == 'STG':
            if src_col == '(DERIVED)':
                src_sql = manual_logic.strip() if manual_logic else staging_col_name
                row['LOGIC_NAME'] = staging_col_name
            else:
                src_sql = src_col
                if automated_logic:
                    automated_logic = automated_logic.lower()
                    if 'trim' in automated_logic:
                        src_sql = f"TRIM({src_sql})"
                    if 'upper' in automated_logic:
                        src_sql = f"UPPER({src_sql})"
                src_sql = f"NULLIF(TRIM( {src_sql} ), '')"
                row['LOGIC_NAME'] = src_col
        else:
            if src_col == '(DERIVED)':
                if manual_logic:
                    if 'HASH' in manual_logic.upper() or 'COMPOSITE' in manual_logic.upper():
                        hashcol = manual_logic.replace('HASH:', '').replace('COMPOSITE:', '').replace('\n', '').replace(' ', '').strip()
                        temp1 = hashcol.split(',')
                        joincol = (', ').join(temp1)
                        hashcol = ("','").join(temp1)
                        col1 = f"{{ dbt_utils.generate_surrogate_key ( [ '{hashcol}' ] ) }}"
                        src_sql = f"{{{col1}}}"
                    else:
                        src_sql = manual_logic.strip()
                elif datatype == 'DATE':
                    src_sql = f"CASE WHEN {staging_col_name} is null then '-1' WHEN {staging_col_name} < '1901-01-01' then '-2' WHEN {staging_col_name} > '2099-12-31' then '-3' ELSE regexp_replace( {staging_col_name}, '[^0-9]+', '') END :: INTEGER"
                else:
                    src_sql = staging_col_name
                row['LOGIC_NAME'] = staging_col_name
            else:
                src_sql = src_col
                row['LOGIC_NAME'] = src_col

        row['SQL'] = src_sql
        row['RENAME'] = staging_col_name
        row['IS_DERIVED'] = src_col == '(DERIVED)'

    return columns

def make_scd( schema, table_name, scd_dict):
    '''
    scd={
        'table':'STG_POLICYTEST',
        'schema':'X10057301',
        'scd1':['POLICY_ID','CLAIM_ID'],
        'scd2':['PTYPE'],
        'scd3':['OWNER_NAME','ADDR'],
        'keycol':'POLICY_ID'
    }
    '''

    keycol = 'KEY'

    #print(scd_dict)
    #cols=scd_dict['scd2']+scd_dict['scd3']+scd_dict['scd0']
    combined_scd_cols = scd_dict[ 'scd2' ] + scd_dict[ 'scd3' ] + scd_dict[ 'scd0' ]

    scd2_sql = build_scd2( schema, table_name,  combined_scd_cols, scd_dict[ keycol ])

    scd6_sql = build_scd6( schema, table_name,  scd_dict[ keycol ],
        scd1_cols = scd_dict[ 'scd1' ], scd2_cols = scd_dict[ 'scd2' ], scd3_cols = scd_dict[ 'scd3' ], scd0_cols = scd_dict[ 'scd0' ])

    return scd2_sql, scd6_sql


def make_yml(table_name, tables, columns):
    print(f'==== Building YAML for {table_name} ...')
    
    model_tests = []
    column_tests = {}

    for row in columns.values():
        col_name = row.get('STAGING LAYER COLUMN NAME')
        if col_name:
            model_tests.extend(get_model_tests(row))
            column_tests[col_name] = get_column_tests(row, table_name)

    schema_dict = {
        'version': 2,
        'models': [
            {
                'name': table_name,
                'tests': model_tests,
                'columns': []
            }
        ]
    }

    for col_name, tests in column_tests.items():
        if tests:
            column_entry = {
                'name': col_name,
                'tests': tests
            }
            schema_dict['models'][0]['columns'].append(column_entry)

    debug_print(f"Generated YAML for {table_name}:")
    debug_print(yaml.dump(schema_dict, sort_keys=False))

    return schema_dict


def format_select_cols(columns, layer):
    select_cols = []
    for idx, col in enumerate(columns.values()):
        col_name = col.get('STAGING LAYER COLUMN NAME')
        src_sql = col.get('SQL')
        
        if col_name:
            if layer == 'FINAL':
                col_str = col_name
            else:
                if '\n' in src_sql:
                    src_sql = format_manual_logic(src_sql, ' ' * 12)
                col_str = f"{src_sql:<60} as {col_name:>50}"

            if idx == 0:
                select_cols.append(f"        {col_str}")
            else:
                select_cols.append(f"      , {col_str}")
    
    return '\n'.join(select_cols)


def format_manual_logic(logic, indent):
    if not logic or '\n' not in logic:
        return logic
    lines = logic.split('\n')
    return f"{lines[0]}\n" + '\n'.join([f"{indent}{line.strip()}" for line in lines[1:]])


def build(args, modeldir, table_name, tables, columns, wb, alter_sql='', config_sql=''):
    print(f'\t\t-- Building {table_name} ...')
    yml = make_yml(table_name, tables, columns)
    all_sql = []

    try:
        current_layer = table_name.split('_')[0]
        columns = get_source_sql(table_name, columns, tables)

        # Config and SRC LAYER
        all_sql.append(config_sql)
        if config_sql:
            all_sql.append('\n')  # Add a newline after the config block
        all_sql.append('---- SRC LAYER ----\nWITH')
        src_sql = build_src_layer(args, tables, current_layer)
        all_sql.append(src_sql)

        # LOGIC LAYER
        all_sql.append('---- LOGIC LAYER ----')
        logic_sql = build_logic_layer(columns, tables, current_layer)
        all_sql.append(logic_sql)

        # RENAME LAYER
        all_sql.append('---- RENAME LAYER ----')
        rename_sql = build_rename_layer(columns)
        all_sql.append(rename_sql)

        # FILTER LAYER
        all_sql.append('---- FILTER LAYER ----')
        filter_sql = build_filter_layer(tables)
        all_sql.append(filter_sql)

        # JOIN and FINAL LAYER
        join_sql = build_join_layer(tables, columns, current_layer)
        all_sql.append(join_sql)

        sql_str = '\n'.join(all_sql)

        # Write files
        yml_file = modeldir / (table_name + '.yml')
        sql_file = modeldir / (table_name + '.sql')

        with open(yml_file, 'w') as fw:
            yaml.dump(yml, fw, sort_keys=False)
            print(f'\t\t-- Created YAML file: {yml_file}')

        with open(sql_file, 'w') as fw:
            fw.write(sql_str)
            print(f'\t\t-- Created SQL file: {sql_file}')

    except Exception as e:
        print(f'\t\t-- Error building {table_name}: {str(e)}')
        traceback.print_exc()

    print(f'\t\t-- Finished building {table_name}')


def build_src_layer(args, tables, current_layer):
    '''
    {1: {'SOURCE SCHEMA': 'STAGING', 'SOURCE TABLE': 'STG_DEP_MEMBERS', 'ALIAS': 'M', 'FILTER CONDITIONS': None, 'Parent Join Number': 2, 'PARENT TABLE JOIN': None, 'Child Table Join': None, 'JOIN TYPE': None}, 2: {'SOURCE SCHEMA': 'STAGING', 'SOURCE TABLE': 'STG_DEP_ADDRESS', 'ALIAS': 'A', 'FILTER CONDITIONS': None, 'Parent Join Number': 2, 'PARENT TABLE JOIN': 'CONCAT(M.PEACH_BASE_NUMBER, M.PEACH_SUFFIX_NUMBER)', 'Child Table Join': 'CONCAT(A_PEACH_BASE_NUMBER, A_PEACH_SUFFIX_NUMBER)', 'JOIN TYPE': 'left join'}, 3: {'SOURCE SCHEMA': 'STAGING', 'SOURCE TABLE': 'STG_DEP_US_ZIPCODE_LIST', 'ALIAS': 'Z', 'FILTER CONDITIONS': None, 'Parent Join Number': 1, 'PARENT TABLE JOIN': 'A.ZIP_CODE', 'Child Table Join': 'Z.ZIPCODE', 'JOIN TYPE': 'left join'}}
    '''
    debug_print(f'==== Build SOURCE SQL for {current_layer} layer.')
    all_sql = []

    if current_layer == 'STG':
        # For STG layer, we create a single CTE with the alias 'STG'
        source_schema = next(iter(tables.values())).get('SOURCE SCHEMA')
        source_table = next(iter(tables.values())).get('SOURCE TABLE')
        if source_schema and source_table:
            sql = f"STG as ( SELECT * FROM {{{{ source('{source_schema}', '{source_table}') }}}} )"
            all_sql.append(sql)
        else:
            debug_print(f"WARNING: Missing SOURCE SCHEMA or SOURCE TABLE for STG layer")
    else:
        # For other layers, keep the existing logic
        for tbl in tables.values():
            table = tbl.get('ALIAS')
            if table is not None:
                act_table = tbl.get('SOURCE TABLE')
                sql = f"SRC_{table:<14} as ( SELECT * FROM {{{{ ref('{act_table}') }}}} )"
                all_sql.append(sql)

    all_sql_str = ',\n'.join(all_sql)

    # Generate commented-out source references
    comment_sql = []
    if current_layer == 'STG':
        if source_schema and source_table:
            comment_sql.append(f"STG as ( SELECT * FROM {source_schema}.{source_table} )")
    else:
        for tbl in tables.values():
            table = tbl.get('ALIAS')
            schema = tbl.get('SOURCE SCHEMA')
            if table is not None:
                act_table = tbl.get('SOURCE TABLE')
                comment_sql.append(f"SRC_{table:<14} as ( SELECT * FROM {schema}.{act_table} )")

    comment_sql_str = '\n'.join(comment_sql)

    # Combine both parts
    final_sql = f"{all_sql_str}\n\n/*\n{comment_sql_str}\n*/"

    debug_print(f"DEBUG: Generated Source SQL:\n{final_sql}")

    return final_sql


def build_logic_layer(columns, tables, current_layer):
    print(f'==== Build LOGIC SQL ..')
    alias_list = []
    final_sql = []

    for row in tables.values():
        tbl = row.get('ALIAS')
        if tbl and tbl not in alias_list:
            alias_list.append(tbl)

    for tbl in alias_list:
        src_tbl = f'SRC_{tbl}'
        tgt_tbl = f'LOGIC_{tbl}'
        
        table_cols = []
        for idx, row in enumerate(columns.values()):
            if row.get('SOURCE TABLE') == tbl:
                src_col = row.get('SQL', '')
                logic_name = row.get('LOGIC_NAME', '')
                is_derived = row.get('IS_DERIVED', False)
                
                if is_derived:
                    if 'dbt_utils.generate_surrogate_key' in src_col:
                        # This is a HASH or COMPOSITE derived column
                        col_str = f"{src_col:<60} as {logic_name:>50}"
                    else:
                        col_str = f"{src_col:<60} as {logic_name:>50}"
                elif current_layer == 'STG':
                    col_str = f"{src_col:<60} as {logic_name:>50}"
                else:
                    col_str = logic_name
                
                if idx == 0:
                    table_cols.append(f"        {col_str}")
                else:
                    table_cols.append(f"      , {col_str}")
        
        cols_str = '\n'.join(table_cols)
        
        sql = f"""
, {tgt_tbl} as (
    SELECT
{cols_str}
    FROM {src_tbl}
)"""
        final_sql.append(sql)

    return '\n'.join(final_sql)


def build_rename_layer(columns):
    print(f'==== Build RENAME SQL ...')
    alias_list = []
    final_sql = []

    for row in columns.values():
        tbl = row.get('SOURCE TABLE')
        if tbl and tbl not in alias_list:
            alias_list.append(tbl)

    for tbl in alias_list:
        src_tbl = f'LOGIC_{tbl}'
        tgt_tbl = f'RENAME_{tbl}'
        table_cols = []

        for idx, row in enumerate(columns.values()):
            if row.get('SOURCE TABLE') == tbl:
                original_col = row.get('SOURCE COLUMN', '')
                target_col = row.get('STAGING LAYER COLUMN NAME', '')
                is_derived = row.get('IS_DERIVED', False)
                
                if is_derived:
                    col_str = target_col  # Derived columns already have their final name from the LOGIC layer
                elif original_col != target_col:
                    col_str = f"{original_col:<60} as {target_col:>50}"
                else:
                    col_str = original_col
                
                if idx == 0:
                    table_cols.append(f"        {col_str}")
                else:
                    table_cols.append(f"      , {col_str}")

        if table_cols:
            cols_str = '\n'.join(table_cols)
            sql = f"""
, {tgt_tbl} as (
    SELECT
{cols_str}
    FROM {src_tbl}
)"""
            final_sql.append(sql)

    return '\n'.join(final_sql)


def build_filter_layer(tables):
    print(f'==== Build FILTER SQL ....')
    filter_sql = []
    
    for row in tables.values():
        alias = row['ALIAS']
        from_table = alias or row['SOURCE TABLE']
        source_table = f'RENAME_{from_table}'

        filters = row['FILTER CONDITIONS'].split(';') if row['FILTER CONDITIONS'] else []
        filters_str = ' AND '.join(filters)

        sql = f"""
, FILTER_{from_table} as (
    SELECT *
    FROM {source_table}"""

        if filters_str:
            sql += f"""
    WHERE {filters_str}"""

        sql += "\n)"
        filter_sql.append(sql)

    return '\n'.join(filter_sql)


def build_join_layer(tables, columns):
    debug_print("Entering build_join_layer function")
    join_sql = []
    
    first_table = next(iter(tables.values()), {})
    current_layer = first_table.get('SOURCE TABLE', '').split('_')[0]
    
    debug_print(f"Current layer: {current_layer}")

    formatted_cols = format_select_cols(columns, 'FINAL')

    if current_layer == 'STG':
        debug_print("Processing STG layer")
        join_sql.append(f"""---- JOIN LAYER ----
-- No joins for STG layer

---- FINAL LAYER ----
SELECT
{formatted_cols}
FROM STG""")
    else:
        debug_print("Processing non-STG layer")
        if len(tables) == 0: 
            print(' --- NOTHING TO PROCESS ---')
            return ''

        final_alias = ''
        aliases = {}
        for row in tables.values():
            if 'SOURCE SCHEMA' not in row or 'SOURCE TABLE' not in row:
                debug_print(f"WARNING: Missing required fields in row: {row}")
                continue
            alias = row.get('ALIAS')
            aliases[row['SOURCE TABLE']] = alias
            if not row.get('JOIN TYPE'):
                final_alias = alias

        join_lvls = build_join_levels(tables)
        all_sql = []

        all_sql.append("---- JOIN LAYER ----")
        
        made_cte = set()
        for lvl in sorted(join_lvls):
            parent_table, join_sql = do_join(made_cte, join_lvls[lvl])
            made_cte.add(parent_table)
            join_sql_str = '\n        '.join(join_sql)
            
            sql = f"""
, {parent_table} as (
    SELECT *
    FROM FILTER_{parent_table}
        {join_sql_str}
)"""
            all_sql.append(sql)

        final_sql = f"""
---- FINAL LAYER ----
SELECT
{formatted_cols}
FROM {final_alias}"""

        all_sql.append(final_sql)

        all_sql_str = '\n'.join(all_sql)

        final_filter_str = ''
        for row in tables.values():
            filter_comment = row.get('FINAL LAYER FILTER', '')
            if filter_comment:
                filter_comment = filter_comment.replace("\n", " ")
                final_filter_str += f'\n/* {filter_comment} */\n'

        all_sql_str += final_filter_str
        join_sql.append(all_sql_str)
                
    debug_print(f"Exiting build_join_layer function. Generated SQL:\n{' '.join(join_sql)}")
    return '\n'.join(join_sql)


def build_join_levels(tables):
    lvls = {}
    for row in tables.values():
        if not row[ 'PARENT TABLE JOIN' ]:
            continue
        lvl = row[ 'PARENT JOIN NUMBER' ]
        if not lvl: lvl = 0
        if lvl not in lvls:
            lvls[ lvl ] = []
        # Add Parent Join Alias to determine if it's a sub join
        row[ 'PARENT JOIN ALIAS' ] = row[ 'PARENT TABLE JOIN'].split('.')[0]
        lvls[ lvl ].append(row)

    return lvls


def build_join_layer(tables, columns, current_layer):
    debug_print("Entering build_join_layer function")
    join_sql = []
    
    first_table = next(iter(tables.values()), {})
    final_alias = first_table.get('ALIAS') or first_table.get('SOURCE TABLE')

    if len(tables) <= 1:
        # No joins needed, just select from the FILTER layer
        join_sql.append(f"""
---- JOIN LAYER ----
, JOIN_{final_alias} as (
    SELECT *
    FROM FILTER_{final_alias}
)""")
    else:
        debug_print("Processing non-STG layer")
        if len(tables) == 0: 
            print(' --- NOTHING TO PROCESS ---')
            return ''

        final_alias = ''
        aliases = {}
        for row in tables.values():
            if 'SOURCE SCHEMA' not in row or 'SOURCE TABLE' not in row:
                debug_print(f"WARNING: Missing required fields in row: {row}")
                continue
            alias = row.get('ALIAS')
            aliases[row['SOURCE TABLE']] = alias
            if not row.get('JOIN TYPE'):
                final_alias = alias

        join_lvls = build_join_levels(tables)
        all_sql = []

        all_sql.append("---- JOIN LAYER ----")
        
        made_cte = set()
        for lvl in sorted(join_lvls):
            parent_table, join_sql = do_join(made_cte, join_lvls[lvl])
            made_cte.add(parent_table)
            join_sql_str = '\n        '.join(join_sql)
            
            sql = f"""
, {parent_table} as (
    SELECT *
    FROM FILTER_{parent_table}
        {join_sql_str}
)"""
            all_sql.append(sql)

    # Handle HASHDIFF
    hashdiff_cols = []
    print("DEBUG: Checking HASHDIFF columns")
    for idx, col in enumerate(columns.values()):
        try:
            hashdiff_value = col.get('HASHDIFF')
            colname = col.get('STAGING LAYER COLUMN NAME', '')
            debug_print(f"DEBUG: Column {idx}: Name={colname}, HASHDIFF value='{hashdiff_value}'")
            
            # Check if hashdiff_value is not None and has non-whitespace characters
            if hashdiff_value is not None and hashdiff_value.strip():
                hashdiff_cols.append(f"IFNULL(NULLIF(TRIM({colname}), ''), '^^')")
        except AttributeError as e:
            print(f"DEBUG: Error processing column {idx}: {str(e)}")
        except Exception as e:
            print(f"DEBUG: Unexpected error processing column {idx}: {str(e)}")

    formatted_cols = format_select_cols(columns, 'FINAL')
    
    debug_print(f"DEBUG: Number of HASHDIFF columns: {len(hashdiff_cols)}")
    if hashdiff_cols:  # Only generate HASHDIFF if we have at least one column
        hashdiff_str = ' ||\n            '.join(hashdiff_cols)
        formatted_cols += f"""
      , SHA2(
            {hashdiff_str}
        )                                                   as HASHDIFF"""
    else:
        debug_print("DEBUG: No HASHDIFF columns found, skipping HASHDIFF generation")

    final_sql = f"""
---- FINAL LAYER ----
SELECT
{formatted_cols}
FROM JOIN_{final_alias}"""

    join_sql.append(final_sql)

    debug_print(f"Exiting build_join_layer function. Generated SQL:\n{' '.join(join_sql)}")
    return '\n'.join(join_sql)

        # final_filter_str = ''
        # for row in tables.values():
        #     filter_comment = row.get('FINAL LAYER FILTER', '')
        #     if filter_comment:
        #         filter_comment = filter_comment.replace("\n", " ")
        #         final_filter_str += f'\n/* {filter_comment} */\n'

        # all_sql_str += final_filter_str
        # join_sql.append(all_sql_str)
                


def test_dot( str_val ):
    valid = False
    pos_dot = str_val.find(".")
    pos_ubar = str_val.find("_")

    if pos_dot != -1:
        if pos_dot < pos_ubar:
            valid = True
    return valid


def do_join(made_cte, tables ):
    '''
        select *
            from FPO
                LEFT JOIN FPAD USING( POLICY_ID )
                LEFT JOIN FPAD2 on FPO.POLICY_ID = FPAD2.POLICY_IDX;
    '''
    join_sql = []
    parent_join_list = []
    # The sorted_rows group items together by Parent Join Alias with INNER JOINS being outdented
    sorted_rows = []
    sorted_rows = sorted(tables, key = itemgetter( 'PARENT JOIN ALIAS' ))
    
    # for row in table_rows:
    for row in sorted_rows:
        join_type = row[ 'JOIN TYPE' ]
        if not join_type:
            continue
        join_type = join_type.replace('_', ' ').upper()
        source_table = row[ 'ALIAS' ]
        parent_join = row[ 'PARENT TABLE JOIN' ]
        child_join = row[ 'CHILD TABLE JOIN' ]
        parent_table = parent_join.split('.')[ 0 ]

        using_join = ''
        parent_join_col = parent_join.split('.')[ -1 ]
        child_join_col = child_join.split('.')[ -1 ]

        if parent_join_col == child_join_col:
            using_join = f' using( {parent_join_col} ) '

        if 'coalesce' in parent_table:
            tempparent = parent_join.upper().replace('COALESCE(','').strip()
            parent_join = f'coalesce( FILTER_{tempparent}'
        else:
            parent_join = f'FILTER_{parent_join}'

        # Check to see if the child table join is an alias or column reference
        alias_child = test_dot( child_join)
        if alias_child:
            child_join = f'FILTER_{child_join}'
            
        if source_table in made_cte:
            sql = f'\t\t\t\t{join_type} {source_table} ON  {parent_join} = {child_join} '
            if using_join:
                sql = f'\t\t\t\t{join_type} {source_table} {using_join} '
        else:
            if parent_table in parent_join_list and not join_type.lower().startswith('inner'):
                sql = f'\t\t\t\t\t\t{join_type} FILTER_{source_table} ON  {parent_join} =  {child_join} '
            else:
                sql = f'\t\t\t\t{join_type} FILTER_{source_table} ON  {parent_join} =  {child_join} '
                if using_join:
                    sql = f'\t\t\t\t{join_type} {join_type}  FILTER_{source_table} {using_join} '
                parent_join_list.append( parent_table )

        join_sql.append( sql )

        # print( f'\n\n==== DO JOIN\nPT:{parent_table}\nSQL:{join_sql}')
    return parent_table, join_sql



def main():
    args.count, args.err = 0, 0
    plural, scd6_sql, dml_block, dml_sql  = '', '', '',''
    print('\n' + '-'*30)

    #Process each xls file in the directory
    # for xls in args.mapdir.iterdir():
    for xls in args.mapdir.glob('*.xls*'):
        print(f'\t<< Processing: {xls}')
        if 'xls' not in xls.suffix.lower():
            print(f'\t-- Skipping non-XLS file: {xls}')
            continue

        print(f'-'*30)
        print(f"Currently processing {os.path.basename(xls)}")

        if 'UNIFIED' in xls.name.upper():
            print(f'\t-- UNIFIED MAPPING: {xls.name}')
            err = process_unified(args, xls)
        else:
            print(f'\t-- Standard MAPPING: {xls.name}')
            err = process_std_map(args, xls)
            
        args.count += 1
        args.err += err

    plural = 's.' if args.count > 1 else '.'

    print(f'Processed: {args.count} file{plural}')
    if args.count > 0 : print(f'Mapping docs (XLS) in: {args.mapdir} \nModels written to: {args.modeldir}')
    if args.err > 0:
        plural =  'S.' if args.err > 1 else '.'
        print(f'\n\t!!! Encountered {args.err} ERROR{plural}... ')

    input('[ Press Enter to exit ] ')


if __name__ == '__main__':
    args = process_args()
    setup(args)

    main()
