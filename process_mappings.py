import os, argparse
from pathlib import Path

# Other necessary libs
import openpyxl, yaml               # pip install openpyxl (pyyaml)
import pandas as pd
from operator import itemgetter

# Global variables

# Functions
def process_args():
    parser = argparse.ArgumentParser(description='Command Line Arguments')
    parser.add_argument('--rootdir', default=Path.cwd(),
                        help='The Working Directory for maps and Models (defaults to Current Directory)')
    # boolean
    parser.add_argument( '--yml', action='store_true', help='If flagged as tests, will only generate yml files')
    parser.add_argument( '--silent', action='store_true', help='If flagged will only print file names')
    parser.add_argument( '--debug', action='store_true', help='If flagged will add verbosity to output')

    args = parser.parse_args()
    args.schema = 'DEV_EDW'

    return args


def setup(args):
    #use global args to store the rootdir and subdirs
    dir_list = {'mapdir':'mappings', 'modeldir':'models', 'snapdir':'snapshots', 'archivedir':'models_archive'}

    for dir in dir_list.items():
        setattr(args, dir[0], args.rootdir / dir[1])
        if not getattr(args, dir[0]).exists(): getattr(args, dir[0]).mkdir()

    print( '--Setting up directories--')
    for dir in dir_list:
        print( f'{dir} directory: {getattr(args, dir)}' )
    print( f'{"-"*30}' )

    print( 'Mapping docs (XLS) in:', args.mapdir )
    print( 'Output (SQL, YML) going to:', args.modeldir )

    # Show existing XLS files in Mapdir
    for afile in args.mapdir.glob('*.xlsx'):
        print(f'\tMapping found: {afile}')


def write_snap_step1( table, snapdir, scd2_sql, modeldir, scd6_sql):
    snapfile = snapdir/(table+'_SNAPSHOT_STEP1.sql')
    print('\t-- Writing to', snapfile)
    with open( snapfile, 'w') as fw:
        fw.write( scd2_sql )

    #table_split,split=table.split('_STEP1')
    if table.startswith('DIM'):
        model_file = modeldir/(table+'.sql')
    else:
        model_file = modeldir/(table+'_alt.sql')

    print('\t   Writing to', model_file)
    with open( model_file, 'w' ) as fw:
        fw.write( scd6_sql )
    # print( '========== done' )


def write_scd_step2( args, tgt_table, scd_dict ):
    modeldir = args.modeldir
    
    scd1_cols = scd_dict[ 'scd1' ] 
    exclude_col=[ 'UNIQUE_ID_KEY', 'CURRENT_INDICATOR', 'EFFECTIVE_DATE', 'END_DATE', 'LOAD_DATETIME', 'UPDATE_DATETIME', 'EFFECTIVE_TIMESTAMP', 'END_TIMESTAMP',
    'LOAD_BATCH_ID', 'UPDATE_BATCH_ID', 'PRIMARY_SOURCE_SYSTEM' ]

    scd1_cols = [ col.upper() for col in scd1_cols ]
    exclude_col = [ col.upper() for col in exclude_col ]

    # determine which columns are in scd1 but not in exclude_col without using a set
    scd1_remaining_cols = [ col for col in scd1_cols if col not in exclude_col ]

    scd1_str = ', '.join( scd1_remaining_cols )
    keycol = scd_dict[ 'KEY' ]
    
    src_table = tgt_table.replace( 'DIM_', 'DSV_' )

    # prefix = ''' {{ config(schema = 'EDW_STAGING', tags=["dim"] ) }}\n ----SRC LAYER----\nWITH\n'''
    prefix = '''\n ----SRC LAYER----\nWITH\n'''

    sql1 = prefix+f'''SCD1 as ( SELECT {scd1_str} , {keycol}    \n\t--, '1901-01-01' as DBT_VALID_FROM, '2099-12-31' as DBT_VALID_TO\n\tfrom      {{{{ ref( '{src_table}') }}}} )'''
    sql2 = f'''SCD2 as ( SELECT *    \n\tFROM      {{{{ ref('{tgt_table}_SNAPSHOT_STEP1') }}}} )'''
    sql3 = f'''FINAL as ( SELECT * 
            FROM  SCD2 
                INNER JOIN SCD1 USING( {keycol} )  )\nselect * from FINAL'''

    all_sql_str = ',\n'.join([ sql1, sql2, sql3 ])

    final_file = modeldir/(tgt_table+'_SCDALL_STEP2.sql')

    print( '\t   Writing to', final_file )
    with open( final_file, 'w' ) as fw:
        if not args.yml: fw.write( all_sql_str )  
    # print('========== done')



def process_unified( args, xls ):
    print(f'\t-- UNIFIED MAPPING: {xls.name}')
    src_file = xls.name.split('.')[0]
    layers = ['STG','DST','DSV','DIM']  # FCT is not currently part of uniified, but could be
    schemas = {'STG':'STAGING', 'DST':'STAGING', 'DSV':'STAGING', 'DIM':'EDW_STAGING_DIM', 'FCT':'_MART' }
    addtl = ['DML','DDL','Alter']

    # read the XLS file and list the found sheets
    # use pandas to read the xls file
    # wb = pd.ExcelFile(xls)
    # get the sheet names using pandas
    # found_sheets = wb.sheet_names

    wb = openpyxl.load_workbook(xls)
    found_sheets = wb.sheetnames

    print(f'Sheets found: {found_sheets}\n\n')

    for this_layer in layers:
        print(f'\t-- Attempting to Process {this_layer} Layer of UNIFIED mapping: {xls.name}')
        tables_tab  = f'{this_layer} Tables'
        columns_tab = f'{this_layer} Columns'
        schema = schemas[ this_layer ]
        table_name = f'{src_file.replace("UNIFIED",this_layer )}'

        if tables_tab in found_sheets:
            print(f'\t\t-- Processing {tables_tab}  in: {xls.name}' )
            tables  = process_tables( args, wb, tables_tab )

        if columns_tab in found_sheets:
            print(f'\t\t-- Processing {columns_tab} in: {xls.name}' )
            columns = process_columns( args, wb, columns_tab )

        # BUILD EACH LAYER
        if this_layer == 'STG':
            pass
            print(f'STG table: {table_name} ')  # Should be the XLS filename
            build(args, args.modeldir, table_name, tables, columns)

        elif this_layer == 'DST':
            print(f'\nDST table: {table_name} ')  # Should be the XLS filename
            build(args, args.modeldir, table_name, tables, columns)

        elif this_layer == 'DSV':
            print(f'\nDSV table: {table_name} ')  # Should be the XLS filename
            build(args, args.modeldir, table_name, tables, columns)

        elif this_layer == 'DIM':
            print(f'\nDIM table: {table_name} ')  # Should be the XLS filename
            # Need to check for SCD1, SCD2 columns

            # Need to process Alter tab for constraints
            alter_sql = process_alter( args, wb, 'Alter' )
            # Need to process DML tab for INSERT statements
            dml_block = process_dml( args, wb, 'DML' )

            # Combine the alter/dml to build the configuration
            config = get_config( cur_layer= table_name, alter_sql= alter_sql, dml_block= dml_block )

            # Determine if there are any SCD columns
            is_scd = False
            for col in columns.values():
                if col.get('SCD','') == None:
                    is_scd = True
                    break
            
            scd_dict = get_scdinfo( schema, table_name, columns )
            if is_scd:
                scd2_sql, scd6_sql = make_scd( schema, table_name, scd_dict )
                scd6_sql = config + scd6_sql
            else:
                scd6_sql = config

            write_snap_step1( table_name, args.modeldir, scd2_sql, args.modeldir, scd6_sql )
            write_scd_step2( args, table_name, scd_dict )

            # print( config_sql )
            # build(args, args.modeldir, table_name, tables, columns, alter_sql= alter_sql, dml_block= dml_block, config_sql = config_sql )
            build(args, args.modeldir, table_name, tables, columns )




def build_scd2( schema, table, cols, keycol ):
    schema = 'EDW_STAGING_SNAPSHOT'
    all_cols = cols + [ keycol ]
    
    
    cols_str = ', '.join( all_cols )
    src_table = table.replace( 'DIM_' , 'DSV_' )
    snapshot_sql=f'''
    
    {{%% snapshot %s_SNAPSHOT_STEP1 %%}}

    {{{{
        config(
        target_schema='%s',
        unique_key='%s',
        strategy='check',
        check_cols={cols},
        )
    }}}}

    select {cols_str} 
        from  {{{{ref('%s') }}}}

    {{%% endsnapshot %%}}
    '''%(table,schema,keycol,src_table)

    return snapshot_sql

def build_scd6 ( schema, table, keycol, scd1_cols = [], scd2_cols = [], scd3_cols = [], scd0_cols = [] ):
    keycol = keycol.upper()
    if not keycol: keycol = 'UNIQUE_ID_KEY'

    exclude_col = ['UNIQUE_ID_KEY', 'CURRENT_INDICATOR', 'EFFECTIVE_DATE', 'END_DATE', 'LOAD_DATETIME', 
    'UPDATE_DATETIME', 'EFFECTIVE_TIMESTAMP', 'END_TIMESTAMP',
    'LOAD_BATCH_ID', 'UPDATE_BATCH_ID', 'PRIMARY_SOURCE_SYSTEM']

    start_sql = f'''\n\n WITH  SCD AS ( \n\tSELECT  {keycol}'''

    sql2 = []
    for col in scd2_cols:
        if col in exclude_col: continue
        sql = f'''\n {col}'''
        sql2.append( sql )

    sql3 = []
    for col in scd3_cols:
        if col in exclude_col: continue
        sql = f'''\n\t, lag({col}) over  
           (partition by {keycol} 
             order by dbt_updated_at
        ) as {col}'''
        sql3.append(sql)

    sql1 = []
    for col in scd1_cols:
        if col in exclude_col: continue
        sql=f'''\n\t, last_value({col}) over 
            (partition by {keycol} 
              order by dbt_updated_at 
                ROWS BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING
            ) as {col}'''
        sql1.append(sql)
    
    sql0 = []
    for col in scd0_cols:
        if col in exclude_col: continue
        sql=f'''\n\t, first_value({col}) over 
            (partition by {keycol} 
              order by dbt_updated_at 
                ROWS BETWEEN UNBOUNDED PRECEDING AND UNBOUNDED FOLLOWING
            ) as {col}'''
        sql0.append( sql )

    end_sql = f'''\n\t    , DBT_VALID_FROM AS EFFECTIVE_TIMESTAMP
    \t, DBT_VALID_TO   AS END_TIMESTAMP
    \t,
    \t  CASE WHEN CAST(DBT_VALID_FROM AS DATE) = '1901-01-01' then cast( DBT_VALID_FROM AS DATE )
    \t    WHEN cast( DBT_VALID_FROM AS DATE ) <> '1901-01-01' THEN dateadd(day,1,CAST(DBT_VALID_FROM AS DATE))
    \t  else cast( DBT_VALID_FROM AS DATE ) end as EFFECTIVE_DATE
    \t, cast( DBT_VALID_TO AS DATE ) as END_DATE
    \t, CASE WHEN DBT_VALID_TO IS NULL THEN 'Y' ELSE 'N' END AS CURRENT_INDICATOR'''
    #CAST(DBT_VALID_FROM AS DATE) as EFFECTIVE_DATE\n,
    sql1[ 0 ] = start_sql + sql1[ 0 ]
    scd_all = sql1 + sql2 + sql3 + sql0
    #+ end_sql

    scd_str = ' '.join( scd_all )
    scd_str = scd_str.strip(',')

    scd_str = scd_str + end_sql + '\n'+"     \n\tFROM {{ ref('" + table + "_SCDALL_STEP2') }})\n\n"
    scd_str += 'select * from SCD\n'

    return scd_str

def get_scdinfo( schema, table, cols_dict):
    '''
    {'Order#': None, 'SCD TYPE': None, 'Staging Layer Column Name': None, ' 'CPI': None, 
    'Unique': None, 'Not Null': None, 'Test Expression': None, 'Accepted Values': None}
    '''
    scd_dict = {
        'table': table,
        'schema': schema,
        'KEY':'',
        'scd1':[], 'scd2':[], 'scd3':[], 'scd0':[],
    }

    keycol = 'KEY'
    for row in cols_dict.values():
        if '1' in str(row.get( 'SCD' )).split(','):
            scd_dict[ 'scd1' ].append( row.get( 'Staging Layer Column Name' ))
        if '2' in str(row.get( 'SCD' )).split(','):
            scd_dict[ 'scd2' ].append( row.get( 'Staging Layer Column Name' ))
        if '3' in str(row.get( 'SCD' )).split(','):
            scd_dict[ 'scd3' ].append( row.get( 'Staging Layer Column Name' ))
        if '0' in str(row.get( 'SCD' )).split(','):
            scd_dict[ 'scd0' ].append( row.get( 'Staging Layer Column Name' ))
        if row.get( 'SCD' ) == 'KEY':
            scd_dict[ 'KEY' ] = row.get( 'Staging Layer Column Name' ).upper() 
            
    return scd_dict

def make_scd( schema, table_name, scd_dict):
    '''
    scd={
        'table':'STG_POLICYTEST',
        'schema':'X10057301',
        'scd1':['POLICY_ID','CLAIM_ID'],
        'scd2':['PTYPE'],
        'scd3':['OWNER_NAME','ADDR'],
        'keycol':'POLICY_ID'
    }
    '''

    keycol = 'KEY'

    #print(scd_dict)
    #cols=scd_dict['scd2']+scd_dict['scd3']+scd_dict['scd0']
    combined_scd_cols = scd_dict[ 'scd2' ] + scd_dict[ 'scd3' ] + scd_dict[ 'scd0' ]

    scd2_sql = build_scd2( schema, table_name,  combined_scd_cols, scd_dict[ keycol ])

    scd6_sql = build_scd6( schema, table_name,  scd_dict[ keycol ],
        scd1_cols = scd_dict[ 'scd1' ], scd2_cols = scd_dict[ 'scd2' ], scd3_cols = scd_dict[ 'scd3' ], scd0_cols = scd_dict[ 'scd0' ])

    return scd2_sql, scd6_sql


def process_tables( args, wb, tables_tab ):
    print(f'\t\t-- {tables_tab} Tables ...' )
    # read the tables tab and convert to a dictionary
    expected_cols = ['Source Schema', 'Source Table', 'Alias', 'Filter Conditions' ,'Parent Join Number', 'Parent Table Join' ,'Child Table Join', 'Join Type']
    missing_cols, extra_cols, found_cols = [],[],[]

    ws = wb[tables_tab]
    tables_dict = {}

    # Check for extra columns
    for col in ws.iter_cols(min_row=1, max_row=1, values_only=True):
        for cell in col:
            if cell != '':
                found_cols.append(cell)

    # Check for missing columns and convert to a list
    missing_cols=[x for x in expected_cols if x not in found_cols]
    extra_cols=[x for x in found_cols if x not in expected_cols]

    # Summarize the results
    if len(extra_cols) > 0:
        print(f'\t\t\t-- Found Extra columns: {extra_cols}')
    if len(missing_cols) > 0:
        print(f'\t\t\t-- Missing columns: {missing_cols}')

    # Read the table rows and add to the dictionary
    df = pd.DataFrame(ws.values)

    # Use the 'Alias' to define the dictionary key and add all other column header/values as pairs
    '''Index(['Source Schema', 'Source Table', 'Alias', 'Filter Conditions',
       'Parent Join Number', 'Parent Table Join', 'Child Table Join',
       'Parent Table Join2', 'Child Table Join2', 'Join Type'],
      dtype='object', name=0)'''

    col_lookup = {}
    for i, col in enumerate(found_cols):
        col_lookup[col] = i+1

    for row in df.itertuples():

        if row.Index == 0: continue
        if row[1] != None:
            rown = row.Index
            tables_dict[rown] = {}
            for col in found_cols:
                # Add dicionary entry for each column using the col_lookup dictionary to get the column number
                tables_dict[rown][col] = row[ col_lookup.get(col) ]
                # add the missing cols to the dictionary
                for col in missing_cols:
                    tables_dict[rown][col] = ''

    return tables_dict

def process_columns( args, wb, columns_tab ):
    print(f'\t\t-- {columns_tab} Columns ...' )
    # read the tables tab and convert to a dictionary
    expected_cols = ['Source Schema', 'Source Table', 'Source Column', 'Datatype', 'Automated Logic', 'Manual Logic' ,'Mapping Rule', 'Order#',	'Staging Layer Column Name', 'Staging Layer Datatype', 'CPI', 'Unique', 'Not Null',	'Remove Column', 'Test Expression', 'Accepted Values']
    dim_cols = ['Model Equality', 'Model Equal Rowcount', 'SCD']
    missing_cols, extra_cols, found_cols = [],[],[]
    cur_layer = columns_tab.split(' ')[0]

    if columns_tab.startswith('DIM'):
        expected_cols = expected_cols + dim_cols

    ws = wb[columns_tab]
    columns_dict = {}

    # Check for extra columns
    for col in ws.iter_cols(min_row=1, max_row=1, values_only=True):
        for cell in col:
            if cell != '':
                found_cols.append(cell)

    # Check for missing columns and convert to a list
    missing_cols=[x for x in expected_cols if x not in found_cols]
    extra_cols=[x for x in found_cols if x not in expected_cols]

    # Summarize the results
    if len(extra_cols) > 0:
        print(f'\t\t\t-- Found Extra columns: {extra_cols}')
    if len(missing_cols) > 0:
        print(f'\t\t\t-- Missing columns: {missing_cols}')

    # Read the table rows and add to the dictionary
    df = pd.DataFrame(ws.values)

    # Use the 'Order#' to define the dictionary key and add all other column header/values as pairs
    '''Index(['Source Schema', 'Source Table', 'Source Column', 'Datatype',
       'Automated Logic', 'Manual Logic', 'Mapping Rule', 'Order#',
       'Staging Layer Column Name', 'Staging Layer Datatype', 'CPI', 'Unique',
       'Not Null', 'Test Expression', 'Model Equality', 'Model Equal Rowcount',
       'Accepted Values', 'SCD'],'''
    # column order is based on the order of columns in the found_cols list
    # make a dictionary of the columns and their order
    col_lookup = {}
    for i, col in enumerate(found_cols):
        col_lookup[col] = i+1

    for row in df.itertuples():

        if row.Index == 0: continue
        rown = row.Index
        
        columns_dict[rown] = {}
        for col in found_cols:
            # Add dicionary entry for each column using the col_lookup dictionary to get the column number
            columns_dict[rown][col] = row[ col_lookup.get(col) ]
            # add the missing cols to the dictionary
            for col in missing_cols:
                columns_dict[rown][col] = ''

    # if the Staging Layer Column Name is blank, remove the row from the dictionary
    # Clean up extra rows created by the blank rows in the excel sheet
    blank_rows = []
    for rown, col in enumerate( columns_dict.values() ):
        if cur_layer == 'DST' and col.get( 'Staging Layer Column Name') == None:
            blank_rows.append(rown+1)
    for rown in blank_rows:
        del columns_dict[rown]

    return columns_dict 


def process_alter( args, wb, alter_tab ):
    print(f'\t\t-- Processing ALTER Statements ...' )

    ws = wb[alter_tab]
    alter_sql = []

    # read each row of the alter_tab and append to the alter_sql list if it's not blank
    for row in ws.iter_rows(min_row=0, values_only=True):
        if row[0].strip() != '':
            alter_sql.append(row[0])
    
    return alter_sql

def process_dml( args, wb, dml_tab ):
    print(f'\t\t-- Processing DML Statements ...' )

    ws = wb[dml_tab]
    dml_sql = []

    # read the dml_tab and append each row to the dml_sql list
    for row in ws.iter_rows(min_row=0, values_only=True):
        if row != '':
            dml_sql.append(row[0])

    return dml_sql

def get_config( cur_layer, alter_sql='', dml_block='' ):
    '''
    Builds the CONFIG dynamically based on the file name
        # {{ config(   only if...:
        # ADD: materialized = 'view' if the tgt_table starts with DSV
        # ADD: tags = [ "fact" ]  if the tgt_table starts with FACT%
        # Only add post_hook = (" if there are alter statements.
        # ") <- close post_hook 
        # }} <- close config
    '''
    print( '-- COMBINING ALTER/DML statements ')
    concat_config = '' 
    config = {}
    compileif = ['view','tags','alter']
    cur_layer = str(cur_layer)

    if cur_layer.startswith('DSV'):
        config['view'] = f"materialized = 'view'"
    elif cur_layer.startswith( 'FACT' ):
        config['tags'] = f'tags = [ "fact" ]'

    # Dimension, Fact and Factless Fact tables will get post hook (Alter statements for Referential Integrity)
    if cur_layer.startswith('DIM') or cur_layer.startswith('FACT') or cur_layer.startswith('FLF') :
        if len(alter_sql) > 1:
            alter_sql = '\n'.join(alter_sql)
            # replace all of the carriage returns with spaces and then join into a single string
            dml_block = [x.replace('\n',' ') for x in dml_block]
            dml_block = ' '.join(dml_block)
            config['alter'] = f'post_hook = ("\n{alter_sql}\n{dml_block}\n") '

    compile_cnt = 0
    for checkval in compileif:
        if checkval in config and compile_cnt == 0:
            concat_config = f'{{{{ config( \n{config.get(checkval)} '
            compile_cnt += 1
        elif checkval in config:
            concat_config += f'\n,\t {config.get(checkval)}'
            compile_cnt += 1
    if compile_cnt > 0:
        concat_config  += '\n) }}'

    config_sql = ''.join( concat_config) 

    return config_sql


def get_model_tests(row):
    '''
    https://github.com/fishtown-analytics/dbt-utils
    '''
    test_dict = {}
    test_cols = ['Unique', 'Model Equality', 'Model Equal Rowcount', 'Test Expression',  ]
    # Globally replacing the relationship with relationship_where

    tests = []
    add = {'severity': 'warn'}

    for atest in test_cols:
        if atest not in row:
            continue
        if not row[atest]:
            continue

        if atest == 'Unique':
            if not row.get(atest).lower() == 'unique' :
                test =   str(row[ atest ]).replace('COMPOSITE:','').strip()
                row[atest] = {
                    'dbt_utils.unique_combination_of_columns':
                    {
                        'combination_of_columns': [each.strip() for each in test.split(',')],   
                        'severity': 'warn'
                    }
                }
            else: continue

        if atest == 'Model Equality':
            # print(f' $$$$$$$$ FOUND Model Equality ->  {row[ atest ]}')               # DEBUG
            test =   str(row[ atest ]).strip()
            row[atest] = {
                'dbt_utils.equality': 
                {
                    'compare_model': f"ref( {row[ 'Staging Layer Column Name' ]} )",
                    'compare_columns': [each.strip() for each in test.split(',')]                    
                    # 'compare_columns': [ f'ref({row[ atest ]!r})' ]
                }
            }

        if atest == 'Model Equal Rowcount':
            # print(f' $$$$$$$$ FOUND Model Equal Rowcount -=>  {row[ atest ]}')        # DEBUG
            row[atest] = {
                'dbt_utils.equal_rowcount': 
                {
                    'compare_model': f'ref({row[ atest ]!r})'
                }
            }

        # Relationship moved to column test...

        if atest == 'Test Expression':
            if row.get('Staging Layer Column Name','x') in row.get('Test Expression','y'):
                # If true, then should be a model test, else column test
                test =   str(row[ atest ]).strip()
                row[atest] = {
                    'dbt_utils.expression_is_true':
                    {
                        'severity': 'warn',
                        'expression': test
                    }
                }
            else: continue
        tests.append(row[atest])

    return tests

def get_column_tests(row, table_name):
    '''
    https://github.com/fishtown-analytics/dbt-utils
    '''
    test_dict = {}
    test_cols = [ 'Unique', 'Not Null', 'Test Expression', 'Accepted Values', 'Relationship', 'Set Default']
    # 'Equal Rowcount',  'Mutually Exclusive Ranges', 'Unique Combination of Columns', 'Accepted Range']

    tests = []
    add = {'severity': 'warn'}

    for atest in test_cols:
        if atest not in row:
            continue
        if not row[atest]:
            continue

        # If a FLF or FACT build, do not include the has_default test.  Will require knowing the layer during build.
        if atest == 'Set Default':
            test = str( row[ atest ] ).strip()
            row['Default']= test
            if table_name.startswith('F'):
                row[atest]=''
            else:
                row[atest] = {
                    'has_default':
                    {
                        'severity': 'warn',
                        'hashval': [each.strip() for each in test.split(',')]
                    }
                }

        if atest == 'Unique':
            row[atest] = {
                'unique':
                {
                    'severity': 'warn'
                }
            }

        if atest == 'Not Null':
            if 'where' in row[atest]:
                # TODO: Fix quoting around where
                test = str( row[ atest ] ).upper().replace( 'WHERE','').strip()
                row[atest] = {
                    'dbt_utils.not_null_where':
                    {
                        'where': test
                    }
                }                
            else:
                row[atest] = {
                    'not_null':
                    {
                        'severity': 'warn'
                    }
                }

        if atest == 'Accepted Values':
            if 'not:' in row[atest]:
            # Split to break them into separate rows...
                test = str( row[ atest ] ).upper().replace( 'NOT:','').strip()
                row[atest] = {
                    'not_accepted_values':
                    {
                        'severity': 'warn',
                        'values': [each.strip() for each in test.split(',')]
                    }
                }
            else:
                test = str( row[ atest ] )
                row[atest] = {
                    'accepted_valids':
                    {
                        'ignore': '',
                        'severity': 'warn',
                        'values': [each.strip() for each in test.split(',')]
                    }
                }

        if atest == 'Test Expression':
            # If the column name is not listed, then it can be a column test, else model test
            if not row.get('Staging Layer Column Name','x') in str(row[ atest ]).strip():
                test =   str(row[ atest ]).strip()
                row[atest] = {
                    'expression_is_true_ignore':
                    {
                        'ignore': '',
                        'severity': 'warn',
                        'expression': test
                    }
                }
            else: continue

        if atest == 'Relationship':
            ignore_in = ''
            # Changed to a relationships_where instead to handle ignoring dummy value:  md5('-1111') 
            # Determine if the field is different (Based on TableName.ColumnName )
            if '.' in row[atest]:
                ref_table = row[ atest ].split('.')[0]
                ref_field = row[ atest ].split('.')[1]
                ref_ignore = row[ 'Staging Layer Column Name' ]
            else:
                ref_table = row[ atest ]
                ref_field = row[ 'Staging Layer Column Name' ]
                ref_ignore = row[ 'Staging Layer Column Name' ]

            if ref_field == 'DATE_KEY':
                ignore_in = " -1, -2, -3 "
            else: 
                ignore_in = " '40c5dea533476acdd01f7ef0e84de22f', 'fcbcdcb8f6b1c597c5fdc7a54cd321ae' "

            row[atest] = {
                'dbt_utils.relationships_where':
                {
                    'to': f'ref( {ref_table!r} )',
                    'field': f"{ref_field}",
                    'from_condition': f"{ref_ignore} not in ( {ignore_in} ) "
                }
            }

        if not row[atest] == '':
            tests.append(row[atest])

    return tests

def make_yml(table_name, tables, columns):
    '''
    version: 2

models:
    - name: orders
    tests:
    - dbt_utils.expression_is_true:
            expression: "colA + colB = totalAB"
    - dbt_utils.equality:
            compare_model: ref( 'other_table_name' )
    columns:
    - name: order_id
        tests:
            - unique
            - not_null
            - dbt_utils.expression_is_true:
                expression: "colA + colB = totalAB"
    - name: status
        tests:
            - accepted_values:
                values: [ 'placed', 'shipped', 'completed', 'returned' ]
    - name: REVENUE_CENTER_HKEY
        tests:
            - dbt_utils.relationships_where:
                to: ref('DIM_REVENUE_CENTER')
                field: REVENUE_CENTER_HKEY
                from_condition: REVENUE_CENTER_HKEY <> '40c5dea533476acdd01f7ef0e84de22f'
    '''

    modeltests = []
    for row in columns.values():
        modeltests += get_model_tests(row)
        
    # print( f'<<< Going in with modeltests: {modeltests}' )        # DEBUG

    schema_dict = {
        'version': 2,
        'models': [
            {'name': table_name,
             'tests': modeltests,
             'columns': [],
             }
        ]
    }
    
    # print( f'@@@ ADDED MODELTEST: {schema_dict}' )                  # DEBUG


    ymlcols = []
    cols_processed = set()
    for row in columns.values():
        col = row['Staging Layer Column Name']

        if not col:
            continue
        if col in cols_processed:
            continue
        cols_processed.add(col)

        tests = get_column_tests(row, table_name)

        if tests:
            ymlrow = {
                'name': col,
                'tests': tests,
            }
            ymlcols.append(ymlrow)

    y = {}
    y = schema_dict['models'][0]
    y['columns'] = ymlcols

    return schema_dict

def get_source_sql( target_table, cols, tables ):
    '''
    cols:
    1: {'Source Schema': 'STAGING', 'Source Table': None, 'Source Column': '(DERIVED)', 'Datatype': None, 'Automated Logic': None, 'Manual Logic': 'HASH: ADDRESS_LINE_1\n, ADDRESS_LINE_2\n, CITY_NAME\n, STATE_CODE\n, ZIP_CODE\n, COUNTY_NAME\n, ADMINISTRATION_ADDRESS_IND\n, SHIPPING_ADDRESS_IND\n, PRIMARY_ADDRESS_IND\n, EXAM_LOCATION_IND', 'Mapping Rule': 'EXTRACT THE DISTINCT VALUES OF KEY COLUMNS.\nTable Filter: ADDRESS_LINE_1 IS NOT NULL', 'Order#': 1, 'Staging Layer Column Name': 'UNIQUE_ID_KEY', 'Staging Layer Datatype': 'CHAR(32)', 'CPI': None, 'Unique': 'Unique', 'Not Null': 'not_null', 'Remove Column': None, 'Test Expression': None, 'Accepted Values': None}, 
    '''
    current_layer = target_table.split('.')[0]
    # make a copy of cols to update
    columns = cols.copy()

    # Initialize all of the variables to avoid none type errors
    for idx, row in enumerate( columns.values() ):
        #determine derived first
        src_col, hasname, mlogic = '','',''
        if row.get('Source Column',''):
            src_col = row.get('Source Column','')
        src_sql = src_col
        if src_col:
            hasname = src_col.replace('(DERIVED)', '').strip()
        tgt_col = src_col
        final_col = src_col
        if row.get('Manual Logic', ''):
            mlogic = row.get('Manual Logic', '').upper()

        # if row.get('Staging Layer Column Name') == None or row.get('Remove Column'): break

        if  src_col.startswith('(DERIVED)'):
            #If no Source Table, then only display in FINAL layer.
            row['Derived Source'] = hasname
            if row['Derived Source'] == '':
                row['Derived Source'] = row.get('Staging Layer Column Name')
                if hasname =='': hasname = row.get('Staging Layer Column Name')
            dtype = row.get('Datatype')

            if  dtype == None:
                mlogic = row.get('Manual Logic').upper()
                if 'HASH' in mlogic or 'COMPOSITE' in mlogic:
                    hashcol = ''
                    hashcol = mlogic.replace('HASH:', '').replace('COMPOSITE:', '').replace('\n','').replace(' ','').strip()
                    temp1 = hashcol.split(',')
                    joincol = (', ').join(temp1)
                    hashcol = ("','").join(temp1)
                    col1 = f"{{ dbt_utils.generate_surrogate_key ( [ '{hashcol}' ] ) }}" 
                    surrogate = f"{{{col1}}}"
                # If there is manual logic and it's not a HASH/COMPOSITE, use as is
                elif mlogic:
                    src_sql = mlogic.strip()
                else:
                    src_sql = src_col
            elif  dtype == 'DATE':
                if hasname == '': hasname = src_col
                src_sql = f"CASE WHEN {hasname} is null then '-1' \n\t\t\tWHEN {hasname} < '1901-01-01' then '-2' \n\t\t\tWHEN {hasname} > '2099-12-31' then '-3' \n\t\t\tELSE regexp_replace( {hasname}, '[^0-9]+', '') \n\t\t\t\tEND :: INTEGER {' '*40}"
            else:
                if hasname == '': hasname = src_col
                src_sql = src_col
                

            set_default = row.get('Default')
            if set_default and not ( src_col == 'NULL' or src_col == ''):
                if 'HASH' in mlogic or 'COMPOSITE' in mlogic:
                    if ',' in joincol:
                        src_sql = f"CASE WHEN  nullif(array_to_string(array_construct_compact( {joincol} ),''),'') is NULL \n\t\t\tthen MD5( '{set_default}' ) ELSE {surrogate} \n\t\t\t\tEND {' '*40}"
                    else:
                        src_sql = f"CASE WHEN {joincol} is NULL \n\t\t\tthen MD5( '{set_default}' ) ELSE {surrogate} \n\t\t\t\tEND {' '*40}"
                elif hasname  == '':
                    src_sql = f"MD5( '{set_default}' )"
                elif hasname != '':
                    src_sql = hasname
                    src_sql = f"COALESCE( {hasname},  MD5( '{set_default}' ) ) "
                else:
                    src_sql = f"COALESCE( {src_col}, MD5( '{set_default}' ) ) "
            else:
                if 'HASH' in mlogic or 'COMPOSITE' in mlogic:
                    src_sql = f" {surrogate} "

        # If not DERIVED, then use the Source Column
        # TEXT gets NULLIF and TRIM for STG layer
        if row.get('Datatype') == 'TEXT':
            if current_layer == 'STG':
                src_sql = f"NULLIF( TRIM( {src_col} ),'' )"

        # Automated logic overlaps SRC_SQL
        logic = row.get('Automated Logic')

        if logic:
            logic = logic.lower()
            indent = True
            if 'cast date' in logic:
                src_sql = f'cast( {src_col} as DATE )'
            elif 'cast text' in logic:
                src_sql = f'cast( {src_col} as TEXT )'
            elif 'upper' in logic and 'trim' in logic:
                src_sql = f'upper( trim( {src_col} ))'
            elif logic == 'upper':
                src_sql = f'upper( {src_col} )'
            else:
                print(f'!! Unsupported Operation !! <{logic}>')

        if 'DERIVED' in src_col and src_sql != '' :
            tgt_col = hasname
        elif 'DERIVED' in src_col and mlogic != '':
            src_sql - mlogic
            tgt_col = hasname
        elif hasname != src_col:
            tgt_col = row.get('Derived Source')
        else:
            # Not derived
            tgt_col = src_col

        if row.get('Staging Layer Column Name'):
            final_col = row.get('Staging Layer Column Name')

        # Format the sql to fit on a line if possible
        if len(src_sql)>50:
            src_sql = f"\n\t\t {src_sql:<50}\n{' ':<60} as {tgt_col:>50}"
        else:
            src_sql = f"\n\t\t {src_sql:<50} as {tgt_col:>50}"


        # if row.get('Source Table') == None and row.get('Order#') != 1:
        #     row['SQL'] = src_sql
        #     row['Rename'] = f"\n\t\t {tgt_col:<50} as {final_col:>50}"
        # else:
        row['SQL'] = src_sql
        row['Rename'] = f"\n\t\t {tgt_col:<50} as {final_col:>50}"

    # print('='*80)
    # for row in columns.values():
        # print(row.get('SQL'))
    # print('='*80)
    

    return columns



def build( args, modeldir, tgt_table, tables, columns, alter_sql='', dml_block='', config_sql=''):
    yml = make_yml( tgt_table, tables, columns )
    alias = {}
    all_sql = []
    newvalue = ''

    for thisrow in tables.values():
        thisrow_dict = {}
        for key, value in thisrow.items():
            if not value: value = ''
            if not key: key = ''
            thisrow_dict[key] = value
        thisrow = thisrow_dict

        for key, value in sorted( thisrow.items() ):
            if not key: continue
            if key not in ['Source Table', 'Alias', 'Source Schema']: continue
            if value not in alias:
                if 'Alias' in key:
                    alias[ value ] = []
                    newvalue = value
            if 'Source Table' in key:
                alias[ newvalue ] = value
            if 'Source Schema' in key:
                alias[ 'Schema' ] = value
            if 'Set Default' in key:
                alias[ 'Set Default' ] = value 

    columns = get_source_sql( tgt_table, columns, tables )

    # Build each layer of the SQL
    ################################
    # SRC LAYER
    config = get_config( tgt_table, alter_sql, dml_block )
    config_sql = config_sql if config_sql else config
    all_sql.append( config_sql )
    all_sql[-1] += '\n\n---- SRC LAYER ----\nWITH\n' 

    src_sql = build_src_layer( args, tables )
    all_sql.append( src_sql )
    # all_sql[ 0 ] = config_sql +  '\n\n---- SRC LAYER ----\nWITH\n' + all_sql[ 0 ]

    ################################
    # LOGIC LAYER
    all_sql[ -1 ] += '\n\n---- LOGIC LAYER ----\n'
    logic_sql = build_logic_layer( columns, tables)
    all_sql.append( logic_sql )

    ################################
    # RENAME LAYER
    all_sql[ -1 ] += '\n\n---- RENAME LAYER ----\n'
    rename_sql = build_rename_layer( columns )
    all_sql.append( rename_sql )

    ################################
    # FILTER LAYER
    all_sql[ -1 ] += '\n\n---- FILTER LAYER ----\n'
    filter_sql = build_filter_layer( tables )
    all_sql.append( filter_sql )

    ################################
    # JOIN LAYER
    # Includes the final select
    all_sql[ -1 ] += '\n\n---- JOIN LAYER ----\n'
    join_sql = build_join_layer( tables, columns )
    all_sql.append( join_sql )

    # print( ''.join( all_sql ) )
    # print('='*80)    
    sql_str = '\n'.join( all_sql )

    # If nothing is populated for the Logic Layer then there is nothing to output
    if join_sql == '':
        print( f'\t*** SKIPPING {tgt_table} ***\n\n' )
    else:
        ################################
        # Write out the final finals...
        with open(args.modeldir/(tgt_table + '.yml'), 'w') as fw:
            print( f'\n~~~~> WRITING YML:\n\t\t{tgt_table}.yml' )                          # DEBUG
            fw.write(yaml.dump(yml))

        ##### Output SQL for all other non-DIM sql (STG, DST, DSV)
        if not tgt_table.startswith( 'DIM' ):
            print( f'\n~~~~> WRITING SQL:\n\t\t{tgt_table}.sql\n\n' )                          # DEBUG
            with open(args.modeldir/(tgt_table + '.sql'), 'w') as fw:
                fw.write(sql_str)

        # print(f'\n\n===== CURRENT SQL =\n{sql_str}')                   # DEBUG


def build_src_layer(args, tables):
    '''
    {1: {'Source Schema': 'STAGING', 'Source Table': 'STG_DEP_MEMBERS', 'Alias': 'M', 'Filter Conditions': None, 'Parent Join Number': 2, 'Parent Table Join': None, 'Child Table Join': None, 'Join Type': None}, 2: {'Source Schema': 'STAGING', 'Source Table': 'STG_DEP_ADDRESS', 'Alias': 'A', 'Filter Conditions': None, 'Parent Join Number': 2, 'Parent Table Join': 'CONCAT(M.PEACH_BASE_NUMBER, M.PEACH_SUFFIX_NUMBER)', 'Child Table Join': 'CONCAT(A_PEACH_BASE_NUMBER, A_PEACH_SUFFIX_NUMBER)', 'Join Type': 'left join'}, 3: {'Source Schema': 'STAGING', 'Source Table': 'STG_DEP_US_ZIPCODE_LIST', 'Alias': 'Z', 'Filter Conditions': None, 'Parent Join Number': 1, 'Parent Table Join': 'A.ZIP_CODE', 'Child Table Join': 'Z.ZIPCODE', 'Join Type': 'left join'}}
    '''
    print(f'==== Build SOURCE SQL .')
    all_sql = []
    act_table = ''

    # print( '--- BUILDING SOURCE SQL' )
    for tbl in tables.values():
        table = tbl.get('Alias')
        if table != None:
            act_table = tbl.get('Source Table')
        sql = f"SRC_{table:<14} as ( SELECT *     FROM     {{{{ ref( {act_table!r} ) }}}} ),\n"
        all_sql.append(sql)

    # Add the SQL version as comments to run in SF
    all_sql.append('\n/*\n')

    for tbl in tables.values() :
        table = tbl.get('Alias')
        schema = tbl.get('Source Schema')
        if table != None:
            act_table = tbl.get('Source Table')
        sql = f"SRC_{table:<14} as ( SELECT *     FROM     {schema}.{act_table} ),\n"
        all_sql.append(sql)

    all_sql.append('\n*/')
    all_sql = ''.join(all_sql)

    return all_sql

def build_logic_layer( columns, tables ):
    alias_list = []
    final_sql = []
    sql = ''
    print(f'==== Build LOGIC  SQL ..')          # DEBUG

    for row in tables.values():
        tbl = row.get('Alias')
        if tbl and tbl not in alias_list:
            alias_list.append( tbl )

    # initialize tables_cols for each alias as an empty list
    table_cols = { alias:{} for alias in alias_list }        
    
    for tbl in alias_list:
        src_tbl = f'SRC_{tbl}'
        tgt_tbl = f'LOGIC_{tbl}'
        table_cols[tbl]['src_tbl'] = src_tbl
        table_cols[tbl]['tgt_tbl'] = tgt_tbl
        table_cols[tbl]['cols'] = []

    for row in columns.values():
        row_sql = ''
        tbl = row.get('Source Table')
        if tbl:
            row_sql = row.get('SQL','')
            table_cols[tbl]['cols'].append( row_sql )
    
    for tbl in alias_list:
        src_tbl = table_cols[tbl]['src_tbl']
        tgt_tbl = table_cols[tbl]['tgt_tbl']
        sql += f"\n, {tgt_tbl} as ( \n\tSELECT " 
        for idx,col in enumerate(table_cols[tbl]['cols']):
            if idx == 0:
                sql += f'\t  {col}'
            else:
                sql += f'\t, {col}'
        sql += f"\n\t from {src_tbl} )\n"

    logic_str  = ''.join(sql)

            # if row.get('Source Table','') != tbl:
            #     row_sql,sql = '',''
            # else:
            #     if first == True:
            #         sql = ' '+ row_sql
            #         first = False
            #     else:
            #         sql = ', '+row_sql  if row_sql !='' else ''           
            #         first = False
            #     table_cols.append( sql )
            
        # cols_str = ' '.join(table_cols)
        # sql = f'''\n{tgt_tbl} as ( SELECT {cols_str} \n\t\tFROM SRC_{tbl}
            # )'''

    final_sql.append( logic_str )

    # print(' '.join(final_sql))

    return ',\n'.join( final_sql )

def build_rename_layer( columns ):
    alias_list = []
    final_sql = []
    print(f'==== Build RENAME SQL ...')          # DEBUG

    for row in columns.values():
        tbl = row.get('Source Table')
        if tbl and tbl not in alias_list:
            alias_list.append( tbl )
    
    for tbl in alias_list:
        src_tbl = f'LOGIC_{tbl}'
        tgt_tbl = f'RENAME_{tbl}'
        table_cols = []
        first = True
        for row in columns.values():
            if row.get('Source Table') == tbl:
                if first == True:
                    sql = ' '+ row.get('Rename','')
                    first = False
                else:
                    sql = ', '+row.get('Rename','')                

                # if row.get('Remove Column') == None:
                table_cols.append( sql )
            
        cols_str = ' '.join(table_cols)
        sql = f'''\n, {tgt_tbl} as ( SELECT {cols_str} \n\t\tFROM LOGIC_{tbl}
            )'''

        final_sql.append( sql )

        # print(' '.join(final_sql))

    return '\n'.join( final_sql )

def build_filter_layer(tables):
    print(f'==== Build FILTER SQL ....')          # DEBUG

    filter_sql = []
    for row in tables.values():
        if not row['Filter Conditions']:
            filters = []
        else:
            filters = row['Filter Conditions'].split(';')
        filters_str = ''
        if filters:
            filters_str = f"\n{'WHERE ':>50}"
            filters_str += ' AND '.join(filters)

        alias = row['Alias']
        from_table = alias or row['Source Table']

        if not alias:
            if not row['Source Table']:
                continue
        source_table = f'RENAME_{from_table}'

        sql = f'FILTER_{from_table:<30} as ( SELECT * FROM    {source_table} {filters_str}  ),'
        filter_sql.append(sql)

    return ',\n'.join( filter_sql )

def build_join_layer(tables, columns):
    '''Includes both Tables and Columns because it will generate the final select statement'''
    join_sql = []
    if len(tables) == 0: 
        print(' --- NOTHING TO PROCESS ---')
        return ''

    #{ 'POLICY_CLAIMS':'P', 'STG_POLICYTEST':'T', 'POLICY_OWNERS':'O', 'POLICY_OWNERS_ADDRESS':'POA' }
    final_alias = ''
    aliases = {}
    for row in tables.values():
        if not row['Source Schema']:
            continue
        alias = row['Alias']
        aliases[row['Source Table']] = alias
        if not row['Join Type']:
            final_alias = alias

    join_lvls = build_join_levels(tables)
    all_sql = []

    made_cte = set()
    # print( f'\n<< JOIN LEVELS: {join_lvls}\n')                               # DEBUG
    # All join_lvls should have a value.  If not, uncomment the above and see where the issue is.

    for lvl in sorted(join_lvls):
        # print( f'<< Currently on lvl: {lvl}')                               # DEBUG
        parent_table, join_sql = do_join( made_cte, join_lvls[ lvl ] )
        made_cte.add( parent_table )
        join_sql_str = '\n\t\t'.join(join_sql)
        
        if 'COALESCE(' in parent_table.upper():
            wascoalesce = parent_table.upper().replace('COALESCE(','').strip()
            sql = f'{wascoalesce} as ( SELECT * \n\t\t\t\tFROM  FILTER_{wascoalesce}\n{join_sql_str} )'
        else:
            sql = f'{parent_table} as ( SELECT * \n\t\t\t\tFROM  FILTER_{parent_table}\n{join_sql_str} )'
        all_sql.append(sql)

    ########################################
    # Build the final select statement
    select_cols = []
    for row in columns.values():
        if not row['Remove Column']:
            if row['Source Table'] == None:
                select_cols.append( row.get('SQL') )
            else:
                select_cols.append( row['Staging Layer Column Name'] )
    
    if len(select_cols) == len(columns):
        select = '*'
    else:
        select = ''
        for col in select_cols:
            if col == select_cols[ 0 ]:
                select += f'\n\t\t  {col}'
            else:
                select += f'\n\t\t, {col}'

    final_sql = f'\nSELECT {select} \nFROM {final_alias}'

    if not all_sql:
        sql = f' JOIN_{alias:<10}  as  ( SELECT * \n\t\t\t\tFROM  FILTER_{alias} )'
        all_sql = [sql + '\n' + f' SELECT * FROM  JOIN_{alias}']
    else:
        all_sql[-1] = all_sql[-1] + final_sql

    all_sql_str = ',\n'.join(all_sql)

    # print( f'\n\n=== Build JOIN SQL\n{all_sql}')                   # DEBUG
    return all_sql_str


def do_join(made_cte, tables ):
    '''
        select *
            from FPO
                LEFT JOIN FPAD USING( POLICY_ID )
                LEFT JOIN FPAD2 on FPO.POLICY_ID = FPAD2.POLICY_IDX;
    '''
    join_sql = []
    parent_join_list = []
    # The sorted_rows group items together by Parent Join Alias with INNER JOINS being outdented
    sorted_rows = []
    sorted_rows = sorted(tables, key = itemgetter( 'Parent Join Alias' ))
    
    # for row in table_rows:
    for row in sorted_rows:
        join_type = row[ 'Join Type' ]
        if not join_type:
            continue
        join_type = join_type.replace('_', ' ').upper()
        source_table = row[ 'Alias' ]
        parent_join = row[ 'Parent Table Join' ]
        child_join = row[ 'Child Table Join' ]
        parent_table = parent_join.split('.')[ 0 ]

        using_join = ''
        parent_join_col = parent_join.split('.')[ -1 ]
        child_join_col = child_join.split('.')[ -1 ]

        if parent_join_col == child_join_col:
            using_join = f' using( {parent_join_col} ) '

        if 'coalesce' in parent_table:
            tempparent = parent_join.upper().replace('COALESCE(','').strip()
            parent_join = f'coalesce( FILTER_{tempparent}'
        else:
            parent_join = f'FILTER_{parent_join}'
            
        if source_table in made_cte:
            sql = f'\t\t\t\t{join_type} {source_table} ON  {parent_join} = {child_join} '
            if using_join:
                sql = f'\t\t\t\t{join_type} {source_table} {using_join} '
        else:
            if parent_table in parent_join_list and not join_type.lower().startswith('inner'):
                sql = f'\t\t\t\t\t\t{join_type} FILTER_{source_table} ON  {parent_join} =  FILTER_{child_join} '
            else:
                sql = f'\t\t\t\t{join_type} FILTER_{source_table} ON  {parent_join} =  FILTER_{child_join} '
                if using_join:
                    sql = f'\t\t\t\t{join_type} {join_type}  FILTER_{source_table} {using_join} '
                parent_join_list.append( parent_table )

        join_sql.append( sql )

        # print( f'\n\n==== DO JOIN\nPT:{parent_table}\nSQL:{join_sql}')
    return parent_table, join_sql

def build_join_levels(tables):
    lvls = {}
    for row in tables.values():
        if not row[ 'Parent Table Join' ]:
            continue
        lvl = row[ 'Parent Join Number' ]
        if not lvl: lvl = 0
        if lvl not in lvls:
            lvls[ lvl ] = []
        # Add Parent Join Alias to determine if it's a sub join
        row[ 'Parent Join Alias' ] = row[ 'Parent Table Join'].split('.')[0]
        lvls[ lvl ].append(row)

    return lvls


def main():
    count = 0
    plural, scd6_sql, dml_block, dml_sql  = '', '', '',''
    print('\n' + '-'*30)

    #Process each xls file in the directory
    # for xls in args.mapdir.iterdir():
    for xls in args.mapdir.glob('*.xls*'):
        print(f'\t<< Processing: {xls}')
        if xls.name.startswith('XDIM_'):
            print(f'\t\tNOT STAGE MAPPING -- skipping: {xls.name}')
            continue
        elif 'xls' not in xls.suffix.lower():
            print(f'\t-- Skipping non-XLS file: {xls}')
            continue

        print(f'-'*30)
        print(f"Currently processing {os.path.basename(xls)}")

        # Determine if it is UNIFIED or not
        if 'UNIFIED' in xls.name.upper():
            print(f'\t-- UNIFIED MAPPING: {xls.name}')
            process_unified( args, xls )
        else:
            print(f'\t-- Standard MAPPING: {xls.name}')
            # process_std_map( args, xls )
            layer = xls.name.split('_')[0]
            
        
        count += 1

    plural = 's.' if count > 1 else '.'

    print(f'Processed: {count} file{plural}')
    if count > 0 : print(f'Models written to: {args.modeldir}')

    input('[ Press Enter to exit ] ')


if __name__ == '__main__':
    args = process_args()
    setup(args)

    main()
