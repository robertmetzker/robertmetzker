from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

wb = Workbook()
ws = wb.active
ws.title = "Grades"

sample= [
{'resource_type': 'test', 'depends_on': {'macros': ['macro.dbt.test_unique'], 'nodes': ['model.edw.STG_POLICY_PROFILE']}, 'unique_id': 'test.edw.unique_STG_POLICY_PROFILE_PLCY_PRFL_ID.72703b3b12', 'original_file_path': 'models\\STAGING\\STG_POLICY_PROFILE.yml', 'name': 'unique_STG_POLICY_PROFILE_PLCY_PRFL_ID'}, 
{'resource_type': 'test', 'depends_on': {'macros': ['macro.dbt.test_unique'], 'nodes': ['model.edw.STG_POLICY_STATUS_REASON_HISTORY']}, 'unique_id': 'test.edw.unique_STG_POLICY_STATUS_REASON_HISTORY_PLCY_PRD_ID_PSH_HIST_EFF_DTM_PSH_HIST_END_DTM_PSR_HIST_EFF_DTM_PSR_HIST_END_DTM.4c0f5f077c', 'original_file_path': 'models\\STAGING\\STG_POLICY_STATUS_REASON_HISTORY.yml', 'name': 'unique_STG_POLICY_STATUS_REASON_HISTORY_PLCY_PRD_ID_PSH_HIST_EFF_DTM_PSH_HIST_END_DTM_PSR_HIST_EFF_DTM_PSR_HIST_END_DTM'}, 
{'resource_type': 'test', 'depends_on': {'macros': ['macro.dbt.test_unique'], 'nodes': ['model.edw.STG_POLICY_SUMMARY_DETAIL']}, 'unique_id': 'test.edw.unique_STG_POLICY_SUMMARY_DETAIL_PLCY_SUM_DTL_ID.3ea8e7f135', 'original_file_path': 'models\\STAGING\\STG_POLICY_SUMMARY_DETAIL.yml', 'name': 'unique_STG_POLICY_SUMMARY_DETAIL_PLCY_SUM_DTL_ID'}, 
{'resource_type': 'test', 'depends_on': {'macros': ['macro.dbt.test_unique'], 'nodes': ['model.edw.STG_PREMIUM_PERIOD']}, 'unique_id': 'test.edw.unique_STG_PREMIUM_PERIOD_PREM_PRD_ID.b0ab38dcfe', 'original_file_path': 'models\\STAGING\\STG_PREMIUM_PERIOD.yml', 'name': 'unique_STG_PREMIUM_PERIOD_PREM_PRD_ID'}, 
{'resource_type': 'test', 'depends_on': {'macros': ['macro.dbt.test_unique'], 'nodes': ['model.edw.STG_QUOTE']}, 'unique_id': 'test.edw.unique_STG_QUOTE_AGRE_ID.92a712b4e4', 'original_file_path': 'models\\STAGING\\STG_QUOTE.yml', 'name': 'unique_STG_QUOTE_AGRE_ID'}, 
{'resource_type': 'test', 'depends_on': {'macros': ['macro.dbt.test_unique'], 'nodes': ['model.edw.STG_TASK']}, 'unique_id': 'test.edw.unique_STG_TASK_TASK_ID.1ab18527fa', 'original_file_path': 'models\\STAGING\\STG_TASK.yml', 'name': 'unique_STG_TASK_TASK_ID'}, 
{'resource_type': 'test', 'depends_on': {'macros': ['macro.dbt.test_unique'], 'nodes': ['model.edw.STG_TAX_IDENTIFIER']}, 'unique_id': 'test.edw.unique_STG_TAX_IDENTIFIER_TAX_ID_ID.99d4d0cc1a', 'original_file_path': 'models\\STAGING\\STG_TAX_IDENTIFIER.yml', 'name': 'unique_STG_TAX_IDENTIFIER_TAX_ID_ID'}, 
{'resource_type': 'test', 'depends_on': {'macros': ['macro.dbt.test_unique'], 'nodes': ['model.edw.STG_TCDERFC']}, 'unique_id': 'test.edw.unique_STG_TCDERFC_ICD_CODE_ICDV_CODE_EM_RFRL_BGNG_DATE.e57064db92', 'original_file_path': 'models\\STAGING\\STG_TCDERFC.yml', 'name': 'unique_STG_TCDERFC_ICD_CODE_ICDV_CODE_EM_RFRL_BGNG_DATE'}, 
{'resource_type': 'test', 'depends_on': {'macros': ['macro.dbt.test_unique'], 'nodes': ['model.edw.STG_TMPPRDT']}, 'unique_id': 'test.edw.unique_STG_TMPPRDT_PRVDR_ID_PRDT_CRT_DTTM_PRVDR_BASE_NMBR_PRVDR_SFX_NMBR.4497b45562', 'original_file_path': 'models\\STAGING\\STG_TMPPRDT.yml', 'name': 'unique_STG_TMPPRDT_PRVDR_ID_PRDT_CRT_DTTM_PRVDR_BASE_NMBR_PRVDR_SFX_NMBR'}, 
{'resource_type': 'test', 'depends_on': {'macros': ['macro.dbt.test_unique'], 'nodes': ['model.edw.STG_USERS']}, 'unique_id': 'test.edw.unique_STG_USERS_USER_ID.222af55d86', 'original_file_path': 'models\\STAGING\\STG_USERS.yml', 'name': 'unique_STG_USERS_USER_ID'}, 
{'resource_type': 'test', 'depends_on': {'macros': ['macro.dbt.test_unique'], 'nodes': ['model.edw.STG_WC_CLASS_RATE_TIER']}, 'unique_id': 'test.edw.unique_STG_WC_CLASS_RATE_TIER_WC_CLS_RT_TIER_ID.8e96d1c959', 'original_file_path': 'models\\STAGING\\STG_WC_CLASS_RATE_TIER.yml', 'name': 'unique_STG_WC_CLASS_RATE_TIER_WC_CLS_RT_TIER_ID'}, 
{'resource_type': 'test', 'depends_on': {'macros': ['macro.dbt.test_unique'], 'nodes': ['model.edw.STG_WC_CLASS']}, 'unique_id': 'test.edw.unique_STG_WC_CLASS_WC_CLS_SUFX_ID.ec86d45b67', 'original_file_path': 'models\\STAGING\\STG_WC_CLASS.yml', 'name': 'unique_STG_WC_CLASS_WC_CLS_SUFX_ID'}, 
{'resource_type': 'test', 'depends_on': {'macros': ['macro.dbt.test_unique'], 'nodes': ['model.edw.STG_WC_COVERAGE_PREMIUM']}, 'unique_id': 'test.edw.unique_STG_WC_COVERAGE_PREMIUM_WC_COV_PREM_ID.ae421bd29e', 'original_file_path': 'models\\STAGING\\STG_WC_COVERAGE_PREMIUM.yml', 'name': 'unique_STG_WC_COVERAGE_PREMIUM_WC_COV_PREM_ID'}
]

firstrow = sample[0]
headings = list( firstrow.keys() )
ws.append( headings )


for row in sample:
    row_stuff = list( row.values() )
    row_str = []
    for item in row_stuff:
        if type(item) is dict:
            item = str(item)
            row_str.append( item )
        else:
            row_str.append( item )
            
    ws.append( row_str )
    # Loop through values, if more than one in a list, convert to a single string
    # ws.append( row_str )

wb.save("teststuff.xlsx")
