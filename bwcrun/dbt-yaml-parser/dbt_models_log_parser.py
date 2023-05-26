import yaml, re, os, openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from pathlib import Path
# re will be used to wrap quotes around:         null, false, true


def parse_row( record, parse_row_dict ):
    # temp_row = {}
    temp_row = {'resource_type':'', 'name':'', 'unique_id':'', 'depends_on':'', 'check_cols':'', 'original_file_path':'' }
    save_cols = ['id','resource_type','name','original_file_path','unique_id','depends_on']
    save_list = dict( zip(save_cols, [None]*len(save_cols) ))

    # Loop through the record, setting each empty_list[key] to a save_col.
    # When moving to the next item, append the whole record to the save_list.
    for val in parse_row_dict.keys():
        get_rowvalue = record.get( val )
        if val in save_cols:
            temp_row[val] = get_rowvalue
            if val == 'depends_on':
                unique_node = []
                node = get_rowvalue.get('nodes')
                [unique_node.append(x) for x in node if x not in unique_node]
                temp_row[val] = ', '.join(unique_node)   
            # Capture check_cols for snapshots (per Jeff)
            if temp_row['resource_type'] == 'snapshot':
                get_rowvalue = record.get( 'config', 'unknown' )
                check = []
                cols = get_rowvalue.get('check_cols')
                [check.append(x) for x in cols if x not in check]
                temp_row['check_cols'] = ', '.join( check )   

            print( f"*   {val:<36}: '{get_rowvalue}'" )
        else:
            print( f"{val:<40}: '{get_rowvalue}'" )

    # Append a COPY of the temporary row to the empty_list
    print(f"## SAVING ##\n{temp_row}")
    print(f'\n-----------------')

    return temp_row


def save_output(empty_list):
    wb = Workbook()
    ws = wb.active
    ws.title = "Models"
    output = 'Output-parsed.xlsx'

    firstrow = empty_list[0]
    headings = list( firstrow.keys() )
    ws.append( headings )

    for row in empty_list:
        row_stuff = list( row.values() )
        row_str = []
        for item in row_stuff:
            if type(item) is dict:
                item = str(item)
                row_str.append( item )
            else:
                row_str.append( item )
                
        ws.append( row_str )
        # Loop through values, if more than one in a list, convert to a single string
        # ws.append( row_str )

    print( f'--SAVING TO: {output}' )
    wb.save( output ) 


def main():
    print(f"############ STARTING ")
    empty_list = []

    # Parse the dbt_models.log file created through:   dbt ls --output json > dbt_models.log
    # There are the following layouts:  models, seeds, snapshots, tests, uniqueids

    # NOTE: Fields must be corrected to wrap quotes around:  "null", "false", "true"

    # The layout names will be determined by the "resource_type" as the first entry of each dict, 
    # with the exception of unique_ids whose first entry is:  "unique_id":
    layouts = ['model','seed','snapshot','test','source']

    sample_records=[ 
        {"resource_type": "model", "depends_on": {"macros": ["macro.edw.ref"], "nodes": ["model.edw.STG_MIRA_VALIDATION_EXTRACT", "model.edw.STG_MIRA_STOP_CONDITIONS"]}, "config": {"enabled": "true", "alias": "null", "schema": "ODS_ACTUARIAL", "database": "null", "tags": ["weekly"], "meta": {}, "materialized": "table", "persist_docs": {}, "quoting": {}, "column_types": {}, "full_refresh": "null", "unique_key": "null", "on_schema_change": "ignore", "transient": "false", "post-hook": [], "pre-hook": []}, "unique_id": "model.edw.BWC_STG_MIRA_RESERVE_EXTRACT", "package_name": "edw", "original_file_path": "models\\ODS_ACTUARIAL\\BWC_STG_MIRA_RESERVE_EXTRACT.sql", "name": "BWC_STG_MIRA_RESERVE_EXTRACT", "alias": "BWC_STG_MIRA_RESERVE_EXTRACT", "tags": ["weekly"]} ,
        {"resource_type": "seed", "depends_on": {"macros": [], "nodes": []}, "config": {"enabled": "true", "alias": "null", "schema": "EDW_STAGING", "database": "null", "tags": [], "meta": {}, "materialized": "seed", "persist_docs": {}, "quoting": {}, "column_types": {}, "full_refresh": "null", "unique_key": "null", "on_schema_change": "ignore", "quote_columns": "null", "post-hook": [], "pre-hook": []}, "unique_id": "seed.edw.DIM_DUMMY", "package_name": "edw", "original_file_path": "data\\DIM_DUMMY.csv", "name": "DIM_DUMMY", "alias": "DIM_DUMMY", "tags": []},
        {"resource_type": "snapshot", "depends_on": {"macros": ["macro.edw.ref"], "nodes": ["model.edw.DSV_AMBULATORY_PAYMENT_CLASSIFICATION"]}, "config": {"enabled": "true", "alias": "null", "schema": "null", "database": "null", "tags": [], "meta": {}, "materialized": "snapshot", "persist_docs": {}, "quoting": {}, "column_types": {}, "full_refresh": "null", "unique_key": "UNIQUE_ID_KEY", "on_schema_change": "ignore", "strategy": "check", "target_schema": "EDW_STAGING_SNAPSHOT", "target_database": "null", "updated_at": "null", "check_cols": ["APC_DESC", "APC_CODE_STATUS_CODE", "APC_CODE_STATUS_DESC", "APC_RELATIVE_WEIGHT_RATE", "APC_AMOUNT", "APC_EFFECTIVE_DATE", "APC_EXPIRATION_DATE"], "transient": "false", "post-hook": [], "pre-hook": []}, "unique_id": "snapshot.edw.DIM_AMBULATORY_PAYMENT_CLASSIFICATION_SNAPSHOT_STEP1", "package_name": "edw", "original_file_path": "snapshots\\DIM_AMBULATORY_PAYMENT_CLASSIFICATION_SNAPSHOT_STEP1.sql", "name": "DIM_AMBULATORY_PAYMENT_CLASSIFICATION_SNAPSHOT_STEP1", "alias": "DIM_AMBULATORY_PAYMENT_CLASSIFICATION_SNAPSHOT_STEP1", "tags": []},
        {"resource_type": "test", "depends_on": {"macros": ["macro.edw.test_accepted_valids", "macro.dbt.get_where_subquery"], "nodes": ["model.edw.DIM_CLAIM"]}, "config": {"enabled": "true", "alias": "null", "schema": "dbt_test__audit", "database": "null", "tags": [], "meta": {}, "materialized": "test", "severity": "warn", "store_failures": "null", "where": "null", "limit": "null", "fail_calc": "count(*)", "warn_if": "!= 0", "error_if": "!= 0"}, "unique_id": "test.edw.accepted_valids_DIM_CLAIM_COMBINED_IND__Y__N.3f8febe0bd", "package_name": "edw", "original_file_path": "models\\EDW_STAGING_DIM\\DIM_CLAIM.yml", "name": "accepted_valids_DIM_CLAIM_COMBINED_IND__Y__N", "alias": "accepted_valids_DIM_CLAIM_COMBINED_IND__Y__N", "tags": []},
        {"unique_id": "source.edw.MEDICAL.ACCOUNT_STATUS_TYPE", "package_name": "edw", "original_file_path": "models\\source.yml", "name": "ACCOUNT_STATUS_TYPE", "source_name": "MEDICAL", "resource_type": "source", "tags": [], "config": {"enabled": "true"}},
        ]

    dict_parsers = { 
    "model_dict":    {"resource_type": "", "depends_on": {"macros": [], "nodes": []}, "config": {"enabled": "", "alias": "", "schema": "", "database": "", "tags": [], "meta": {}, "materialized": "", "persist_docs": {}, "quoting": {}, "column_types": {}, "full_refresh": "", "unique_key": "", "on_schema_change": "", "transient": "", "post-hook": [], "pre-hook": []}, "unique_id": "", "package_name": "", "original_file_path": "", "name": "", "alias": "", "tags": []} ,
    "seed_dict" :     {"resource_type": "", "depends_on": {"macros": [], "nodes": []}, "config": {"enabled": "", "alias": "", "schema": "", "database": "", "tags": [], "meta": {}, "materialized": "", "persist_docs": {}, "quoting": {}, "column_types": {}, "full_refresh": "", "unique_key": "", "on_schema_change": "", "quote_columns": "", "post-hook": [], "pre-hook": []}, "unique_id": "", "package_name": "", "original_file_path": "", "name": "", "alias": "", "tags": []},
    "snapshot_dict" : {"resource_type": "", "depends_on": {"macros": [], "nodes": []}, "config": {"enabled": "", "alias": "", "schema": "", "database": "", "tags": [], "meta": {}, "materialized": "", "persist_docs": {}, "quoting": {}, "column_types": {}, "full_refresh": "", "unique_key": "", "on_schema_change": "", "strategy": "", "target_schema": "", "target_database": "", "updated_at": "", "check_cols": [], "transient": "", "post-hook": [], "pre-hook": []}, "unique_id": "", "package_name": "", "original_file_path": "", "name": "", "alias": "", "tags": []},
    "test_dict" :     {"resource_type": "", "depends_on": {"macros": [], "nodes": []}, "config": {"enabled": "", "alias": "", "schema": "", "database": "", "tags": [], "meta": {}, "materialized": "", "severity": "", "store_failures": "", "where": "", "limit": "", "fail_calc": "", "warn_if": "", "error_if": ""}, "unique_id": "", "package_name": "", "original_file_path": "", "name": "", "alias": "", "tags": []},
    "source_dict" :   {"unique_id": "", "package_name": "", "original_file_path": "", "name": "", "source_name": "", "resource_type": "", "tags": [], "config": {"enabled": ""}}
    }

    cwd = os.path.dirname( os.path.realpath(__file__))
    # filename  = Path( cwd + "/dbt_models.log")
    # filename  = Path( cwd + "/snapshots.yml")
    filename  = Path( cwd+ "/models.yml")
    # filename = Path( "/Users/rmetzk/Desktop/python/yamls/dbt_models.log" )
    records = {}

    id = 0
    with open(filename) as f:
        for line in f:
            id += 1
            print(id, line)
            # Clean up strings
            newline = line.replace(' true',' "true"')
            # newline = newline.replace(' true}',' "true"}')
            newline = newline.replace(' false',' "false"')
            newline = newline.replace(' null',' "null"')
            newline = newline.replace('[null',' ["null"')
            # newline = re.sub('(true),','"\\1"',line).strip()
            # newline = re.sub('(false),','"\\1",', line)
            # newline = re.sub('(null)','"\\1"',line)
            newline = re.sub('(\[\]|\{\})','""',newline)
            newline = newline.replace('\n',' ')
            newline = newline.replace('\\','/')
            newline = newline.replace('/n',' ')
            newline = newline.replace('//','/')

            record= eval( newline.strip("'") )

            # for item, record in enumerate( sample_records ):
            temp_row ={}
            entry_type = record.get("resource_type","unknown")

            if entry_type:
                dict2use = f'{entry_type}_dict'
                print( f'## {entry_type} found:  using {entry_type}_dict')
                parse_row_dict =  dict_parsers.get( dict2use )
                new_row = parse_row(record, parse_row_dict)
                empty_list.append( new_row.copy() )

    print(f"############ COMPLETE ")
    # print( '------ ')
    # print( empty_list )
    # print( '------ ')
    save_output( empty_list )


if __name__ == '__main__':
    main()

