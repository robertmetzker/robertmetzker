# default imports
import os, openpyxl, sheet2dict
from pathlib import Path

# specific imports
import shutil, datetime, math, pandas as pd
from openpyxl import Workbook

def countlines( file_tomove ):
    with open( file_tomove, 'r') as fp:
        for count, line in enumerate(fp):
            pass
    
    return count + 1


def copyfile( from_loc, to_loc, file_tomove, created ):
    src_path = from_loc +'/'+ file_tomove
    tgt_path = to_loc +'/'+ file_tomove     
    num_lines = -1
    cdate = 0
    file_created = None
    
    # Write out the creation timestamp for each file
    os.makedirs( Path(to_loc), exist_ok = True)
    print(f'\n > Move {file_tomove:<20} from {from_loc}')

    try:
        # truncate the milliseconds from the mtime since xls does not store them at the same precision for comparison
        # file_created = datetime.datetime.fromtimestamp( os.path.getmtime( src_path ))
        # e.g. 1645649851.817844
        cdate = math.trunc(  os.path.getmtime( src_path ) )
        file_created = datetime.datetime.fromtimestamp( cdate )
        if not created == file_created:
            num_lines = countlines( src_path )
            shutil.copy2( src_path, tgt_path ) 
            print(f'\tsuccessfully copied to {to_loc}... ')
        else:
            print(f' ### NOTE: >>> {file_tomove:<20}  was skipped as already copied...')
    except FileNotFoundError:
        print(f' ### ERROR >>> {file_tomove:<20}  was not found in {from_loc}')
    except:
        print(f'\\nFAILED to copy to {to_loc}... ')

    return num_lines, file_created


def combinefile( from_loc, to_loc, file_tomove, combine2, created ):
    # TODO - Change this to use iterations/generator instead of PANDAS to combine
    # with open(filename, 'r') as readfile:
    #   if Path(tgt_path).is_file():
    #       next(readfile)
    #   else:
    #     shutil.copyfileobj(readfile, outfile)

    
    src_path = from_loc +'/'+ file_tomove
    tgt_path = to_loc +'/'+ combine2    
    num_lines = -1
    cdate = 0
    file_created = None
    
    # Write out the creation timestamp for each file
    os.makedirs( Path(to_loc), exist_ok = True)
    print(f'\n > Move {file_tomove:<20} from {from_loc}')

    try:
        # truncate the milliseconds from the mtime since xls does not store them at the same precision for comparison
        # file_created = datetime.datetime.fromtimestamp( os.path.getmtime( src_path ))
        # e.g. 1645649851.817844
        cdate = math.trunc(  os.path.getmtime( src_path ) )
        file_created = datetime.datetime.fromtimestamp( cdate )
        if not created == file_created:
            num_lines = countlines( src_path )

            if Path(tgt_path).is_file():
                tmp = pd.read_csv( src_path )
                tmp.to_csv(tgt_path, index=False, header=False, mode='a')
                print(f'\tsuccessfully appended to {to_loc}... ')
            else:
                tmp = pd.read_csv( src_path )
                tmp.to_csv(tgt_path, index=False, header=True, mode='w')
                print(f'\tsuccessfully created at {to_loc}... ')

        else:
            print(f' ### NOTE: >>> {file_tomove:<20}  was skipped as already copied...')
    except FileNotFoundError:
        print(f' ### ERROR >>> {file_tomove:<20}  was not found in {from_loc}')
    except:
        print(f'\nFAILED to copy to {to_loc}... ')

    return num_lines, file_created

    
def main():
    # TODO - Change file objects to Path objects
    ext_file = Path( 'c:/Users/A85275/Desktop/file-locations.xlsx' )
    print( f"=== Using information from: {os.path.basename( ext_file )}" )

    curwb = openpyxl.load_workbook( ext_file )
    wb = Workbook( ext_file )
    cur_sheet = curwb.active
    total_rows = cur_sheet.max_row + 1
    combine2 = None

    # TODO - Change this to read obtain the column numbers from the header row
    print( f'---- Found Workbooks: {wb.sheetnames}' )
    fn = 1  #    1- 'Filename'
    sp = 2  #    2- 'Source-path'
    tp = 3  #    3- 'Target-path' 
    ci = 4  #    4- 'Concat-into'
    cd = 5  #    5- 'Create-date'
    nl = 6  #    6- 'Num-Lines'
    md = 7  #    7- 'Move-date'

    for row_number in range( 2, total_rows ):
        file_tomove = cur_sheet.cell(row=row_number, column=fn ).value
        from_loc = cur_sheet.cell(row=row_number, column=sp ).value
        to_loc   = cur_sheet.cell(row=row_number, column=tp ).value
        created  = cur_sheet.cell(row=row_number, column=cd ).value
        combine2  = cur_sheet.cell(row=row_number, column=ci ).value

        if not combine2:
            num_lines, file_created = copyfile( from_loc, to_loc, file_tomove, created )
        else:
            num_lines, file_created = combinefile( from_loc, to_loc, file_tomove, combine2, created )
            
        
        try:
            cur_sheet.cell(row = row_number, column= cd).value = file_created
            cur_sheet.cell(row = row_number, column= nl).value = num_lines
            if num_lines > 0:
                cur_sheet.cell(row = row_number, column= md ).value = datetime.datetime.now()
            curwb.save( ext_file )
        except PermissionError:
            print(f"Permission denied.  {ext_file} is currently open.")            
        except Exception as e:
            print( f'#### ERROR:  {e} ' )
            
            
        
if __name__ == '__main__':
    main()