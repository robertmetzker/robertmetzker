# Python file to create a mockup/template for file management.
import os
from openpyxl import Workbook
from openpyxl.styles import Font
from pathlib import Path


def main():
    # Defaults
    desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
    cwd = os.getcwd()
    
    ext_file = 'file-locations.xlsx'

    print( f'=== Using information from: {os.path.basename( ext_file )}' )
    wb = Workbook()
    ws = wb.active
    print( f'==== Configuring Headers in {ws}' )
    
    ws['A1'] = 'Filename'
    ws['B1'] = 'Source-path'
    ws['C1'] = 'Target-path'
    ws['D1'] = 'Concat-into'
    ws['E1'] = 'Create-date'
    ws['F1'] = 'Num-Lines'
    ws['G1'] = 'Move-date'

    # Bold Headers
    ws['A1'].font =  Font(bold=True)
    ws['B1'].font =  Font(bold=True)
    ws['C1'].font =  Font(bold=True)
    ws['D1'].font =  Font(bold=True)
    ws['E1'].font =  Font(bold=True)
    ws['F1'].font =  Font(bold=True)
    ws['G1'].font =  Font(bold=True)

    # Set width on Path columns
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20

    # Dummy up some data
    print( f'==== Configuring Data in {ws}' )    
    ws['A2'] = 'sample-1.csv'
    ws['B2'] = f'{cwd}'
    ws['C2'] = f'{cwd}/subdir'

    ws['A3'] = 'sample-2.txt'
    ws['B3'] = f'{cwd}'
    ws['C3'] = f'{cwd}/subdir'
    
    ws['A4'] = 'sample-3'
    ws['B4'] = f'{cwd}'
    ws['C4'] = f'{cwd}/subdir'

    ws['A5'] = 'cat-1.txt'
    ws['B5'] = f'{cwd}'
    ws['C5'] = f'{cwd}/subdir2'
    ws['D5'] = 'combined.txt'

    ws['A6'] = 'cat-2.txt'
    ws['B6'] = f'{cwd}'
    ws['C6'] = f'{cwd}/subdir2'
    ws['D6'] = 'combined.txt'

    ws['A7'] = 'cat-3.txt'
    ws['B7'] = f'{cwd}'
    ws['C7'] = f'{cwd}/subdir2'
    ws['D7'] = 'combined.txt'

    ws['A8'] = '*'
    ws['B8'] = f'{cwd}/subdir'
    ws['C8'] = f'{cwd}/subdir3'

    ws['A9'] = '*'
    ws['B9'] = f'{cwd}/subdir2'
    ws['C9'] = f'{cwd}/subdir3'

    ws['A10'] = '*.txt'
    ws['B10'] = f'{cwd}/subdir3'
    ws['C10'] = f'{cwd}/subdir4'
    ws['D10'] = 'samples-combined.txt'

    wb.save( f'{desktop}/{ext_file}' )
    print( '--- Workbook complete' )

if __name__ == '__main__':
    main()